from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from database import db
from auth_utils import get_current_user, hash_password
from datetime import datetime, timezone
import os
import httpx

def get_financial_year() -> int:
    d = datetime.now(timezone.utc)
    return d.year if d.month >= 4 else d.year - 1
from bson import ObjectId
import csv
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

router = APIRouter()

ROLES = ["hr_admin", "management", "managers", "employee", "field_agent"]
STATUSES = ["active", "probation", "notice_period", "exited"]

DEPARTMENTS = [
    "Accounts",
    "Administration",
    "Compliance",
    "Human Resources",
    "IT",
    "Management",
    "Operations",
    "Risk and Credit",
]

# Field Team
FIELD_DESIGNATIONS = [
    "Divisional Manager",
    "Area Manager",
    "Senior Branch Manager",
    "Branch Manager",
    "Senior Field Officer",
    "Field Officer",
]

# Management
MANAGEMENT_DESIGNATIONS = [
    "Director",
    "Chief Executive Officer",
    "Chief Operating Officer",
]

# Risk Team (reports to management)
RISK_DESIGNATIONS = [
    "Audit Manager",
    "Credit Officer",
]

# Head Office
HO_DESIGNATIONS = [
    "Company Secretary",
    "HR Manager",
    "Accounts Manager",
    "Senior Manager",
    "Manager",
    "Assistant Manager",
    "Senior Executive",
    "Executive",
    "Assistant",
]

DESIGNATIONS = MANAGEMENT_DESIGNATIONS + HO_DESIGNATIONS + FIELD_DESIGNATIONS + RISK_DESIGNATIONS

DESIGNATION_GROUPS = {
    "Management": MANAGEMENT_DESIGNATIONS,
    "Head Office": HO_DESIGNATIONS,
    "Field Team": FIELD_DESIGNATIONS,
    "Risk Team": RISK_DESIGNATIONS,
}


def emp_to_dict(emp):
    emp["id"] = str(emp.pop("_id"))
    return emp


def _strip_salary_unless_authorised(emp_dict: dict, current_user: dict) -> dict:
    """Salary / CTC visibility rules:
       - hr_admin and management: see everything (full access).
       - the employee themselves: see their own salary (needed for self-portal & payslips).
       - everyone else (managers, peers, field_agent): salary is stripped.
    Applied to /employees list & detail responses. Payslip / payroll endpoints have their own
    independent ACL.
    """
    role = current_user.get("role")
    if role in ("hr_admin", "management"):
        return emp_dict
    if current_user.get("employee_id") and emp_dict.get("employee_id") == current_user.get("employee_id"):
        return emp_dict
    emp_dict.pop("salary", None)
    emp_dict.pop("ctc_monthly", None)
    emp_dict.pop("ctc_annual", None)
    return emp_dict


async def get_next_employee_id():
    last = await db.employees.find_one({}, sort=[("employee_id", -1)])
    if not last:
        return "RMF0001"
    last_id = last.get("employee_id", "RMF0000")
    try:
        num = int(last_id.replace("RMF", "")) + 1
        return f"RMF{num:04d}"
    except Exception:
        return "RMF0001"


class EmployeeCreate(BaseModel):
    first_name: str
    last_name: str
    email: str
    mobile: str
    department: str
    designation: str
    role: str
    reporting_to: Optional[str] = None
    joining_date: str
    basic: float
    hra: float
    special_allowance: float = 0
    canteen_allowance: float = 0
    conveyance_allowance: float = 0
    bank_name: Optional[str] = None
    account_number: Optional[str] = None
    ifsc_code: Optional[str] = None
    address_current: Optional[str] = None
    address_permanent: Optional[str] = None
    aadhaar_number: Optional[str] = None
    pan_number: Optional[str] = None
    emergency_contact_name: Optional[str] = None
    emergency_contact_mobile: Optional[str] = None
    date_of_birth: Optional[str] = None
    gender: Optional[str] = None
    blood_group: Optional[str] = None
    uan_number: Optional[str] = None
    esi_number: Optional[str] = None
    create_user_account: bool = True
    password: Optional[str] = None
    branch: Optional[str] = None
    shift_id: Optional[str] = None


class EmployeeUpdate(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: Optional[str] = None
    mobile: Optional[str] = None
    department: Optional[str] = None
    designation: Optional[str] = None
    role: Optional[str] = None
    reporting_to: Optional[str] = None
    joining_date: Optional[str] = None
    date_of_birth: Optional[str] = None
    gender: Optional[str] = None
    father_or_husband_name: Optional[str] = None
    aadhaar_number: Optional[str] = None
    pan_number: Optional[str] = None
    blood_group: Optional[str] = None
    emergency_contact_name: Optional[str] = None
    emergency_contact_mobile: Optional[str] = None
    basic: Optional[float] = None
    hra: Optional[float] = None
    special_allowance: Optional[float] = None
    canteen_allowance: Optional[float] = None
    conveyance_allowance: Optional[float] = None
    ctc_monthly: Optional[float] = None
    bank_name: Optional[str] = None
    account_number: Optional[str] = None
    ifsc_code: Optional[str] = None
    address_current: Optional[str] = None
    address_permanent: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    joining_location: Optional[str] = None
    branch: Optional[str] = None
    shift_id: Optional[str] = None
    multi_session_attendance: Optional[bool] = None
    status: Optional[str] = None
    uan_number: Optional[str] = None
    esi_number: Optional[str] = None
    epf_employee: Optional[float] = None


@router.get("/next-id")
async def next_employee_id(current_user: dict = Depends(get_current_user)):
    return {"next_id": await get_next_employee_id()}


@router.get("/designations")
async def get_designations(current_user: dict = Depends(get_current_user)):
    return {
        "departments": DEPARTMENTS,
        "groups": DESIGNATION_GROUPS,
        "all": DESIGNATIONS,
    }


@router.get("")
async def list_employees(
    status: str = "all",
    department: str = None,
    role: str = None,
    search: str = None,
    current_user: dict = Depends(get_current_user),
):
    query = {}
    if status and status != "all":
        query["status"] = status
    if department:
        query["department"] = department
    if role:
        query["role"] = role

    # Scope by role
    user_role  = current_user.get("role")
    my_emp_id  = current_user.get("employee_id")
    if user_role in ["employee", "field_agent"]:
        # Can only see themselves
        query["employee_id"] = my_emp_id
    elif user_role == "managers":
        # Can see themselves + their FULL reporting sub-tree (direct + indirect reports)
        from services.hierarchy import get_descendant_employee_ids
        descendants = await get_descendant_employee_ids(my_emp_id) if my_emp_id else set()
        allowed_ids = list(descendants) + ([my_emp_id] if my_emp_id else [])
        existing_filter = query.pop("$or", None)
        scope_clause = {"employee_id": {"$in": allowed_ids}}
        if existing_filter:
            query["$and"] = [scope_clause, {"$or": existing_filter}]
        else:
            query.update(scope_clause)
    # hr_admin, management: no restriction

    if search and user_role in ["hr_admin", "management", "managers"]:
        search_clause = {"$or": [
            {"first_name": {"$regex": search, "$options": "i"}},
            {"last_name": {"$regex": search, "$options": "i"}},
            {"employee_id": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
        ]}
        if "$or" in query:
            query = {"$and": [query, search_clause]}
        else:
            query.update(search_clause)
    elif search:
        query["$or"] = [
            {"first_name": {"$regex": search, "$options": "i"}},
            {"last_name": {"$regex": search, "$options": "i"}},
            {"employee_id": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
        ]

    emps = await db.employees.find(query).sort("employee_id", 1).to_list(1000)
    return [_strip_salary_unless_authorised(emp_to_dict(e), current_user) for e in emps]


@router.post("")
async def create_employee(data: EmployeeCreate, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")
    existing = await db.employees.find_one({"email": data.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Employee with this email already exists")
    employee_id = await get_next_employee_id()
    gross = data.basic + data.hra + data.special_allowance + data.canteen_allowance + data.conveyance_allowance
    emp_doc = {
        "employee_id": employee_id,
        "first_name": data.first_name,
        "last_name": data.last_name,
        "email": data.email.lower(),
        "mobile": data.mobile,
        "department": data.department,
        "designation": data.designation,
        "role": data.role,
        "reporting_to": data.reporting_to,
        "joining_date": data.joining_date,
        "status": "probation",
        "salary": {
            "basic": data.basic,
            "hra": data.hra,
            "special_allowance": data.special_allowance,
            "canteen_allowance": data.canteen_allowance,
            "conveyance_allowance": data.conveyance_allowance,
            "gross": gross,
            "epf_employee": data.epf_employee,
        },
        "bank_details": {
            "bank_name": data.bank_name,
            "account_number": data.account_number,
            "ifsc_code": data.ifsc_code,
        },
        "address": {"current": data.address_current, "permanent": data.address_permanent},
        "aadhaar_number": data.aadhaar_number,
        "pan_number": data.pan_number,
        "emergency_contact": {"name": data.emergency_contact_name, "mobile": data.emergency_contact_mobile},
        "date_of_birth": data.date_of_birth,
        "gender": data.gender,
        "blood_group": data.blood_group,
        "uan_number": data.uan_number,
        "esi_number": data.esi_number,
        "branch": data.branch,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user.get("employee_id"),
    }
    result = await db.employees.insert_one(emp_doc)
    # Initialize leave balance
    await db.leave_balances.insert_one({
        "employee_id": employee_id,
        "year": get_financial_year(),
        "CL":       {"total": 7,  "used": 0, "remaining": 7},
        "SL":       {"total": 15, "used": 0, "remaining": 15},
        "EL":       {"total": 0,  "used": 0, "remaining": 0},
        "Marriage": {"total": 5,  "used": 0, "remaining": 5},
    })
    if data.create_user_account:
        password = data.password or "Welcome@123"
        user_doc = {
            "username": employee_id,
            "email": data.email.lower(),
            "password_hash": hash_password(password),
            "name": f"{data.first_name} {data.last_name}",
            "role": data.role,
            "employee_id": employee_id,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.users.insert_one(user_doc)
    emp_doc["id"] = str(result.inserted_id)
    emp_doc.pop("_id", None)
    return emp_doc


@router.get("/{employee_id}")
async def get_employee(employee_id: str, current_user: dict = Depends(get_current_user)):
    emp = await db.employees.find_one({"employee_id": employee_id})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    return _strip_salary_unless_authorised(emp_to_dict(emp), current_user)


@router.put("/{employee_id}")
async def update_employee(employee_id: str, data: EmployeeUpdate, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")
    emp = await db.employees.find_one({"employee_id": employee_id})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}

    # shift_id == "" means "clear override" (use role default).
    if update_data.get("shift_id") == "":
        update_data.pop("shift_id")
        update_data["__unset_shift_id"] = True

    # Validate IFSC
    if update_data.get("ifsc_code"):
        import re as _re_ifsc
        ifsc = update_data["ifsc_code"].upper().strip()
        if not _re_ifsc.match(r"^[A-Z]{4}0[A-Z0-9]{6}$", ifsc):
            raise HTTPException(status_code=400, detail="Invalid IFSC code. Format: 4 letters + 0 + 6 alphanumeric.")
        update_data["ifsc_code"] = ifsc

    # Validate reporting_to
    if update_data.get("reporting_to"):
        rep = update_data["reporting_to"].strip().upper()
        if rep == employee_id:
            raise HTTPException(status_code=400, detail="An employee cannot report to themselves.")
        rep_emp = await db.employees.find_one({"employee_id": rep})
        if not rep_emp:
            raise HTTPException(status_code=400, detail=f"Reporting To: no employee with ID {rep} found.")
        update_data["reporting_to"] = rep

    # Validate email uniqueness if changed
    if update_data.get("email"):
        new_email = update_data["email"].lower().strip()
        if new_email != emp.get("email"):
            if await db.employees.find_one({"email": new_email, "employee_id": {"$ne": employee_id}}):
                raise HTTPException(status_code=400, detail=f"Email {new_email} is already in use.")
            update_data["email"] = new_email
            # Update the linked user account too
            await db.users.update_one({"employee_id": employee_id}, {"$set": {"email": new_email}})

    # Salary recalculation (or auto-distribute from CTC)
    salary_keys = ["basic", "hra", "special_allowance", "canteen_allowance", "conveyance_allowance", "epf_employee"]
    salary_changed = any(k in update_data for k in salary_keys) or "ctc_monthly" in update_data
    if salary_changed:
        salary = emp.get("salary", {}) or {}
        for k in salary_keys:
            if k in update_data:
                salary[k] = update_data.pop(k)
        if "ctc_monthly" in update_data:
            ctc = update_data.pop("ctc_monthly") or 0
            salary["ctc_monthly"] = ctc
            salary["ctc_annual"] = round(ctc * 12, 2)
            update_data["ctc_monthly"] = ctc
            update_data["ctc_annual"] = round(ctc * 12, 2)
        gross_keys = ["basic", "hra", "special_allowance", "canteen_allowance", "conveyance_allowance"]
        salary["gross"] = sum(salary.get(k, 0) or 0 for k in gross_keys)
        if not salary.get("ctc_monthly"):
            salary["ctc_monthly"] = salary["gross"]
            salary["ctc_annual"] = round(salary["gross"] * 12, 2)
        update_data["salary"] = salary

    # Bank consolidation
    bank_keys = ["bank_name", "account_number", "ifsc_code"]
    if any(k in update_data for k in bank_keys):
        bank = emp.get("bank_details", {}) or {}
        old_account = bank.get("account_number", "")
        old_ifsc    = bank.get("ifsc_code", "")
        for k in bank_keys:
            if k in update_data:
                bank[k] = update_data.pop(k)
        # If account number or IFSC changed, clear previous verification
        if bank.get("account_number", "") != old_account or bank.get("ifsc_code", "") != old_ifsc:
            bank.pop("verified", None)
            bank.pop("verified_name", None)
            bank.pop("verified_at", None)
            bank.pop("name_match_score", None)
            bank.pop("verification_raw", None)
        update_data["bank_details"] = bank

    # Address consolidation
    addr_keys = ["address_current", "address_permanent"]
    if any(k in update_data for k in addr_keys):
        addr = emp.get("address", {}) or {}
        for k in addr_keys:
            if k in update_data:
                addr[k.replace("address_", "")] = update_data.pop(k)
        update_data["address"] = addr

    # Emergency contact consolidation
    if "emergency_contact_name" in update_data or "emergency_contact_mobile" in update_data:
        ec = emp.get("emergency_contact", {}) or {}
        if "emergency_contact_name" in update_data:
            ec["name"] = update_data.pop("emergency_contact_name")
        if "emergency_contact_mobile" in update_data:
            ec["mobile"] = update_data.pop("emergency_contact_mobile")
        update_data["emergency_contact"] = ec

    # If role changes, sync user account too
    if "role" in update_data:
        await db.users.update_one({"employee_id": employee_id}, {"$set": {"role": update_data["role"]}})

    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    unset_shift = update_data.pop("__unset_shift_id", False)
    mongo_update = {"$set": update_data}
    if unset_shift:
        mongo_update["$unset"] = {"shift_id": ""}
    await db.employees.update_one({"employee_id": employee_id}, mongo_update)
    emp = await db.employees.find_one({"employee_id": employee_id})
    return emp_to_dict(emp)


EMPLOYEE_TEMPLATE_COLUMNS = [
    # Personal
    ("employee_id", "Employee ID (leave blank to auto-assign)"),
    ("first_name", "First Name *"),
    ("last_name", "Last Name *"),
    ("email", "Email *"),
    ("mobile", "Mobile *"),
    # Job
    ("department", "Department *"),
    ("designation", "Designation *"),
    ("role", "Role *"),
    ("reporting_to", "Reporting To (Employee ID)"),
    ("joining_date", "Joining Date (YYYY-MM-DD) *"),
    ("status", "Status *"),
    # Salary — all manual
    ("ctc_annual", "CTC Annual (₹)"),
    ("basic", "Basic (₹/mo)"),
    ("hra", "HRA (₹/mo)"),
    ("special_allowance", "Special Allowance (₹/mo)"),
    ("canteen_allowance", "Canteen Allowance (₹/mo)"),
    ("conveyance_allowance", "Conveyance Allowance (₹/mo)"),
    ("epf_employee", "EPF Employee (₹/mo)"),
    # Statutory
    ("uan_number", "UAN Number"),
    ("esi_number", "ESIC / ESI Number"),
    ("pan", "PAN"),
    ("aadhaar", "Aadhaar"),
    # Bank
    ("bank_name", "Bank Name"),
    ("account_number", "Account Number"),
    ("ifsc_code", "IFSC Code"),
    # Login
    ("password", "Initial Password (default: Welcome@123)"),
]


@router.get("/bulk-upload/template")
async def download_template(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["hr_admin"]:
        raise HTTPException(status_code=403, detail="Access denied")

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    # Header styling
    header_fill = PatternFill("solid", fgColor="1E2A47")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(border_style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Write headers (row 1)
    for col_idx, (_, label) in enumerate(EMPLOYEE_TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(label) // 2 + 10)
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "A2"

    # Sample row (row 2)
    sample_values = {
        "employee_id": "",
        "first_name": "John",
        "last_name": "Doe",
        "email": "john.doe@radhyamfi.com",
        "mobile": "9876543210",
        "department": "Operations",
        "designation": "Field Officer",
        "role": "field_agent",
        "reporting_to": "",
        "joining_date": "2026-03-01",
        "status": "probation",
        "ctc_annual": 300000,
        "basic": 12000,
        "hra": 4800,
        "special_allowance": 2400,
        "canteen_allowance": 1500,
        "conveyance_allowance": 1500,
        "epf_employee": 1440,
        "uan_number": "",
        "esi_number": "",
        "pan": "ABCDE1234F",
        "aadhaar": "123412341234",
        "bank_name": "SBI",
        "account_number": "1234567890",
        "ifsc_code": "SBIN0001234",
        "password": "Welcome@123",
    }
    for col_idx, (key, _) in enumerate(EMPLOYEE_TEMPLATE_COLUMNS, start=1):
        ws.cell(row=2, column=col_idx, value=sample_values.get(key, "")).alignment = Alignment(horizontal="left")

    # Hidden lookup sheet — avoids the 255-char Excel inline formula limit
    ws_lkp = wb.create_sheet("_Lookups")
    ws_lkp.sheet_state = "hidden"

    LOOKUP_LISTS = {
        "department": DEPARTMENTS,
        "designation": DESIGNATIONS,
        "role": ROLES,
        "status": STATUSES,
    }
    lkp_col_map = {}  # key -> column letter in _Lookups
    for lkp_col_idx, (lkp_key, lkp_vals) in enumerate(LOOKUP_LISTS.items(), start=1):
        lkp_letter = get_column_letter(lkp_col_idx)
        lkp_col_map[lkp_key] = lkp_letter
        for row_i, val in enumerate(lkp_vals, start=1):
            ws_lkp.cell(row=row_i, column=lkp_col_idx, value=val)

    # Data validation dropdowns referencing the hidden sheet
    def _add_dv(col_key: str, values: list):
        col_idx = next(i for i, (k, _) in enumerate(EMPLOYEE_TEMPLATE_COLUMNS, start=1) if k == col_key)
        col_letter = get_column_letter(col_idx)
        lkp_letter = lkp_col_map[col_key]
        n = len(values)
        formula = f"_Lookups!${lkp_letter}$1:${lkp_letter}${n}"
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
        dv.error = f"Pick one of the allowed values"
        dv.errorTitle = "Invalid value"
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}1001")

    _add_dv("department", DEPARTMENTS)
    _add_dv("designation", DESIGNATIONS)
    _add_dv("role", ROLES)
    _add_dv("status", STATUSES)

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    ws2["A1"] = "Radhya Micro Finance — Employee Bulk Upload Template"
    ws2["A1"].font = Font(bold=True, size=14, color="E85B1E")
    instructions = [
        "",
        "How to use:",
        "1. Delete the sample row on the 'Employees' sheet if you do not want it imported.",
        "2. Fill one row per employee. Fields marked * are required.",
        "3. Department, Designation, Role and Status use dropdowns — click the cell to pick.",
        "4. Leave 'Employee ID' blank to auto-assign the next RMF number.",
        "5. Salary values are per month unless otherwise stated.",
        "6. 'EPF Employee' is the monthly employee-side PF contribution (usually 12% of Basic, capped).",
        "7. ESIC (both sides) and Gratuity are auto-computed by the system — do not fill them here.",
        "8. UAN = 12-digit Universal Account Number. ESI = 17-digit ESIC number. Leave blank if not yet issued.",
        "9. Save as .xlsx and upload on the Employees page.",
        "",
        "Role options: " + ", ".join(ROLES),
        "Status options: " + ", ".join(STATUSES),
        "Department options: " + ", ".join(DEPARTMENTS),
        "Designation options: " + ", ".join(DESIGNATIONS),
    ]
    for i, line in enumerate(instructions, start=2):
        ws2.cell(row=i, column=1, value=line)
    ws2.column_dimensions["A"].width = 120

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="employee_bulk_upload_template.xlsx"'},
    )


def _row_to_dict_from_excel(headers_labels: list, row_values: tuple) -> dict:
    """Map Excel-sheet row to canonical field keys. headers_labels = list of col keys."""
    return {headers_labels[i]: row_values[i] for i in range(min(len(headers_labels), len(row_values)))}


@router.post("/bulk-upload")
async def bulk_upload(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["hr_admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    content = await file.read()
    filename = (file.filename or "").lower()

    # Parse either .xlsx (preferred) or legacy .csv
    rows: list[dict] = []
    if filename.endswith(".xlsx"):
        wb = load_workbook(filename=io.BytesIO(content), data_only=True)
        ws = wb["Employees"] if "Employees" in wb.sheetnames else wb.active
        # We assume the template order. Map by column key.
        col_keys = [k for k, _ in EMPLOYEE_TEMPLATE_COLUMNS]
        # Skip header row (row 1); if the first real row equals the sample, we still attempt to import it
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            d = _row_to_dict_from_excel(col_keys, row)
            rows.append(d)
    else:
        reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
        for row in reader:
            rows.append(row)

    created, skipped, errors = 0, 0, []
    for row in rows:
        # Normalise all values to strings/floats as expected
        email = str(row.get("email") or "").lower().strip()
        if not email:
            continue
        try:
            existing = await db.employees.find_one({"email": email})
            if existing:
                skipped += 1
                continue
            employee_id = (str(row.get("employee_id") or "").strip()
                           or await get_next_employee_id())

            def _f(k: str) -> float:
                v = row.get(k)
                if v in (None, ""):
                    return 0.0
                try:
                    return float(v)
                except (TypeError, ValueError):
                    return 0.0

            basic = _f("basic")
            hra = _f("hra")
            special = _f("special_allowance")
            canteen = _f("canteen_allowance")
            conveyance = _f("conveyance_allowance")
            epf_emp = _f("epf_employee") if row.get("epf_employee") not in (None, "") else None
            ctc_annual = _f("ctc_annual") if row.get("ctc_annual") not in (None, "") else None
            gross = basic + hra + special + canteen + conveyance

            emp_doc = {
                "employee_id": employee_id,
                "first_name": str(row.get("first_name") or "").strip(),
                "last_name": str(row.get("last_name") or "").strip(),
                "email": email,
                "mobile": str(row.get("mobile") or "").strip(),
                "department": str(row.get("department") or "").strip(),
                "designation": str(row.get("designation") or "").strip(),
                "role": str(row.get("role") or "employee").strip(),
                "reporting_to": str(row.get("reporting_to") or "").strip() or None,
                "joining_date": str(row.get("joining_date") or "").strip().split(" ")[0].split("T")[0],
                "status": str(row.get("status") or "active").strip(),
                "pan": str(row.get("pan") or "").strip() or None,
                "aadhaar": str(row.get("aadhaar") or "").strip() or None,
                "uan_number": str(row.get("uan_number") or "").strip() or None,
                "esi_number": str(row.get("esi_number") or "").strip() or None,
                "salary": {
                    "ctc_annual": ctc_annual,
                    "basic": basic, "hra": hra,
                    "special_allowance": special,
                    "canteen_allowance": canteen,
                    "conveyance_allowance": conveyance,
                    "epf_employee": epf_emp,
                    "gross": gross,
                },
                "bank_details": {
                    "bank_name": str(row.get("bank_name") or "").strip(),
                    "account_number": str(row.get("account_number") or "").strip(),
                    "ifsc_code": str(row.get("ifsc_code") or "").strip(),
                },
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.employees.insert_one(emp_doc)
            await db.leave_balances.insert_one({
                "employee_id": employee_id,
                "year": get_financial_year(),
                "CL":       {"total": 7,  "used": 0, "remaining": 7},
                "SL":       {"total": 15, "used": 0, "remaining": 15},
                "EL":       {"total": 0,  "used": 0, "remaining": 0},
                "Marriage": {"total": 5,  "used": 0, "remaining": 5},
            })
            password = str(row.get("password") or "Welcome@123").strip() or "Welcome@123"
            await db.users.update_one(
                {"username": employee_id},
                {"$setOnInsert": {
                    "username": employee_id,
                    "email": email,
                    "password_hash": hash_password(password),
                    "name": f"{emp_doc['first_name']} {emp_doc['last_name']}".strip(),
                    "role": emp_doc["role"],
                    "employee_id": employee_id,
                    "is_active": True,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                }},
                upsert=True,
            )
            created += 1
        except Exception as e:
            errors.append(f"Row {email}: {str(e)}")
    return {"created": created, "skipped": skipped, "errors": errors}


# ─────────────────────────────────────────────
#  Perfios bank account verification
# ─────────────────────────────────────────────
PERFIOS_URL = "https://hub.perfios.com/api/kyc/v3/bankacc-verification"

@router.post("/{employee_id}/verify-bank")
async def verify_bank_account(employee_id: str, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")

    emp = await db.employees.find_one({"employee_id": employee_id})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    bank = emp.get("bank_details") or {}
    account_number = bank.get("account_number", "").strip()
    ifsc_code      = bank.get("ifsc_code", "").strip().upper()
    emp_name       = f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip()

    if not account_number or not ifsc_code:
        raise HTTPException(status_code=400, detail="Bank account number and IFSC code are required before verifying")

    api_key = os.environ.get("PERFIOS_API_KEY", "")
    if not api_key:
        raise HTTPException(status_code=500, detail="Perfios API key not configured")

    payload = {
        "accountNumber":       account_number,
        "accountHolderName":   emp_name,
        "ifsc":                ifsc_code,
        "consent":             "Y",
        "nameMatchType":       "Entity",
        "useCombinedSolution": "Y",
        "allowPartialMatch":   True,
        "preset":              "G",
        "suppressReorderPenalty": True,
        "clientData":          {"caseId": employee_id},
    }

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(
                PERFIOS_URL,
                json=payload,
                headers={"x-auth-key": api_key, "Content-Type": "application/json"},
            )
        raw = resp.json()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Perfios API unreachable: {e}")

    # Parse Perfios response structure:
    # raw.statusCode == 101 → success
    # raw.result.data.source[0].data.accountName → registered name
    # raw.result.data.source[0].isValid → account valid
    # raw.result.comparisionData.inputVsSource.flags.accountHolderName.score → name match
    r = raw.get("result") or {}
    data = r.get("data") or {}
    sources = data.get("source") or []
    src_data = sources[0].get("data") if sources else {}
    is_valid = sources[0].get("isValid") if sources else False
    comp = r.get("comparisionData") or {}
    validity = (comp.get("inputVsSource") or {}).get("validity", "")
    name_flag = ((comp.get("inputVsSource") or {}).get("flags") or {}).get("accountHolderName") or {}

    verified   = bool(is_valid) or validity == "VALID" or raw.get("statusCode") == 101
    reg_name   = src_data.get("accountName") or src_data.get("registeredName") or ""
    name_match = name_flag.get("score")
    if name_match is not None:
        name_match = round(float(name_match) * 100, 1)  # convert 0-1 → 0-100

    # Persist result on employee
    await db.employees.update_one(
        {"employee_id": employee_id},
        {"$set": {
            "bank_details.verified":             verified,
            "bank_details.verified_name":        reg_name,
            "bank_details.name_match_score":     name_match,
            "bank_details.verified_at":          datetime.now(timezone.utc).isoformat(),
            "bank_details.verification_raw":     raw,
        }},
    )

    return {
        "verified":        verified,
        "verified_name":   reg_name,
        "name_match_score": name_match,
        "raw":             raw,
    }
