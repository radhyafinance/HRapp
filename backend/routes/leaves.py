from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
from database import db
from auth_utils import get_current_user
from datetime import datetime, timezone, date, timedelta
from bson import ObjectId
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter()

LEAVE_TYPES = ["CL", "SL", "EL", "Maternity", "Paternity", "Marriage", "Comp-Off", "LWP"]


# ──────────────────────────────────────────────────────────────
#  Notification helpers (fire-and-forget; never raise)
# ──────────────────────────────────────────────────────────────
async def _notify_leave_applied(leave: dict, employee_doc: dict):
    """When an employee applies for leave, notify their reporting manager
    (if set) plus all hr_admin / management users."""
    try:
        from routes.notifications import create_notification as _notify
        targets = set()
        rt = (employee_doc or {}).get("reporting_to")
        if rt:
            targets.add(rt)
        hr_users = await db.users.find(
            {"role": {"$in": ["hr_admin", "management"]}},
            {"_id": 0, "employee_id": 1, "username": 1}
        ).to_list(50)
        for u in hr_users:
            target = u.get("employee_id") or u.get("username")
            if target:
                targets.add(target)
        # Don't notify the applicant themselves
        targets.discard(leave["employee_id"])
        name = f"{employee_doc.get('first_name','')} {employee_doc.get('last_name','')}".strip() or leave["employee_id"]
        msg = f"{name} applied for {leave['leave_type']} ({leave.get('days',1)}d): {leave['start_date']} → {leave['end_date']}"
        for emp_id in targets:
            await _notify(
                employee_id=emp_id,
                title="Leave Application",
                message=msg,
                type="leave",
                link="/leaves",
                meta={"leave_id": str(leave.get("_id") or ""), "applicant": leave["employee_id"]},
            )
    except Exception:
        pass


async def _notify_leave_decision(leave: dict, status: str, remarks: str = None, approval_type: str = None):
    """Notify the applicant when leave is approved/rejected."""
    try:
        from routes.notifications import create_notification as _notify
        title = "Leave Approved" if status == "approved" else "Leave Rejected"
        base = f"{leave['leave_type']} ({leave.get('days',1)}d): {leave['start_date']} → {leave['end_date']}"
        if status == "approved" and approval_type == "el":
            base += " · deducted from EL"
        elif status == "approved" and approval_type == "salary_deduction":
            base += " · salary deduction"
        if remarks:
            base += f" · Note: {remarks}"
        await _notify(
            employee_id=leave["employee_id"],
            title=title,
            message=base,
            type="leave",
            link="/leaves",
            meta={"leave_id": str(leave.get("_id") or ""), "status": status},
        )
    except Exception:
        pass



LEAVE_BALANCE_TEMPLATE = {
    "CL":       {"total": 7,  "used": 0, "remaining": 7},
    "SL":       {"total": 15, "used": 0, "remaining": 15},
    "EL":       {"total": 0,  "used": 0, "remaining": 0},
    "Marriage": {"total": 5,  "used": 0, "remaining": 5},
}


def leave_to_dict(l):
    l["id"] = str(l.pop("_id"))
    # Normalize legacy from_date/to_date schema → start_date/end_date so the UI doesn't break
    if not l.get("start_date") and l.get("from_date"):
        l["start_date"] = l["from_date"]
    if not l.get("end_date") and l.get("to_date"):
        l["end_date"] = l["to_date"]
    return l


def calc_days(
    start_str: str,
    end_str: str,
    day_type: str = "full_day",
    start_half: bool = False,
    end_half: bool = False,
) -> float:
    start = date.fromisoformat(start_str)
    end = date.fromisoformat(end_str)
    if start == end:
        # Single day: half-day types count as 0.5
        return 0.5 if day_type in ("first_half", "second_half") else 1.0
    total = float((end - start).days + 1)
    if start_half:
        total -= 0.5
    if end_half:
        total -= 0.5
    return max(0.5, total)


def get_financial_year(dt: datetime = None) -> int:
    """Return the starting year of the current financial year (April–March).
    e.g. Apr 2026–Mar 2027 → 2026; Jan–Mar 2026 → 2025.
    """
    d = dt or datetime.now(timezone.utc)
    return d.year if d.month >= 4 else d.year - 1


async def validate_leave_application(data, employee: dict):
    """Enforce all 10 company leave policy rules."""
    days = calc_days(data.start_date, data.end_date, getattr(data, "day_type", "full_day"), getattr(data, "start_half", False), getattr(data, "end_half", False))
    today = date.today()
    start = date.fromisoformat(data.start_date)
    end = date.fromisoformat(data.end_date)

    # Rule 1: Marriage leave — 5 days max, once per employment, own marriage
    if data.leave_type == "Marriage":
        if days > 5:
            raise HTTPException(400, "Marriage leave cannot exceed 5 days.")
        existing = await db.leave_applications.count_documents({
            "employee_id": data.employee_id,
            "leave_type": "Marriage",
            "status": {"$ne": "rejected"},
        })
        if existing > 0:
            raise HTTPException(400, "Marriage leave can only be availed once during employment.")

    # Rule 2: CL max 2 days at a time
    if data.leave_type == "CL" and days > 2:
        raise HTTPException(
            400,
            "Casual Leave (CL) cannot exceed 2 consecutive days at a time. "
            "Additional days will be treated as Earned Leave (EL) or unauthorised leave."
        )

    # Rule 3: SL max 3 consecutive days
    if data.leave_type == "SL" and days > 3:
        raise HTTPException(
            400,
            "Sick Leave (SL) cannot exceed 3 consecutive days. "
            "For longer illness, please apply for Earned Leave (EL) or LWP."
        )

    # Rule 5: SL and CL cannot be clubbed together
    if data.leave_type in ["SL", "CL"]:
        other_type = "CL" if data.leave_type == "SL" else "SL"
        window_start = (start - timedelta(days=1)).isoformat()
        window_end = (end + timedelta(days=1)).isoformat()
        conflict = await db.leave_applications.count_documents({
            "employee_id": data.employee_id,
            "leave_type": other_type,
            "status": {"$ne": "rejected"},
            "start_date": {"$lte": window_end},
            "end_date": {"$gte": window_start},
        })
        if conflict > 0:
            raise HTTPException(
                400,
                f"SL and CL cannot be clubbed together under any circumstances. "
                f"You have an existing {other_type} application adjacent to or overlapping these dates."
            )

    # Rule 6: EL accrual starts after 6 months of employment
    if data.leave_type == "EL":
        joining_str = employee.get("joining_date") or employee.get("salary", {}).get("joining_date")
        if joining_str:
            try:
                joining = date.fromisoformat(joining_str)
                eligible_from = joining + timedelta(days=183)  # 6 months
                if today < eligible_from:
                    remaining_days = (eligible_from - today).days
                    raise HTTPException(
                        400,
                        f"Earned Leave (EL) starts accruing after 6 months of employment. "
                        f"You will be eligible in {remaining_days} more day(s) "
                        f"(from {eligible_from.strftime('%d %b %Y')})."
                    )
            except (ValueError, TypeError):
                pass
        # Rule 10: No EL during approved maternity leave period
        maternity_leaves = await db.leave_applications.find({
            "employee_id": data.employee_id,
            "leave_type": "Maternity",
            "status": "approved",
        }).to_list(100)
        for ml in maternity_leaves:
            ml_start = date.fromisoformat(ml["start_date"])
            ml_end = date.fromisoformat(ml["end_date"])
            if start <= ml_end and end >= ml_start:
                raise HTTPException(
                    400,
                    "Earned Leave (EL) cannot be availed during the maternity leave period as per company policy."
                )

    # Rule 8: Paternity — 15 days advance notice, max 2 times
    if data.leave_type == "Paternity":
        notice = (start - today).days
        if notice < 15:
            raise HTTPException(
                400,
                f"Paternity leave requires at least 15 days advance notice. "
                f"Please apply at least 15 days before your expected start date."
            )
        count = await db.leave_applications.count_documents({
            "employee_id": data.employee_id,
            "leave_type": "Paternity",
            "status": {"$ne": "rejected"},
        })
        if count >= 2:
            raise HTTPException(400, "Paternity leave can only be availed up to 2 times during employment.")

    # Rule 9: Maternity — 30 days advance notice, max 2 times
    if data.leave_type == "Maternity":
        notice = (start - today).days
        if notice < 30:
            raise HTTPException(
                400,
                f"Maternity leave requires at least 30 days advance notice. "
                f"Please apply at least 30 days before your expected start date."
            )
        count = await db.leave_applications.count_documents({
            "employee_id": data.employee_id,
            "leave_type": "Maternity",
            "status": {"$ne": "rejected"},
        })
        if count >= 2:
            raise HTTPException(400, "Maternity leave can only be availed up to 2 times during employment.")


class LeaveApplyRequest(BaseModel):
    employee_id: str
    leave_type: str
    start_date: str
    end_date: str
    reason: str
    day_type: str = "full_day"   # "full_day" | "first_half" | "second_half" (single-day)
    start_half: bool = False      # multi-day: first day is 2nd half only → −0.5
    end_half: bool = False        # multi-day: last day is 1st half only → −0.5


class CertificateUploadRequest(BaseModel):
    data_base64: str       # base64-encoded file content
    mime_type: str         # image/jpeg, image/png, application/pdf
    file_name: Optional[str] = None


class LeaveApproveRequest(BaseModel):
    action: str                        # approve / reject
    remarks: Optional[str] = None
    approval_type: Optional[str] = None  # sl | el | salary_deduction (for SL > 2 days without cert)


class LeaveAdminEditRequest(BaseModel):
    leave_type: Optional[str] = None
    reason: Optional[str] = None
    remarks: Optional[str] = None
    approval_type: Optional[str] = None  # sl | el | salary_deduction


class LeaveAdminCancelRequest(BaseModel):
    reason: Optional[str] = None


class EncashmentRequest(BaseModel):
    employee_id: str
    days_to_encash: int
    remarks: Optional[str] = None


@router.post("")
async def apply_leave(data: LeaveApplyRequest, current_user: dict = Depends(get_current_user)):
    if data.leave_type not in LEAVE_TYPES:
        raise HTTPException(status_code=400, detail=f"Invalid leave type. Allowed: {LEAVE_TYPES}")

    # Fetch employee for joining date (Rule 6) and other checks
    emp_id = data.employee_id or current_user.get("employee_id")
    employee = await db.employees.find_one({"employee_id": emp_id}) or {}

    # Detect admin applying on behalf of another employee
    is_admin_for_other = (
        current_user.get("role") in ["hr_admin", "management"]
        and bool(data.employee_id)
        and data.employee_id != (current_user.get("employee_id") or "")
    )

    # Skip policy validation when admin applies on behalf of an employee (admin override)
    if not is_admin_for_other:
        await validate_leave_application(data, employee)

    days = calc_days(data.start_date, data.end_date, data.day_type, data.start_half, data.end_half)

    # Balance check for quota-tracked types (skip for admin applying on behalf — they know best)
    if not is_admin_for_other and data.leave_type not in ["Maternity", "Paternity", "LWP", "Comp-Off"]:
        balance = await db.leave_balances.find_one(
            {"employee_id": emp_id, "year": get_financial_year()}
        )
        if balance and data.leave_type in balance:
            remaining = balance[data.leave_type].get("remaining", 0)
            if days > remaining:
                raise HTTPException(
                    status_code=400,
                    detail=f"Insufficient {data.leave_type} balance. Remaining: {remaining} day(s)."
                )

    now_str = datetime.now(timezone.utc).isoformat()
    approval_type = data.leave_type.lower().replace("-", "_").replace(" ", "_")
    approver_id = current_user.get("employee_id") or current_user.get("username")

    doc = {
        "employee_id": emp_id,
        "leave_type": data.leave_type,
        "start_date": data.start_date,
        "end_date": data.end_date,
        "days": days,
        "day_type": data.day_type,
        "start_half": data.start_half,
        "end_half": data.end_half,
        "reason": data.reason,
        "medical_certificate": None,
        "certificate_uploaded_at": None,
        # Admin applying on behalf → auto-approved immediately
        "status": "approved" if is_admin_for_other else "pending",
        "applied_at": now_str,
        "applied_by_admin": is_admin_for_other,
        "approved_by": approver_id if is_admin_for_other else None,
        "approval_date": now_str if is_admin_for_other else None,
        "remarks": f"Applied by {approver_id} (HR Admin)" if is_admin_for_other else None,
        "approval_type": approval_type if is_admin_for_other else None,
    }
    result = await db.leave_applications.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    doc.pop("_id", None)

    # Auto-approve: deduct balance immediately
    if is_admin_for_other:
        BALANCE_TRACKED = ["CL", "SL", "EL", "Marriage"]
        if data.leave_type in BALANCE_TRACKED:
            fy = get_financial_year()
            bal_exists = await db.leave_balances.find_one(
                {"employee_id": emp_id, "year": fy}, {"_id": 1}
            )
            if not bal_exists:
                await db.leave_balances.insert_one({
                    "employee_id": emp_id,
                    "year": fy,
                    **{k: dict(v) for k, v in LEAVE_BALANCE_TEMPLATE.items()},
                })
            await db.leave_balances.update_one(
                {"employee_id": emp_id, "year": fy},
                {"$inc": {
                    f"{data.leave_type}.used": days,
                    f"{data.leave_type}.remaining": -days,
                }},
            )
    else:
        await _notify_leave_applied({**doc, "_id": result.inserted_id}, employee)
    return doc


@router.get("/calendar-overlay")
async def calendar_overlay(
    date_from: str,
    date_to: str,
    current_user: dict = Depends(get_current_user),
):
    """Return APPROVED leaves overlapping the given date window.
    Scope:
      - hr_admin / management → all employees
      - managers → direct reports + self
      - employee / field_agent → only own
    Used by the Calendar page to overlay team-coverage gaps.
    """
    role = current_user.get("role")
    me_id = current_user.get("employee_id")

    q = {
        "status": "approved",
        "$and": [
            {"$or": [
                {"start_date": {"$lte": date_to}},
                {"from_date": {"$lte": date_to}},
            ]},
            {"$or": [
                {"end_date": {"$gte": date_from}},
                {"to_date": {"$gte": date_from}},
            ]},
        ],
    }

    if role in ["hr_admin", "management"]:
        pass
    elif role == "managers":
        from services.hierarchy import get_descendant_employee_ids
        scope_ids = list(await get_descendant_employee_ids(me_id)) if me_id else []
        if me_id:
            scope_ids.append(me_id)
        q["employee_id"] = {"$in": scope_ids} if scope_ids else "__none__"
    else:
        q["employee_id"] = me_id

    leaves = await db.leave_applications.find(q).to_list(2000)
    if not leaves:
        return []

    # Enrich with employee name
    emp_ids = list({l["employee_id"] for l in leaves})
    employees = await db.employees.find(
        {"employee_id": {"$in": emp_ids}},
        {"_id": 0, "employee_id": 1, "first_name": 1, "last_name": 1, "designation": 1},
    ).to_list(2000)
    emap = {e["employee_id"]: e for e in employees}

    out = []
    for l in leaves:
        emp = emap.get(l["employee_id"], {})
        name = f"{emp.get('first_name','')} {emp.get('last_name','')}".strip() or l["employee_id"]
        # Initials for the dot
        initials = "".join([w[0] for w in name.split()[:2] if w])[:2].upper() or l["employee_id"][:2]
        out.append({
            "id": str(l["_id"]),
            "employee_id": l["employee_id"],
            "employee_name": name,
            "initials": initials,
            "designation": emp.get("designation", ""),
            "leave_type": l.get("leave_type"),
            "from_date": l.get("from_date") or l.get("start_date"),
            "to_date": l.get("to_date") or l.get("end_date"),
            "start_date": l.get("start_date") or l.get("from_date"),
            "end_date": l.get("end_date") or l.get("to_date"),
            "days": l.get("days"),
            "reason": l.get("reason"),
            "day_type": l.get("day_type", "full_day"),
            "start_half": l.get("start_half", False),
            "end_half": l.get("end_half", False),
        })
    return out


@router.get("")
async def list_leaves(
    employee_id: str = None,
    status: str = None,
    current_user: dict = Depends(get_current_user),
):
    query = {}
    role = current_user.get("role")
    me_id = current_user.get("employee_id")

    if role in ["employee", "field_agent"]:
        # Employees can only see their own leaves
        query["employee_id"] = me_id
    elif role == "managers":
        if employee_id:
            # Viewing a specific employee — validate they are in the sub-tree
            from services.hierarchy import get_descendant_employee_ids
            allowed = await get_descendant_employee_ids(me_id) if me_id else set()
            if me_id:
                allowed.add(me_id)
            if employee_id not in allowed:
                raise HTTPException(status_code=403, detail="Access denied")
            query["employee_id"] = employee_id
        else:
            # Default: scope to entire sub-tree + self
            from services.hierarchy import get_descendant_employee_ids
            scope_ids = list(await get_descendant_employee_ids(me_id)) if me_id else []
            if me_id:
                scope_ids.append(me_id)
            query["employee_id"] = {"$in": scope_ids} if scope_ids else "__none__"
    elif employee_id:
        # hr_admin / management filtering by specific employee
        query["employee_id"] = employee_id

    if status:
        query["status"] = status
    leaves = await db.leave_applications.find(query).sort("applied_at", -1).to_list(1000)
    return [leave_to_dict(l) for l in leaves]


async def _enrich_leaves_with_employee(leaves: list) -> list:
    """Attach name, designation, department, branch to each leave row."""
    if not leaves:
        return []
    emp_ids = list({l.get("employee_id") for l in leaves if l.get("employee_id")})
    employees = await db.employees.find(
        {"employee_id": {"$in": emp_ids}},
        {"_id": 0, "employee_id": 1, "first_name": 1, "last_name": 1,
         "designation": 1, "department": 1, "branch": 1},
    ).to_list(2000)
    emap = {e["employee_id"]: e for e in employees}
    out = []
    for l in leaves:
        d = leave_to_dict(l)
        emp = emap.get(d.get("employee_id"), {})
        d["employee_name"] = f"{emp.get('first_name','')} {emp.get('last_name','')}".strip()
        d["designation"] = emp.get("designation", "")
        d["department"] = emp.get("department", "")
        d["branch"] = emp.get("branch", "")
        out.append(d)
    return out


@router.get("/pending")
async def pending_leaves(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["hr_admin", "management", "managers"]:
        raise HTTPException(status_code=403, detail="Access denied")

    query = {"status": "pending"}
    # Managers see leaves of every employee in their sub-tree (direct + indirect
    # reports). HR/Management see everything.
    if current_user.get("role") == "managers":
        from services.hierarchy import get_descendant_employee_ids
        me_id = current_user.get("employee_id")
        if not me_id:
            return []
        scope = await get_descendant_employee_ids(me_id)
        if not scope:
            return []
        query["employee_id"] = {"$in": list(scope)}

    leaves = await db.leave_applications.find(query).sort("applied_at", -1).to_list(500)
    enriched = await _enrich_leaves_with_employee(leaves)

    # Attach remaining balance for the requested leave type so approvers can see it
    fy = get_financial_year()
    emp_ids = list({l["employee_id"] for l in leaves})
    balances = await db.leave_balances.find(
        {"employee_id": {"$in": emp_ids}, "year": fy}
    ).to_list(500)
    bal_map = {b["employee_id"]: b for b in balances}

    # Comp-Off balances live in a separate collection — batch-fetch for any Comp-Off leaves
    co_emp_ids = [l["employee_id"] for l in enriched if l.get("leave_type") == "Comp-Off"]
    co_remaining_map = {}
    if co_emp_ids:
        co_pipeline = [
            {"$match": {"employee_id": {"$in": co_emp_ids}, "status": "approved"}},
            {"$group": {"_id": "$employee_id", "count": {"$sum": 1}}},
        ]
        co_rows = await db.comp_off_grants.aggregate(co_pipeline).to_list(1000)
        co_remaining_map = {r["_id"]: r["count"] for r in co_rows}

    for l in enriched:
        lt = l.get("leave_type", "")
        if lt == "Comp-Off":
            l["remaining_balance"] = co_remaining_map.get(l["employee_id"], 0)
        else:
            bal = bal_map.get(l["employee_id"], {})
            if lt in bal and isinstance(bal[lt], dict):
                l["remaining_balance"] = bal[lt].get("remaining")
            else:
                l["remaining_balance"] = None
    return enriched


@router.get("/approved")
async def approved_leaves(current_user: dict = Depends(get_current_user)):
    """All approved leaves enriched with employee name/designation/branch.
    HR Admin / Management only."""
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")
    leaves = await db.leave_applications.find({"status": "approved"}).sort("start_date", -1).to_list(2000)
    return await _enrich_leaves_with_employee(leaves)


@router.get("/balances/all")
async def all_leave_balances(current_user: dict = Depends(get_current_user)):
    """Return leave balances for all employees — HR Admin and Management only."""
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")
    fy = get_financial_year()
    # Fetch all employees (name + id)
    employees = await db.employees.find(
        {"status": {"$in": ["active", "probation", "notice_period"]}},
        {"employee_id": 1, "first_name": 1, "last_name": 1, "department": 1, "designation": 1, "_id": 0}
    ).sort("employee_id", 1).to_list(1000)

    emp_ids = [e["employee_id"] for e in employees]
    emp_map = {e["employee_id"]: e for e in employees}

    balances = await db.leave_balances.find(
        {"employee_id": {"$in": emp_ids}, "year": fy}
    ).to_list(1000)
    bal_map = {b["employee_id"]: b for b in balances}

    # Live Comp-Off counts (approved, not used, not expired) — one aggregation, not N queries
    comp_off_pipeline = [
        {"$match": {"employee_id": {"$in": emp_ids}, "status": "approved"}},
        {"$group": {"_id": "$employee_id", "count": {"$sum": 1}}},
    ]
    co_rows = await db.comp_off_grants.aggregate(comp_off_pipeline).to_list(1000)
    co_map = {r["_id"]: r["count"] for r in co_rows}

    result = []
    for emp_id in emp_ids:
        emp = emp_map[emp_id]
        bal = bal_map.get(emp_id, LEAVE_BALANCE_TEMPLATE)
        # Prefer manually stored Comp-Off balance; fall back to counting live grants
        stored_co = bal.get("Comp-Off")
        if stored_co and isinstance(stored_co, dict):
            comp_off_bal = {
                "total": stored_co.get("total", 0),
                "used": stored_co.get("used", 0),
                "remaining": stored_co.get("remaining", 0),
            }
        else:
            co_remaining = co_map.get(emp_id, 0)
            comp_off_bal = {"total": co_remaining, "used": 0, "remaining": co_remaining}
        result.append({
            "employee_id": emp_id,
            "name": f"{emp.get('first_name','')} {emp.get('last_name','')}".strip(),
            "department": emp.get("department", ""),
            "designation": emp.get("designation", ""),
            "CL": bal.get("CL", {"total": 7, "used": 0, "remaining": 7}),
            "SL": bal.get("SL", {"total": 15, "used": 0, "remaining": 15}),
            "EL": bal.get("EL", {"total": 0, "used": 0, "remaining": 0}),
            "Marriage": bal.get("Marriage", {"total": 5, "used": 0, "remaining": 5}),
            "Comp-Off": comp_off_bal,
        })
    return result



@router.get("/balance/my")
async def my_leave_balance(current_user: dict = Depends(get_current_user)):
    emp_id = current_user.get("employee_id")
    if not emp_id:
        return {**LEAVE_BALANCE_TEMPLATE, "Comp-Off": {"total": 0, "used": 0, "remaining": 0}}
    balance = await db.leave_balances.find_one(
        {"employee_id": emp_id, "year": get_financial_year()}
    )
    if not balance:
        bal = dict(LEAVE_BALANCE_TEMPLATE)
    else:
        balance.pop("_id", None)
        bal = balance
    # Prefer manually stored Comp-Off balance; fall back to counting live grants
    stored_co = bal.get("Comp-Off")
    if stored_co and isinstance(stored_co, dict):
        # Manual override already set in leave_balances
        pass
    else:
        comp_off_remaining = await db.comp_off_grants.count_documents(
            {"employee_id": emp_id, "status": "approved"}
        )
        bal["Comp-Off"] = {"total": comp_off_remaining, "used": 0, "remaining": comp_off_remaining}
    return bal


@router.get("/balance/{employee_id}")
async def get_leave_balance(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Get leave balance for a specific employee.
    Only the employee themselves, hr_admin, or management may view."""
    role = current_user.get("role")
    me_id = current_user.get("employee_id")
    if role not in ["hr_admin", "management"] and me_id != employee_id:
        raise HTTPException(status_code=403, detail="Access denied")
    balance = await db.leave_balances.find_one(
        {"employee_id": employee_id, "year": get_financial_year()}
    )
    if not balance:
        bal = {"employee_id": employee_id, **LEAVE_BALANCE_TEMPLATE}
    else:
        balance.pop("_id", None)
        bal = balance
    # Prefer manually stored Comp-Off balance; fall back to counting live grants
    stored_co = bal.get("Comp-Off")
    if not (stored_co and isinstance(stored_co, dict)):
        comp_off_remaining = await db.comp_off_grants.count_documents(
            {"employee_id": employee_id, "status": "approved"}
        )
        bal["Comp-Off"] = {"total": comp_off_remaining, "used": 0, "remaining": comp_off_remaining}
    return bal


@router.post("/{leave_id}/certificate")
async def upload_certificate(
    leave_id: str,
    body: CertificateUploadRequest,
    current_user: dict = Depends(get_current_user),
):
    """Upload medical certificate (photo or PDF) for an SL application. Can be done before or after the leave."""
    import base64 as _b64
    leave = await db.leave_applications.find_one({"_id": ObjectId(leave_id)})
    if not leave:
        raise HTTPException(status_code=404, detail="Leave application not found")

    # Employees can upload their own; managers can upload for anyone
    is_manager = current_user.get("role") in ["hr_admin", "management", "managers"]
    if not is_manager and current_user.get("employee_id") != leave["employee_id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    if body.mime_type not in ["image/jpeg", "image/png", "image/jpg", "application/pdf"]:
        raise HTTPException(status_code=400, detail="Only JPEG, PNG, or PDF files are accepted.")

    # Validate base64
    try:
        raw = _b64.b64decode(body.data_base64)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid base64 payload.")

    if len(raw) > 5 * 1024 * 1024:  # 5 MB limit
        raise HTTPException(status_code=400, detail="File size must not exceed 5 MB.")

    await db.leave_applications.update_one(
        {"_id": ObjectId(leave_id)},
        {"$set": {
            "medical_certificate": body.data_base64,
            "certificate_mime_type": body.mime_type,
            "certificate_file_name": body.file_name or "certificate",
            "certificate_uploaded_at": datetime.now(timezone.utc).isoformat(),
            "certificate_uploaded_by": current_user.get("employee_id"),
        }},
    )
    return {"message": "Medical certificate uploaded successfully."}


@router.put("/{leave_id}/approve")
async def approve_leave(leave_id: str, data: LeaveApproveRequest, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["hr_admin", "management", "managers"]:
        raise HTTPException(status_code=403, detail="Access denied")
    leave = await db.leave_applications.find_one({"_id": ObjectId(leave_id)})
    if not leave:
        raise HTTPException(status_code=404, detail="Leave application not found")
    if leave["status"] != "pending":
        raise HTTPException(status_code=400, detail="Leave is already processed")

    # Approval permission rules (per Q3 — direct manager + Admin only):
    #   hr_admin                        → always allowed
    #   management                      → always allowed (for emergencies)
    #   managers (direct reporting_to) → allowed
    #   managers (skip-level)           → CAN view (in pending list) but NOT approve
    if current_user.get("role") == "managers":
        me_id = current_user.get("employee_id")
        applicant = await db.employees.find_one(
            {"employee_id": leave["employee_id"]},
            {"_id": 0, "reporting_to": 1},
        )
        if not me_id or not applicant or applicant.get("reporting_to") != me_id:
            raise HTTPException(
                status_code=403,
                detail="Only the direct reporting manager (or HR Admin) can approve this leave."
            )

    new_status = "approved" if data.action == "approve" else "rejected"

    # Determine which balance to deduct (default to leave type)
    deduct_type = leave["leave_type"]
    # Default approval_type to the actual leave type (lowercased) so messages are correct
    approval_type = data.approval_type or leave["leave_type"].lower().replace("-", "_").replace(" ", "_")

    # SL > 2 days — special handling based on certificate presence
    if new_status == "approved" and leave["leave_type"] == "SL" and leave.get("days", 1) > 2:
        has_cert = bool(leave.get("medical_certificate"))
        if has_cert:
            # Certificate present: approve as normal SL
            deduct_type = "SL"
            approval_type = "sl"
        else:
            # No certificate — admin must specify how to handle
            if data.approval_type == "el":
                # Check EL balance
                balance = await db.leave_balances.find_one(
                    {"employee_id": leave["employee_id"], "year": get_financial_year()}
                )
                el_remaining = (balance or {}).get("EL", {}).get("remaining", 0)
                if el_remaining < leave["days"]:
                    raise HTTPException(
                        400,
                        f"Insufficient EL balance to convert. Available: {el_remaining} day(s). "
                        f"Use 'salary_deduction' instead or approve partially."
                    )
                deduct_type = "EL"
                approval_type = "el"
            elif data.approval_type == "salary_deduction":
                # No balance deduction — payroll team handles salary cut
                deduct_type = None
                approval_type = "salary_deduction"
            else:
                raise HTTPException(
                    400,
                    "SL > 2 days without medical certificate requires an approval decision: "
                    "'el' (deduct from Earned Leave) or 'salary_deduction' (admin will deduct from salary)."
                )

    await db.leave_applications.update_one(
        {"_id": ObjectId(leave_id)},
        {"$set": {
            "status": new_status,
            "approved_by": current_user.get("employee_id") or current_user.get("username"),
            "approval_date": datetime.now(timezone.utc).isoformat(),
            "remarks": data.remarks,
            "approval_type": approval_type,
        }},
    )

    # Deduct balance
    BALANCE_TRACKED = ["CL", "SL", "EL", "Marriage"]
    if new_status == "approved" and deduct_type in BALANCE_TRACKED:
        fy = get_financial_year()
        # Ensure balance document exists — auto-initialize if missing
        bal_exists = await db.leave_balances.find_one(
            {"employee_id": leave["employee_id"], "year": fy}, {"_id": 1}
        )
        if not bal_exists:
            await db.leave_balances.insert_one({
                "employee_id": leave["employee_id"],
                "year": fy,
                **{k: dict(v) for k, v in LEAVE_BALANCE_TEMPLATE.items()},
            })
        await db.leave_balances.update_one(
            {"employee_id": leave["employee_id"], "year": fy},
            {"$inc": {
                f"{deduct_type}.used": leave["days"],
                f"{deduct_type}.remaining": -leave["days"],
            }},
        )

    # Sync attendance for half-day leaves
    if new_status == "approved":
        emp_id_leave = leave["employee_id"]
        day_type = leave.get("day_type", "full_day")
        start_half = leave.get("start_half", False)
        end_half = leave.get("end_half", False)
        half_day_note = "Half day leave (approved)"

        async def _mark_half_day(att_date: str):
            await db.attendance_records.update_one(
                {"employee_id": emp_id_leave, "date": att_date},
                {"$set": {"status": "half_day", "leave_half_day": True, "regularised": True, "regularised_note": half_day_note}},
                upsert=True,
            )

        # Single-day half-leave
        if leave["start_date"] == leave["end_date"] and day_type in ("first_half", "second_half"):
            await _mark_half_day(leave["start_date"])
        else:
            # Multi-day: first day starts from 2nd half
            if start_half:
                await _mark_half_day(leave["start_date"])
            # Multi-day: last day ends at 1st half
            if end_half:
                await _mark_half_day(leave["end_date"])

    msg_map = {
        "sl": "Approved as Sick Leave.",
        "el": "Approved — deducted from Earned Leave balance.",
        "salary_deduction": "Approved — salary deduction noted for payroll.",
        "cl": "Approved as Casual Leave.",
        "el": "Approved as Earned Leave.",
        "marriage": "Approved as Marriage Leave.",
        "maternity": "Approved as Maternity Leave.",
        "paternity": "Approved as Paternity Leave.",
        "comp_off": "Approved as Comp-Off.",
        "lwp": "Approved as Leave Without Pay.",
    }
    await _notify_leave_decision(leave, new_status, data.remarks, approval_type if new_status == "approved" else None)
    return {
        "message": msg_map.get(approval_type, f"Leave {new_status}"),
        "status": new_status,
        "approval_type": approval_type,
    }


@router.put("/{leave_id}/admin-edit")
async def admin_edit_approved_leave(
    leave_id: str,
    data: LeaveAdminEditRequest,
    current_user: dict = Depends(get_current_user),
):
    """hr_admin / management can edit an already-approved leave record.
    Automatically adjusts leave balances when leave_type or approval_type changes."""
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")

    leave = await db.leave_applications.find_one({"_id": ObjectId(leave_id)})
    if not leave:
        raise HTTPException(status_code=404, detail="Leave application not found")
    if leave["status"] != "approved":
        raise HTTPException(status_code=400, detail="Only approved leaves can be edited this way")

    BALANCE_TRACKED = ["CL", "SL", "EL", "Marriage"]
    fy = get_financial_year()
    days = leave.get("days", 1)

    old_type = leave["leave_type"]
    old_approval_type = leave.get("approval_type", old_type.lower().replace("-", "_").replace(" ", "_"))
    new_type = data.leave_type if data.leave_type else old_type
    new_approval_type = data.approval_type if data.approval_type else old_approval_type

    def _deduct_type(ltype, atype):
        if atype == "el":
            return "EL"
        if atype == "salary_deduction":
            return None  # no balance was deducted
        if ltype in BALANCE_TRACKED:
            return ltype
        return None

    old_deduct = _deduct_type(old_type, old_approval_type)
    new_deduct = _deduct_type(new_type, new_approval_type)

    # Adjust balances only when deduction target changed
    if old_deduct != new_deduct:
        if old_deduct in BALANCE_TRACKED:
            await db.leave_balances.update_one(
                {"employee_id": leave["employee_id"], "year": fy},
                {"$inc": {
                    f"{old_deduct}.used": -days,
                    f"{old_deduct}.remaining": days,
                }},
            )
        if new_deduct in BALANCE_TRACKED:
            await db.leave_balances.update_one(
                {"employee_id": leave["employee_id"], "year": fy},
                {"$inc": {
                    f"{new_deduct}.used": days,
                    f"{new_deduct}.remaining": -days,
                }},
            )

    update_payload = {
        "edited_by": current_user.get("employee_id") or current_user.get("username"),
        "edited_at": datetime.now(timezone.utc).isoformat(),
    }
    if data.leave_type:
        update_payload["leave_type"] = data.leave_type
    if data.reason is not None:
        update_payload["reason"] = data.reason
    if data.remarks is not None:
        update_payload["remarks"] = data.remarks
    if data.approval_type:
        update_payload["approval_type"] = data.approval_type

    await db.leave_applications.update_one(
        {"_id": ObjectId(leave_id)},
        {"$set": update_payload},
    )
    return {"message": "Leave updated successfully."}


@router.put("/{leave_id}/admin-cancel")
async def admin_cancel_approved_leave(
    leave_id: str,
    data: LeaveAdminCancelRequest,
    current_user: dict = Depends(get_current_user),
):
    """hr_admin / management can cancel an already-approved leave.
    Reverses the leave balance deduction automatically."""
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")

    leave = await db.leave_applications.find_one({"_id": ObjectId(leave_id)})
    if not leave:
        raise HTTPException(status_code=404, detail="Leave application not found")
    if leave["status"] != "approved":
        raise HTTPException(status_code=400, detail="Only approved leaves can be cancelled")

    BALANCE_TRACKED = ["CL", "SL", "EL", "Marriage"]
    fy = get_financial_year()
    days = leave.get("days", 1)

    leave_type = leave["leave_type"]
    approval_type = leave.get("approval_type", leave_type.lower().replace("-", "_").replace(" ", "_"))

    # Determine which balance was originally deducted
    if approval_type == "el":
        deduct_type = "EL"
    elif approval_type == "salary_deduction":
        deduct_type = None  # nothing was deducted from balance
    elif leave_type in BALANCE_TRACKED:
        deduct_type = leave_type
    else:
        deduct_type = None

    # Reverse the balance deduction
    if deduct_type in BALANCE_TRACKED:
        await db.leave_balances.update_one(
            {"employee_id": leave["employee_id"], "year": fy},
            {"$inc": {
                f"{deduct_type}.used": -days,
                f"{deduct_type}.remaining": days,
            }},
        )

    await db.leave_applications.update_one(
        {"_id": ObjectId(leave_id)},
        {"$set": {
            "status": "cancelled",
            "cancelled_by": current_user.get("employee_id") or current_user.get("username"),
            "cancelled_at": datetime.now(timezone.utc).isoformat(),
            "cancellation_reason": data.reason or "",
        }},
    )
    return {"message": "Leave cancelled and balance restored."}


@router.post("/admin/credit-halfyear")
async def credit_halfyear_leaves(current_user: dict = Depends(get_current_user)):
    """
    Rule 3: SL and CL credited on half-yearly basis.
    HR Admin triggers this on Jan 1 (H1) and Jul 1 (H2) each year.
    Credits 3.5 CL and 7.5 SL per half-year to all active employees.
    """
    if current_user.get("role") != "hr_admin":
        raise HTTPException(status_code=403, detail="HR Admin only")

    today = datetime.now(timezone.utc)
    fy = get_financial_year(today)
    # Financial year half-years: H1 = Apr–Sep, H2 = Oct–Mar
    half = "H1" if 4 <= today.month <= 9 else "H2"
    flag_key = f"credited_{half}_{fy}"

    employees = await db.employees.find(
        {"status": {"$in": ["active", "probation"]}},
        {"employee_id": 1, "_id": 0}
    ).to_list(1000)

    credited = 0
    skipped = 0
    for emp in employees:
        emp_id = emp["employee_id"]
        balance = await db.leave_balances.find_one(
            {"employee_id": emp_id, "year": fy}
        )
        if not balance:
            continue
        if balance.get(flag_key):
            skipped += 1
            continue
        await db.leave_balances.update_one(
            {"employee_id": emp_id, "year": fy},
            {"$inc": {
                "CL.total": 3.5,   # 3.5 CL per half (7 CL per FY)
                "CL.remaining": 3.5,
                "SL.total": 7.5,   # 7.5 SL per half (15 SL per FY)
                "SL.remaining": 7.5,
            }, "$set": {flag_key: True}},
        )
        credited += 1

    return {
        "message": f"{half} FY{fy}-{str(fy+1)[-2:]} leave credit completed.",
        "credited": credited,
        "skipped_already_credited": skipped,
    }


@router.post("/admin/credit-monthly-el")
async def credit_monthly_el(current_user: dict = Depends(get_current_user)):
    """
    Rule 6: EL accrues at 1 day/month after 6 months of employment.
    HR Admin triggers this on the 1st of every month.
    Only credits employees who have completed 6 months and haven't been credited this month.
    """
    if current_user.get("role") != "hr_admin":
        raise HTTPException(status_code=403, detail="HR Admin only")

    today = datetime.now(timezone.utc)
    fy = get_financial_year(today)
    flag_key = f"el_credited_{today.year}_{today.month:02d}"

    employees = await db.employees.find(
        {"status": {"$in": ["active", "probation"]}},
        {"employee_id": 1, "joining_date": 1, "_id": 0}
    ).to_list(1000)

    credited = 0
    skipped_not_eligible = 0
    skipped_already_credited = 0

    for emp in employees:
        emp_id = emp["employee_id"]
        joining_str = emp.get("joining_date")
        if not joining_str:
            skipped_not_eligible += 1
            continue

        try:
            joining = date.fromisoformat(joining_str)
        except (ValueError, TypeError):
            skipped_not_eligible += 1
            continue

        # Must have completed 6 months
        eligible_from = joining + timedelta(days=183)
        if today.date() < eligible_from:
            skipped_not_eligible += 1
            continue

        balance = await db.leave_balances.find_one({"employee_id": emp_id, "year": fy})
        if balance and balance.get(flag_key):
            skipped_already_credited += 1
            continue

        # Auto-create the FY balance row if missing so the EL credit isn't silently dropped.
        if not balance:
            new_balance = {
                "employee_id": emp_id,
                "year": fy,
                **{k: dict(v) for k, v in LEAVE_BALANCE_TEMPLATE.items()},
            }
            await db.leave_balances.insert_one(new_balance)

        await db.leave_balances.update_one(
            {"employee_id": emp_id, "year": fy},
            {
                "$inc": {"EL.total": 1, "EL.remaining": 1},
                "$set": {flag_key: True},
            },
        )
        credited += 1

    return {
        "message": f"Monthly EL credit for {today.strftime('%B %Y')} completed.",
        "credited": credited,
        "skipped_not_eligible": skipped_not_eligible,
        "skipped_already_credited": skipped_already_credited,
    }



async def request_el_encashment(data: EncashmentRequest, current_user: dict = Depends(get_current_user)):
    """
    Rule 7: EL encashment after 3 years of service, min 30 EL accumulated.
    Request must be submitted within 30 days from end of financial year.
    """
    if current_user.get("role") not in ["hr_admin", "management", "managers"]:
        if current_user.get("employee_id") != data.employee_id:
            raise HTTPException(status_code=403, detail="Access denied")

    employee = await db.employees.find_one({"employee_id": data.employee_id})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # Check 3 years of service
    joining_str = employee.get("joining_date")
    if joining_str:
        joining = date.fromisoformat(joining_str)
        three_years = joining + timedelta(days=3 * 365)
        if date.today() < three_years:
            raise HTTPException(
                400,
                f"EL encashment is only allowed after 3 years of service. "
                f"Eligible from {three_years.strftime('%d %b %Y')}."
            )

    # Check minimum 30 EL balance
    balance = await db.leave_balances.find_one(
        {"employee_id": data.employee_id, "year": get_financial_year()}
    )
    el_remaining = (balance or {}).get("EL", {}).get("remaining", 0)
    if el_remaining < 30:
        raise HTTPException(
            400,
            f"Minimum 30 EL required for encashment. Current EL balance: {el_remaining} days."
        )
    if data.days_to_encash > el_remaining:
        raise HTTPException(400, f"Cannot encash more than available EL balance ({el_remaining} days).")

    doc = {
        "employee_id": data.employee_id,
        "type": "EL_encashment",
        "days_to_encash": data.days_to_encash,
        "remarks": data.remarks,
        "status": "pending",
        "requested_at": datetime.now(timezone.utc).isoformat(),
        "processed_by": None,
        "processed_at": None,
    }
    result = await db.leave_applications.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    doc.pop("_id", None)
    return doc


# ──────────────────────────────────────────────────────────────
#  Leave Balance Management (Admin) — Manual edit, Initialize, Bulk Excel
# ──────────────────────────────────────────────────────────────

BALANCE_KEYS = ["CL", "SL", "EL", "Marriage"]


class BalanceUpdateRequest(BaseModel):
    CL_total: float
    CL_used: float
    SL_total: float
    SL_used: float
    EL_total: float
    EL_used: float
    Marriage_total: float
    Marriage_used: float
    CompOff_total: Optional[float] = None   # None = do not touch; 0+ = manual override
    CompOff_used: Optional[float] = None
    reason: str


def _build_balance_snapshot(balance: dict) -> dict:
    """Return a clean {CL:{total,used,remaining}, ...} snapshot for audit."""
    snap = {}
    for k in BALANCE_KEYS:
        b = (balance or {}).get(k, {})
        snap[k] = {
            "total": round(float(b.get("total", 0)), 2),
            "used": round(float(b.get("used", 0)), 2),
            "remaining": round(float(b.get("remaining", 0)), 2),
        }
    return snap


async def _write_audit(
    employee_id: str,
    before: dict,
    after: dict,
    reason: str,
    changed_by: str,
    source: str,
):
    await db.leave_balance_audit.insert_one({
        "employee_id": employee_id,
        "year": get_financial_year(),
        "before": before,
        "after": after,
        "reason": reason,
        "changed_by": changed_by,
        "changed_at": datetime.now(timezone.utc).isoformat(),
        "source": source,  # "manual" | "bulk_upload" | "initialize"
    })


@router.post("/admin/initialize-balances")
async def initialize_balances(current_user: dict = Depends(get_current_user)):
    """Initialize a default leave balance row for any active/probation employee
    who doesn't yet have one for the current Financial Year.
    Defaults: CL 7, SL 15, EL 0, Marriage 5 (all unused)."""
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")

    fy = get_financial_year()
    employees = await db.employees.find(
        {"status": {"$in": ["active", "probation", "notice_period"]}},
        {"employee_id": 1, "_id": 0}
    ).to_list(2000)

    existing = await db.leave_balances.find(
        {"year": fy, "employee_id": {"$in": [e["employee_id"] for e in employees]}},
        {"employee_id": 1, "_id": 0}
    ).to_list(2000)
    existing_ids = {b["employee_id"] for b in existing}

    initialized = 0
    changed_by = current_user.get("employee_id") or current_user.get("username")
    for emp in employees:
        emp_id = emp["employee_id"]
        if emp_id in existing_ids:
            continue
        new_balance = {
            "employee_id": emp_id,
            "year": fy,
            **{k: dict(v) for k, v in LEAVE_BALANCE_TEMPLATE.items()},
        }
        await db.leave_balances.insert_one(new_balance)
        await _write_audit(
            employee_id=emp_id,
            before={},
            after=_build_balance_snapshot(LEAVE_BALANCE_TEMPLATE),
            reason="Initial leave balance creation",
            changed_by=changed_by,
            source="initialize",
        )
        initialized += 1

    return {
        "message": f"Initialized leave balances for FY{fy}–{str(fy+1)[-2:]}.",
        "initialized": initialized,
        "skipped_existing": len(existing_ids),
    }


@router.put("/admin/balance/{employee_id}")
async def update_employee_balance(
    employee_id: str,
    data: BalanceUpdateRequest,
    current_user: dict = Depends(get_current_user),
):
    """Manually edit a single employee's leave balance for the current FY.
    Requires a mandatory reason — logged in leave_balance_audit."""
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")

    if not data.reason or not data.reason.strip():
        raise HTTPException(status_code=400, detail="Reason is required.")

    for k in BALANCE_KEYS:
        total = getattr(data, f"{k}_total")
        used = getattr(data, f"{k}_used")
        if total < 0 or used < 0:
            raise HTTPException(status_code=400, detail=f"{k}: total/used cannot be negative.")
        if used > total:
            raise HTTPException(status_code=400, detail=f"{k}: used ({used}) cannot exceed total ({total}).")

    # Validate optional Comp-Off fields
    if data.CompOff_total is not None or data.CompOff_used is not None:
        co_total = data.CompOff_total if data.CompOff_total is not None else 0.0
        co_used = data.CompOff_used if data.CompOff_used is not None else 0.0
        if co_total < 0 or co_used < 0:
            raise HTTPException(status_code=400, detail="Comp-Off: total/used cannot be negative.")
        if co_used > co_total:
            raise HTTPException(status_code=400, detail=f"Comp-Off: used ({co_used}) cannot exceed total ({co_total}).")

    employee = await db.employees.find_one({"employee_id": employee_id}, {"_id": 0, "employee_id": 1})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found.")

    fy = get_financial_year()
    existing = await db.leave_balances.find_one({"employee_id": employee_id, "year": fy})
    before_snap = _build_balance_snapshot(existing or {})

    new_payload = {}
    for k in BALANCE_KEYS:
        total = getattr(data, f"{k}_total")
        used = getattr(data, f"{k}_used")
        new_payload[k] = {"total": total, "used": used, "remaining": total - used}

    # Optionally store manual Comp-Off balance override in leave_balances
    if data.CompOff_total is not None and data.CompOff_used is not None:
        co_total = data.CompOff_total
        co_used = data.CompOff_used
        new_payload["Comp-Off"] = {"total": co_total, "used": co_used, "remaining": co_total - co_used}

    await db.leave_balances.update_one(
        {"employee_id": employee_id, "year": fy},
        {"$set": new_payload, "$setOnInsert": {"employee_id": employee_id, "year": fy}},
        upsert=True,
    )

    after_snap = _build_balance_snapshot(new_payload)
    if data.CompOff_total is not None and data.CompOff_used is not None:
        after_snap["Comp-Off"] = new_payload["Comp-Off"]
    before_co = (existing or {}).get("Comp-Off")
    if before_co:
        before_snap["Comp-Off"] = before_co

    await _write_audit(
        employee_id=employee_id,
        before=before_snap,
        after=after_snap,
        reason=data.reason.strip(),
        changed_by=current_user.get("employee_id") or current_user.get("username"),
        source="manual",
    )

    return {"message": "Leave balance updated.", "balance": new_payload}


@router.get("/admin/balances-template")
async def download_balances_template(current_user: dict = Depends(get_current_user)):
    """Download an Excel template pre-filled with current leave balances
    for all active employees. Edit totals/used and upload back."""
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")

    fy = get_financial_year()

    employees = await db.employees.find(
        {"status": {"$in": ["active", "probation", "notice_period"]}},
        {"employee_id": 1, "first_name": 1, "last_name": 1, "department": 1, "_id": 0}
    ).sort("employee_id", 1).to_list(2000)

    emp_ids = [e["employee_id"] for e in employees]
    balances = await db.leave_balances.find(
        {"year": fy, "employee_id": {"$in": emp_ids}}
    ).to_list(2000)
    bal_map = {b["employee_id"]: b for b in balances}

    wb = Workbook()
    ws = wb.active
    ws.title = f"Leave Balances FY{fy}"

    header_fill = PatternFill("solid", fgColor="1E2A47")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    headers = [
        "Employee ID", "Name", "Department",
        "CL Total", "CL Used",
        "SL Total", "SL Used",
        "EL Total", "EL Used",
        "Marriage Total", "Marriage Used",
        "Reason (required)",
    ]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    for row_idx, emp in enumerate(employees, 2):
        emp_id = emp["employee_id"]
        name = f"{emp.get('first_name','')} {emp.get('last_name','')}".strip()
        bal = bal_map.get(emp_id, LEAVE_BALANCE_TEMPLATE)
        row = [
            emp_id, name, emp.get("department", ""),
            round(float(bal.get("CL", {}).get("total", 7)), 2), round(float(bal.get("CL", {}).get("used", 0)), 2),
            round(float(bal.get("SL", {}).get("total", 15)), 2), round(float(bal.get("SL", {}).get("used", 0)), 2),
            round(float(bal.get("EL", {}).get("total", 0)), 2), round(float(bal.get("EL", {}).get("used", 0)), 2),
            round(float(bal.get("Marriage", {}).get("total", 5)), 2), round(float(bal.get("Marriage", {}).get("used", 0)), 2),
            "",
        ]
        for col_idx, val in enumerate(row, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = border

    widths = [12, 28, 18, 10, 10, 10, 10, 10, 10, 14, 14, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w

    ws.freeze_panes = "A2"

    # Help sheet
    help_ws = wb.create_sheet("Instructions")
    help_lines = [
        "Radhya MFI — Leave Balance Bulk Update Template",
        "",
        f"Financial Year: FY{fy}–{str(fy+1)[-2:]} (April {fy} – March {fy+1})",
        "",
        "How to use:",
        "  1. Edit the Total and Used columns as required.",
        "  2. Remaining is auto-calculated as Total − Used on save.",
        "  3. Enter a Reason in the last column for EACH row you modify.",
        "  4. Rows with a blank Reason will be SKIPPED (no changes applied).",
        "  5. Upload the saved file via the 'Bulk Upload' button on the Leaves page.",
        "",
        "Validation rules:",
        "  • Used cannot exceed Total.",
        "  • Values must be non-negative integers.",
        "  • Unknown Employee IDs will be reported as skipped.",
        "  • Every change is audit-logged with your user, timestamp, and the reason.",
    ]
    for i, line in enumerate(help_lines, 1):
        help_ws.cell(row=i, column=1, value=line)
    help_ws.column_dimensions["A"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Leave_Balances_FY{fy}.xlsx"'},
    )


@router.post("/admin/balances-upload")
async def upload_balances_template(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """Bulk update leave balances from a filled-in Excel template.
    Rows with blank 'Reason' are skipped. Every applied row writes an audit log."""
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted.")

    content = await file.read()
    try:
        wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Unable to read Excel file. Ensure it is a valid .xlsx file.")

    ws = wb.active
    fy = get_financial_year()
    changed_by = current_user.get("employee_id") or current_user.get("username")

    updated = 0
    skipped_no_reason = 0
    skipped_unknown = 0
    errors = []

    # Expected header: Employee ID(1), Name(2), Department(3), CL_T(4), CL_U(5),
    # SL_T(6), SL_U(7), EL_T(8), EL_U(9), Mar_T(10), Mar_U(11), Reason(12)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if row is None or not any(row):
            continue
        try:
            emp_id = str(row[0] or "").strip()
            if not emp_id:
                continue
            reason = str(row[11] or "").strip() if len(row) > 11 else ""
            if not reason:
                errors.append(f"Row {row_idx} ({emp_id}): Reason column is blank — row skipped.")
                skipped_no_reason += 1
                continue

            emp = await db.employees.find_one({"employee_id": emp_id}, {"_id": 0, "employee_id": 1})
            if not emp:
                skipped_unknown += 1
                errors.append(f"Row {row_idx}: Employee ID '{emp_id}' not found.")
                continue

            values = {}
            for k, col_t, col_u in [
                ("CL", 3, 4), ("SL", 5, 6), ("EL", 7, 8), ("Marriage", 9, 10),
            ]:
                total = round(float(row[col_t] or 0), 2)
                used = round(float(row[col_u] or 0), 2)
                if total < 0 or used < 0 or used > total:
                    raise ValueError(f"{k}: invalid total/used ({total}/{used})")
                values[k] = {"total": total, "used": used, "remaining": total - used}

            existing = await db.leave_balances.find_one({"employee_id": emp_id, "year": fy})
            before_snap = _build_balance_snapshot(existing or {})

            await db.leave_balances.update_one(
                {"employee_id": emp_id, "year": fy},
                {"$set": values, "$setOnInsert": {"employee_id": emp_id, "year": fy}},
                upsert=True,
            )
            await _write_audit(
                employee_id=emp_id,
                before=before_snap,
                after=_build_balance_snapshot(values),
                reason=reason,
                changed_by=changed_by,
                source="bulk_upload",
            )
            updated += 1
        except ValueError as ve:
            errors.append(f"Row {row_idx}: {ve}")
        except Exception as ex:
            errors.append(f"Row {row_idx}: {ex}")

    return {
        "message": f"Bulk update complete. {updated} updated, {skipped_no_reason} skipped (no reason), {skipped_unknown} skipped (unknown ID).",
        "updated": updated,
        "skipped_no_reason": skipped_no_reason,
        "skipped_unknown": skipped_unknown,
        "errors": errors[:50],
    }


@router.get("/admin/balance-audit")
async def balance_audit_log(
    employee_id: Optional[str] = None,
    limit: int = 200,
    current_user: dict = Depends(get_current_user),
):
    """Return the audit trail of leave-balance edits (most recent first)."""
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")
    q = {}
    if employee_id:
        q["employee_id"] = employee_id
    cursor = db.leave_balance_audit.find(q, {"_id": 0}).sort("changed_at", -1).limit(max(1, min(limit, 1000)))
    return await cursor.to_list(limit)


@router.get("/{leave_id}")
async def get_leave(leave_id: str, current_user: dict = Depends(get_current_user)):
    leave = await db.leave_applications.find_one({"_id": ObjectId(leave_id)})
    if not leave:
        raise HTTPException(status_code=404, detail="Not found")
    return leave_to_dict(leave)
