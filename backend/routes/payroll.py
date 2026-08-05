from fastapi import APIRouter, HTTPException, Depends, Response
from pydantic import BaseModel
from typing import Optional, List
from database import db
from auth_utils import get_current_user
from services.shift_rules import resolve_shift_for
from datetime import datetime, timezone, date, timedelta
import calendar as _cal
import uuid
from bson import ObjectId
from pymongo import ReturnDocument
from pymongo.errors import DuplicateKeyError
import io

router = APIRouter()


def _is_non_working_saturday(d: date, sat_rule: Optional[str]) -> bool:
    """Mirror of frontend isWeeklyOff() for Saturdays — see /app/frontend/src/utils/shiftRules.js"""
    if d.weekday() != 5:
        return False
    if not sat_rule or sat_rule == "all_working":
        return False
    if sat_rule == "all_off":
        return True
    week_in_month = (d.day - 1) // 7 + 1  # 1..5
    if sat_rule == "alt_1_3_off":
        return week_in_month in (1, 3, 5)
    if sat_rule == "alt_2_4_off":
        return week_in_month in (2, 4)
    return False


# ── Salary hold ───────────────────────────────────────────────────────────────
# When HR Admin accepts a resignation the employee's salary is held: the payroll
# record is still created and fully calculated, but it is kept out of the NEFT
# file (the only thing that actually moves money) until someone releases it.
#
# The flag is stored ON THE RECORD rather than derived from the employee's live
# status, for two reasons: auto_exit_employees_past_lwd() flips notice_period ->
# exited once the LWD passes, so a derived answer would change underneath us; and
# a stored flag leaves an audit trail of who held/released what, and when.

# Statuses whose money has not left yet, so a hold can still bite.
_UNPAID_STATUSES = ["draft", "processed"]

# The ONLY statuses that may appear in the NEFT file. An allow-list on purpose:
# the previous test was `!= "draft"`, which quietly made `cancelled`, `reversed`,
# an empty string and a missing status field all payable.
_PAYABLE_STATUSES = {"processed", "paid"}


def _oid(record_id: str) -> ObjectId:
    """A malformed id is a 404, not a 500."""
    try:
        return ObjectId(record_id)
    except Exception:
        raise HTTPException(status_code=404, detail="Not found")


def _parse_period(period: str) -> tuple:
    """(year, month) or a 400. Month 13 used to reach monthrange() and 500."""
    try:
        y, m = str(period).split("-")
        year, month = int(y), int(m)
    except (ValueError, AttributeError):
        raise HTTPException(status_code=400, detail="Period must be in YYYY-MM format")
    if not (1 <= month <= 12) or len(m) != 2 or year < 2000 or year > 2100:
        raise HTTPException(status_code=400, detail=f"'{period}' is not a valid month (YYYY-MM).")
    return year, month


def _clean_payment_date(value: str) -> str:
    """A payment date must be a real date, not free text written into paid_at."""
    v = (value or "").strip()
    if not v:
        return ""
    try:
        return date.fromisoformat(v[:10]).isoformat()
    except ValueError:
        raise HTTPException(status_code=400,
                            detail="Payment date must be a real date in YYYY-MM-DD format.")


def _days(n: float) -> str:
    """Day counts read as 5, not 5.0 — but half-days must survive as 0.5."""
    n = float(n or 0)
    return str(int(n)) if n == int(n) else f"{n:.1f}"


def _hold_doc(reason: str, actor: str) -> dict:
    return {
        "on_hold": True,
        "hold_reason": reason,
        "held_at": datetime.now(timezone.utc).isoformat(),
        "held_by": actor,
        "hold_eligible_at": None,   # stamped when the exit reaches `completed`
    }


async def hold_payroll_for_exit(employee_id: str, reason: str, actor: str) -> int:
    """Hold every not-yet-paid payroll record for an employee.

    Called when a resignation is fully approved. Records already marked `paid`
    are left alone — that money is gone and a flag won't bring it back.
    """
    res = await db.payroll_records.update_many(
        {"employee_id": employee_id, "status": {"$in": _UNPAID_STATUSES}, "on_hold": {"$ne": True}},
        {"$set": _hold_doc(reason, actor)},
    )
    return res.modified_count


async def mark_exit_holds_eligible(employee_id: str, actor: str) -> int:
    """Mark an employee's held records as ready to release.

    Eligibility is NOT a release: the money still doesn't move until an admin
    approves it. This only changes what HR sees on the Payroll page.
    """
    res = await db.payroll_records.update_many(
        {"employee_id": employee_id, "on_hold": True},
        {"$set": {"hold_eligible_at": datetime.now(timezone.utc).isoformat(),
                  "hold_eligible_by": actor}},
    )
    return res.modified_count


def _is_payslip_visible_to_employee(record: dict) -> bool:
    """Gating rule for non-admin payslip visibility:

    1. Status must be `paid` — draft and processed records are hidden until HR
       explicitly publishes (marks as paid) the payroll for the period.
    2. The period's month must have ENDED — i.e., today must be on or after
       the 1st of the month after the payroll period.

    HR Admin / Management bypass this and see everything.
    """
    # A held record should never reach `paid` anyway (publish/finalize refuse
    # it) — this is the backstop, so a stray `paid` can't show a payslip for
    # money that was never sent.
    if record.get("on_hold"):
        return False
    if record.get("status") != "paid":
        return False
    period = record.get("period", "") or ""
    try:
        y, m = map(int, period.split("-"))
    except (ValueError, AttributeError):
        return False
    # 1st of the month AFTER the payroll period
    next_y = y + (1 if m == 12 else 0)
    next_m = 1 if m == 12 else m + 1
    return date.today() >= date(next_y, next_m, 1)


def calc_payroll_components(emp: dict, days_in_month: int = 30, lop_days: float = 0.0,
                            non_employed_days: float = 0.0):
    """Compute salary components.

    Args:
        days_in_month: actual calendar days in the payroll month (28-31).
        lop_days: Loss-of-Pay days — days the employee WAS employed but is not
            being paid for (absent, unapproved).
        non_employed_days: days in this month before joining or after the last
            working day. Not LOP: the employee was not on the rolls at all.

    Payable days = days_in_month − non_employed_days − lop_days, floored at 0.
    All earnings and deductions pro-rate on payable_days / days_in_month.

    The two are kept apart deliberately. Both reduce pay identically, but only
    LOP is absence: printing a joiner's pre-joining days as "LOP" tells a new
    employee they were absent for most of their first month, and puts the same
    claim into the salary register.
    """
    salary = emp.get("salary", {})
    basic = salary.get("basic", 0)
    hra = salary.get("hra", 0)
    special = salary.get("special_allowance", 0)
    canteen = salary.get("canteen_allowance", 0)
    conveyance = salary.get("conveyance_allowance", 0)
    gross = basic + hra + special + canteen + conveyance

    # Floored at 0 so a stale manual LOP entry stacked on top of non-employed
    # days can never drive pay negative — worst case the month pays nothing.
    unpaid_days = float(non_employed_days) + float(lop_days)
    payable_days = max(0.0, float(days_in_month) - unpaid_days)
    lop_fraction = payable_days / float(days_in_month) if days_in_month > 0 else 1.0

    # Pro-rata whenever any day of the month is unpaid, for either reason.
    if payable_days < days_in_month and days_in_month > 0:
        gross_payable = round(gross * lop_fraction, 2)
        basic_payable = round(basic * lop_fraction, 2)
    else:
        gross_payable = gross
        basic_payable = basic

    # EPF: 12% of the basic ACTUALLY EARNED this month, capped at EPF_CAP.
    #
    # The cap is on the CONTRIBUTION, not on the wage, and is NOT pro-rated. It
    # therefore only bites once earned basic passes 15,000 (12% of which is the
    # cap). Someone on 21,100 basic who worked 5 of 31 days earned 3,403.23, is
    # nowhere near the ceiling, and contributes the full 12% of that — 408.39.
    #
    # `salary.epf_employee` on the employee master is an ENROLMENT FLAG only:
    # > 0 means PF-covered, blank or 0 means exempt. Its VALUE is deliberately
    # not used as the amount. HR is instructed to pre-cap it at 1,800, which
    # destroys the 12% for anyone whose basic exceeds 15,000 — reading it back
    # would then cap an already-capped number and under-deduct every LOP month.
    # Deriving from basic cannot drift out of sync with the salary either.
    EPF_CAP = 1800
    EPF_RATE = 0.12
    try:
        enrolled = float(salary.get("epf_employee") or 0) > 0
    except (TypeError, ValueError):
        enrolled = False
    if enrolled:
        epf_employee = round(min(basic_payable * EPF_RATE, EPF_CAP), 2)
        epf_employer = epf_employee  # Employer matches employee (same basis + cap)
    else:
        epf_employee = 0
        epf_employer = 0

    # ESIC: applicable when basic <= 21000, calculated on payable basic
    if basic <= 21000:
        esic_employee = round(basic_payable * 0.0075, 2)
        esic_employer = round(basic_payable * 0.0325, 2)
    else:
        esic_employee = 0
        esic_employer = 0

    # Gratuity provision (monthly)
    designation = (emp.get("designation") or "").strip().lower()
    if designation == "director":
        gratuity_monthly = 0
    else:
        gratuity_monthly = round((basic_payable * 15) / 26 / 12)

    ctc_monthly = round(gross_payable + epf_employer + esic_employer + gratuity_monthly)
    net_salary = round(gross_payable - epf_employee - esic_employee)

    return {
        "basic": basic_payable,
        "hra": round(hra * lop_fraction, 2) if lop_fraction < 1 else hra,
        "special_allowance": round(special * lop_fraction, 2) if lop_fraction < 1 else special,
        "canteen_allowance": round(canteen * lop_fraction, 2) if lop_fraction < 1 else canteen,
        "conveyance_allowance": round(conveyance * lop_fraction, 2) if lop_fraction < 1 else conveyance,
        "gross_salary": gross,
        "gross_payable": gross_payable,
        "epf_employee": epf_employee,
        "epf_employer": epf_employer,
        "esic_employee": esic_employee,
        "esic_employer": esic_employer,
        "gratuity_monthly": gratuity_monthly,
        "ctc_monthly": ctc_monthly,
        "net_salary": net_salary,
        "working_days": days_in_month,
        "present_days": payable_days,
        "lop_days": lop_days,
        "non_employed_days": float(non_employed_days),
    }


async def calculate_lop_days(employee_id: str, year: int, month: int, joining_date_str: str,
                             last_working_day_str: Optional[str] = None) -> tuple:
    """Auto-calculate the month's unpaid days, split by reason.

    Returns `(lop_days, non_employed_days)`.

    LOP = working days where the employee WAS on the rolls but has no punch-in,
    no approved leave, and the day is not a public holiday, Sunday, or a
    non-working Saturday (resolved per the employee's shift saturday_rule).
    Half-day attendance is counted as present (not LOP).

    NON-EMPLOYED = days in this month before the joining date or after the last
    working day. Weekly offs and holidays included — they are not on the rolls,
    so those days are not paid either.

    The two reduce pay identically but are NOT interchangeable. Counting
    pre-joining days as LOP would print "LOP 26" on a joiner's first payslip,
    which reads as a month of unauthorised absence; counting them as neither —
    which is what happened before — silently paid a joiner for a full month.
    """
    days_in_month = _cal.monthrange(year, month)[1]
    period = f"{year}-{month:02d}"

    try:
        joining_date = date.fromisoformat(joining_date_str[:10])
    except Exception:
        joining_date = date(year, month, 1)

    try:
        lwd = date.fromisoformat(last_working_day_str[:10]) if last_working_day_str else None
    except Exception:
        lwd = None

    # Resolve the employee's effective shift to get saturday_rule
    emp = await db.employees.find_one({"employee_id": employee_id}, {"_id": 0, "role": 1}) or {}
    shift = await resolve_shift_for(emp.get("role", "employee"), employee_id, db)
    sat_rule = (shift or {}).get("saturday_rule", "all_working")

    # Fetch attendance records
    att_records = await db.attendance_records.find(
        {"employee_id": employee_id, "date": {"$regex": f"^{period}"}}
    ).to_list(35)
    att_by_date = {r["date"]: r for r in att_records}

    # Fetch holidays
    holidays_list = await db.holidays.find(
        {"date": {"$regex": f"^{period}"}}
    ).to_list(35)
    holiday_dates = {h["date"] for h in holidays_list}

    # Fetch all approved leaves and expand into individual dates.
    #
    # LWP is approved absence that is explicitly NOT paid, so it lands in
    # `lwp_weights` and is charged as LOP. Every other leave type is paid and
    # lands in `paid_leave_dates`. Reading them as one bucket — which is what
    # happened before — paid people for leave-without-pay.
    #
    # Weights mirror calc_days() in routes/leaves.py: a single-day first_half or
    # second_half is 0.5, and on a range start_half/end_half take half off the
    # first and last day respectively.
    #
    # Scoped to leaves that OVERLAP this month. Fetching an employee's entire
    # approved history and capping at 500 dropped rows in insertion order — i.e.
    # the newest first — so a long-serving employee's current-month LWP fell off
    # the end and was paid in full.
    last_day = f"{period}-{days_in_month:02d}"
    approved_leaves = await db.leave_applications.find({
        "employee_id": employee_id,
        "status": "approved",
        "start_date": {"$lte": f"{last_day}￿"},
        "end_date": {"$gte": f"{period}-01"},
    }).to_list(500)
    paid_leave_dates: set = set()
    paid_leave_weights: dict = {}
    lwp_weights: dict = {}
    for lv in approved_leaves:
        try:
            s = date.fromisoformat(str(lv.get("start_date", ""))[:10])
            e = date.fromisoformat(str(lv.get("end_date", ""))[:10])
            is_lwp = str(lv.get("leave_type", "")).strip().upper() == "LWP"
            day_type = str(lv.get("day_type", "full_day") or "full_day")
            cur = s
            while cur <= e:
                if cur.year == year and cur.month == month:
                    iso = cur.isoformat()
                    if s == e:
                        w = 0.5 if day_type in ("first_half", "second_half") else 1.0
                    elif cur == s and lv.get("start_half"):
                        w = 0.5
                    elif cur == e and lv.get("end_half"):
                        w = 0.5
                    else:
                        w = 1.0
                    # Overlapping applications: charge the larger claim once,
                    # never the sum — a day cannot cost more than itself.
                    if is_lwp:
                        lwp_weights[iso] = max(lwp_weights.get(iso, 0.0), w)
                    else:
                        # Every paid-leave date still counts as paid, whatever
                        # its weight — unchanged from before. The weight is only
                        # used to offset an LWP sharing the same day.
                        paid_leave_dates.add(iso)
                        paid_leave_weights[iso] = max(paid_leave_weights.get(iso, 0.0), w)
                cur += timedelta(days=1)
        except Exception:
            pass
    # A paid leave sharing a date caps the LWP rather than cancelling it: half
    # CL + half LWP is a routine combination and they cover DIFFERENT halves, so
    # subtracting one from the other would pay half a day for nothing.
    #
    # A day is worth 1.0. Paid leave covers `p` of it, so at most `1 - p` is
    # left to charge — and never more than the LWP itself claims:
    #   0.5 CL + 0.5 LWP -> min(0.5, 0.5) = 0.5   (disjoint halves)
    #   1.0 CL + 1.0 LWP -> min(1.0, 0.0) = 0     (fully covered)
    #   0.5 CL + 1.0 LWP -> min(1.0, 0.5) = 0.5   (half covered)
    for iso, paid_w in paid_leave_weights.items():
        if iso in lwp_weights:
            chargeable = round(min(lwp_weights[iso], max(0.0, 1.0 - paid_w)), 2)
            if chargeable > 0:
                lwp_weights[iso] = chargeable
            else:
                del lwp_weights[iso]

    today = date.today()
    lop = 0.0
    non_employed = 0.0
    for day in range(1, days_in_month + 1):
        d = date(year, month, day)

        # Outside the employment window — before joining, or past the last
        # working day. Weekly offs and holidays included: they are not on the
        # rolls, so those days are not paid. Checked before the `d > today` skip
        # below, because days after the LWD are unpaid whether or not they have
        # happened yet, and the same is true of a future-dated joining date.
        if d < joining_date or (lwd and d > lwd):
            non_employed += 1.0
            continue

        # Skip future dates — the month is not over yet
        if d > today:
            continue

        date_str = d.isoformat()

        # Sunday = weekly off — not LOP
        if d.weekday() == 6:
            continue

        # Non-working Saturday per employee's shift saturday_rule — not LOP
        if _is_non_working_saturday(d, sat_rule):
            continue

        # Public holiday — not LOP
        if date_str in holiday_dates:
            continue

        # Approved leave-without-pay — charged, at half for a half-day, less any
        # paid leave covering part of the same day. Checked BEFORE the paid-leave
        # skip so that half CL + half LWP costs 0.5 rather than nothing. Placed
        # after the Sunday/Saturday/holiday checks on purpose: an LWP spanning a
        # weekly off must not turn that day into a deduction.
        lwp = lwp_weights.get(date_str)
        if lwp:
            lop += lwp
            continue

        # Covered by a PAID approved leave — not LOP
        if date_str in paid_leave_dates:
            continue

        # Check attendance status
        att = att_by_date.get(date_str)
        if att:
            status = att.get("status", "")
            if status in ("present", "half_day", "leave", "weekly_off", "holiday"):
                continue
            # status == "absent" or unexpected → LOP
            lop += 1.0
        else:
            # No attendance record on a working day → absent → LOP
            lop += 1.0

    return lop, non_employed


async def _with_bank_status(records: list) -> list:
    """Tag each record with whether the employee's bank account is verified.

    An unverified account keeps someone out of the NEFT sheet entirely, so HR needs
    to see that on the Payroll page — before the bank run, not by noticing an absence
    in a spreadsheet afterwards.
    """
    ids = [r.get("employee_id") for r in records if r.get("employee_id")]
    if not ids:
        return records
    emps = await db.employees.find(
        {"employee_id": {"$in": ids}}, {"employee_id": 1, "bank_details": 1, "_id": 0}
    ).to_list(None)
    verified = {e["employee_id"] for e in emps
                if (e.get("bank_details") or {}).get("verified") is True}
    for r in records:
        r["bank_verified"] = r.get("employee_id") in verified
    return records


def pay_to_dict(p):
    p["id"] = str(p.pop("_id"))
    # Records written before salary holds existed have no on_hold field — normalise
    # so the UI never has to distinguish "not held" from "missing".
    p["on_hold"] = bool(p.get("on_hold"))
    p["hold_eligible"] = bool(p.get("hold_eligible_at"))
    # Records processed before joiners/exits were separated from LOP have no
    # such field — normalise so the UI can always read it as a number.
    p["non_employed_days"] = float(p.get("non_employed_days") or 0)
    # Always expose a derived total_deductions field for UI/exports
    p["total_deductions"] = round(
        float(p.get("epf_employee") or 0)
        + float(p.get("esic_employee") or 0)
        + float(p.get("tds") or 0)
        + float(p.get("other_deductions") or 0),
        2,
    )
    return p


class ProcessPayrollRequest(BaseModel):
    month: int
    year: int
    employee_ids: Optional[List[str]] = None  # None = process all active


class PayrollUpdateRequest(BaseModel):
    tds: Optional[float] = 0
    other_deductions: Optional[float] = 0
    other_additions: Optional[float] = 0
    # Loss-of-pay days set by HR (supports half-day, e.g. 0.5, 1, 1.5).
    # When provided, pro-rata recomputes Basic/HRA/allowances/EPF/ESIC.
    lop_days: Optional[float] = None
    remarks: Optional[str] = None


@router.post("/process")
async def process_payroll(data: ProcessPayrollRequest, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")
    period = f"{data.year}-{data.month:02d}"

    # Who gets a payroll record this month.
    #
    # `notice_period` staff are still employed and still working, so they are paid
    # on the normal cycle like everyone else — their record is simply born held.
    #
    # The `exited` arm catches the final part-month: auto_exit_employees_past_lwd()
    # flips someone to `exited` the day after their LWD, so an employee who left on
    # the 15th is already `exited` by the time payroll runs on the 30th. Without
    # this they would silently get nothing for the days they did work. Matching on
    # the LWD falling inside this period keeps it to that one month — someone who
    # left in March does not reappear in September's run.
    query = {
        "$or": [
            {"status": {"$in": ["active", "probation", "notice_period"]}},
            {"status": "exited", "last_working_day": {"$regex": f"^{period}"}},
        ]
    }
    if data.employee_ids:
        query["employee_id"] = {"$in": data.employee_ids}
    employees = await db.employees.find(query).to_list(1000)
    days_in_month = _cal.monthrange(data.year, data.month)[1]
    processed = []
    held = []
    for emp in employees:
        emp_id = emp["employee_id"]
        existing = await db.payroll_records.find_one({"employee_id": emp_id, "period": period})
        if existing:
            continue
        # Auto-calculate LOP: absent days with no approved leave, no holiday, not Sunday
        joining_date_str = emp.get("joining_date", f"{data.year}-{data.month:02d}-01") or f"{data.year}-{data.month:02d}-01"
        # Only pass the LWD once it is set (i.e. the exit was approved) so it can
        # cut off pay mid-month. Everyone else is unaffected.
        lwd_str = emp.get("last_working_day") or None
        lop_days, non_employed_days = await calculate_lop_days(
            emp_id, data.year, data.month, joining_date_str, lwd_str)
        components = calc_payroll_components(emp, days_in_month, lop_days, non_employed_days)

        # An accepted resignation means the salary is held from the moment of
        # acceptance through to clearance.
        emp_status = emp.get("status")
        is_exiting = emp_status == "notice_period" or (emp_status == "exited" and lwd_str)
        hold = _hold_doc(
            f"Resignation accepted — held pending exit clearance"
            + (f" (last working day {str(lwd_str)[:10]})" if lwd_str else ""),
            current_user.get("employee_id") or current_user.get("username") or "System",
        ) if is_exiting else {"on_hold": False}

        record = {
            "employee_id": emp_id,
            "employee_name": f"{emp.get('first_name', '')} {emp.get('last_name', '')}",
            "designation": emp.get("designation", ""),
            "department": emp.get("department", ""),
            "period": period,
            "month": data.month,
            "year": data.year,
            **components,
            # Marks the LOP figure as machine-computed. A manual edit flips this
            # to "manual" and Recalculate LOP then leaves the record alone.
            "lop_source": "auto",
            # The salary this month was actually run on. Recalculate LOP prices
            # against THIS, not the live master — otherwise a raise given in July
            # would silently re-rate a recalculated March payslip at July's pay.
            "salary_basis": emp.get("salary", {}),
            "tds": 0,
            "other_deductions": 0,
            "other_additions": 0,
            "bank_account": (emp.get("bank_details") or {}).get("account_number", ""),
            "ifsc_code": (emp.get("bank_details") or {}).get("ifsc_code", ""),
            "bank_name": (emp.get("bank_details") or {}).get("bank_name", ""),
            "status": "draft",
            **hold,
            "processed_at": datetime.now(timezone.utc).isoformat(),
            "processed_by": current_user.get("employee_id"),
        }
        # ONE atomic operation instead of the earlier find_one-then-insert.
        # Two clicks racing each other both passed that check and both inserted,
        # producing two records for one month — two rows in the NEFT file and
        # the salary paid twice. $setOnInsert writes only when nothing matched,
        # and with the unique (employee_id, period) index the loser of the race
        # raises DuplicateKeyError instead of creating a second record.
        try:
            res = await db.payroll_records.update_one(
                {"employee_id": emp_id, "period": period},
                {"$setOnInsert": record},
                upsert=True,
            )
        except DuplicateKeyError:
            continue          # another run got there first — not an error
        if res.upserted_id is None:
            continue          # a record already existed; leave it untouched
        processed.append(emp_id)
        if is_exiting:
            held.append(emp_id)
    return {"processed": len(processed), "employee_ids": processed, "period": period,
            "held": len(held), "held_employee_ids": held}


@router.delete("/period/{period}")
async def delete_payroll_period(period: str, current_user: dict = Depends(get_current_user)):
    """Delete all payroll records for a given period (YYYY-MM).
    Allowed only until the 15th of the following month
    (e.g. April 2026 records can be deleted up to and including 15-May-2026).
    HR Admin / Management only.
    """
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")

    # Parse and validate period
    try:
        y, m = period.split("-")
        period_year, period_month = int(y), int(m)
        if not (1 <= period_month <= 12) or period_year < 2000 or period_year > 2100:
            raise ValueError
    except (ValueError, AttributeError):
        raise HTTPException(status_code=400, detail="Period must be in YYYY-MM format")

    # Cutoff = 15th of the month AFTER the payroll period
    if period_month == 12:
        cutoff_year, cutoff_month = period_year + 1, 1
    else:
        cutoff_year, cutoff_month = period_year, period_month + 1
    cutoff = datetime(cutoff_year, cutoff_month, 15, 23, 59, 59, tzinfo=timezone.utc)

    now = datetime.now(timezone.utc)
    if now > cutoff:
        raise HTTPException(
            status_code=403,
            detail=f"Cannot delete payroll for {period}. The deletion window closed on "
                   f"{cutoff.strftime('%d %b %Y')}. Edit individual records instead.",
        )

    res = await db.payroll_records.delete_many({"period": period})
    # The bank file logged for this period described records that no longer
    # exist. Left standing, it would tell `Mark All Paid` that freshly
    # re-processed records — with different amounts — had already been sent,
    # and mark them paid off a file that never contained them.
    #
    # Voided rather than deleted: a download that really did go to the bank is
    # history worth keeping, and `_exported_amounts` skips voided rows.
    voided = await db.payroll_exports.update_many(
        {"period": period, "voided_at": {"$exists": False}},
        {"$set": {"voided_at": now.isoformat(),
                  "voided_by": current_user.get("employee_id") or current_user.get("username"),
                  "voided_reason": "payroll for this period was deleted and must be re-processed"}},
    )
    return {
        "exports_voided": voided.modified_count,
        "period": period,
        "deleted": res.deleted_count,
        "cutoff": cutoff.isoformat(),
    }


@router.get("")
async def list_payroll(
    period: str = None,
    employee_id: str = None,
    current_user: dict = Depends(get_current_user),
):
    query = {}
    if period:
        query["period"] = period
    role = current_user.get("role")
    my_emp_id = current_user.get("employee_id")
    if role in ["employee", "field_agent", "managers"]:
        # managers see ONLY their own payroll (not their team's)
        query["employee_id"] = my_emp_id
    elif employee_id:
        query["employee_id"] = employee_id
    records = await db.payroll_records.find(query).sort([("period", -1), ("employee_id", 1)]).to_list(500)
    # Gate visibility for non-admin roles: only paid AND past month-end
    if role not in ["hr_admin", "management"]:
        records = [r for r in records if _is_payslip_visible_to_employee(r)]
    return await _with_bank_status([pay_to_dict(r) for r in records])


@router.get("/employee/{employee_id}")
async def employee_payroll(employee_id: str, current_user: dict = Depends(get_current_user)):
    role = current_user.get("role")
    my_emp_id = current_user.get("employee_id")
    # Non-admin roles can only fetch their own payroll history
    if role not in ["hr_admin", "management"] and employee_id != my_emp_id:
        raise HTTPException(status_code=403, detail="Access denied")
    records = await db.payroll_records.find({"employee_id": employee_id}).sort("period", -1).to_list(50)
    # Gate visibility for non-admin roles
    if role not in ["hr_admin", "management"]:
        records = [r for r in records if _is_payslip_visible_to_employee(r)]
    return [pay_to_dict(r) for r in records]


@router.put("/{record_id}")
async def update_payroll(record_id: str, data: PayrollUpdateRequest, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")
    record = await db.payroll_records.find_one({"_id": _oid(record_id)})
    if not record:
        raise HTTPException(status_code=404, detail="Not found")
    # A paid record is a statement that money left the bank. Editing it used to
    # silently set the status back to `processed`, pulling the payslip out of the
    # employee's view and — since LOP edits mark the record manual — excluding it
    # from every future recalculation. Two invisible consequences of one click.
    if record.get("status") == "paid":
        raise HTTPException(status_code=409, detail=(
            "This payslip is marked as paid and cannot be edited. Reopen it first if the "
            "figures genuinely need to change — that is recorded, with a reason."))
    tds = data.tds or 0
    other_ded = data.other_deductions or 0
    other_add = data.other_additions or 0
    # A negative deduction is an addition wearing a disguise: -50,000 in
    # "other deductions" paid out five times the salary. None of these three
    # fields has a meaning below zero.
    for _label, _v in (("TDS", tds), ("Other deductions", other_ded),
                       ("Other additions", other_add)):
        if float(_v) < 0:
            raise HTTPException(status_code=400, detail=(
                f"{_label} cannot be negative. To pay an employee more, use Other Additions; "
                f"to deduct, use Other Deductions."))

    update_doc = {
        "tds": tds,
        "other_deductions": other_ded,
        "other_additions": other_add,
        "remarks": data.remarks,
        "status": "processed",
    }

    # If HR manually overrides LOP days, recompute pro-rata earnings + EPF/ESIC.
    #
    # The record's non-employed days are carried through untouched. They are not
    # HR's to edit here — they come from the joining date and the last working
    # day — and dropping them would silently restore a joiner to a full month.
    warning = None
    if data.lop_days is not None:
        emp = await db.employees.find_one({"employee_id": record.get("employee_id")})
        days_in_month = record.get("working_days", 30) or 30
        non_employed = float(record.get("non_employed_days", 0) or 0)
        lop = max(0.0, float(data.lop_days))
        # Days outside employment are already deducted. HR used to type them into
        # this box by hand because nothing else did it; doing that now deducts
        # them twice. Say so rather than quietly halving someone's salary.
        if non_employed > 0 and lop > 0:
            warning = (
                f"This month already excludes {_days(non_employed)} non-employed "
                f"day(s) (before joining or after the last working day). The "
                f"{_days(lop)} LOP day(s) entered here are deducted on top of "
                f"those — enter only days the employee was on the rolls and absent."
            )
        if lop + non_employed > days_in_month:
            warning = (
                f"{_days(lop)} LOP day(s) plus {_days(non_employed)} non-employed "
                f"day(s) exceeds the {days_in_month} days in the month — this "
                f"record now pays nothing. Check the LOP figure."
            )
        # Price against the salary the month was RUN on, exactly as the
        # recalculation does. Reading the live master here re-rated a closed
        # month at today's pay — and because the payslip form posts lop_days on
        # every save, an unrelated TDS or remarks edit was enough to do it.
        basis = record.get("salary_basis")
        emp_for_pricing = ({"designation": record.get("designation") or (emp or {}).get("designation"),
                            "salary": basis} if isinstance(basis, dict) and basis else emp)
        if emp_for_pricing:
            new_components = calc_payroll_components(
                emp_for_pricing, days_in_month, lop, non_employed)
            update_doc.update(new_components)
        elif emp is None:
            raise HTTPException(status_code=409, detail=(
                "This payslip has no salary snapshot and its employee record no longer exists, "
                "so the figures cannot be recalculated. Restore the employee record first."))
        update_doc["lop_days"] = lop
        # Only a CHANGED figure means HR has taken ownership. The payslip form
        # posts all five fields every time, pre-filled from the record, so
        # marking manual on every save would freeze a record's LOP for good
        # after an unrelated TDS or remarks edit.
        if lop != float(record.get("lop_days") or 0):
            update_doc["lop_source"] = "manual"

    gross_payable = update_doc.get("gross_payable", record.get("gross_payable", 0))
    epf_employee = update_doc.get("epf_employee", record.get("epf_employee", 0))
    esic_employee = update_doc.get("esic_employee", record.get("esic_employee", 0))

    net_salary = round(gross_payable - epf_employee - esic_employee - tds - other_ded + other_add)
    # Deductions larger than the month's pay are a real situation — an advance
    # being recovered — but the answer is to recover it across months, not to
    # write a negative salary. A negative row in the NEFT file is an instruction
    # to the bank to take money OUT of the employee's account.
    if net_salary < 0:
        raise HTTPException(status_code=400, detail=(
            f"These deductions come to more than the salary for this month — the net would be "
            f"₹{net_salary:,.0f}. Reduce the deduction and recover the balance over the "
            f"following months instead; a negative salary cannot be paid."))
    update_doc["net_salary"] = net_salary
    update_doc["total_deductions"] = round(epf_employee + esic_employee + tds + other_ded, 2)

    await db.payroll_records.update_one(
        {"_id": _oid(record_id)},
        {"$set": update_doc},
    )
    record = await db.payroll_records.find_one({"_id": _oid(record_id)})
    out = pay_to_dict(record)
    if warning:
        out["warning"] = warning
    return out


class FinalizeRequest(BaseModel):
    reason: Optional[str] = None
    payment_date: Optional[str] = None


class ReopenRequest(BaseModel):
    reason: str


@router.post("/{record_id}/finalize")
async def finalize_payroll(record_id: str, data: FinalizeRequest = None,
                           current_user: dict = Depends(get_current_user)):
    """Mark one record paid.

    This is the exception path. If the employee was in a NEFT file it behaves as
    before. If they were NOT — no bank details, cheque, manual transfer — it
    demands a written reason, because that is a claim no export can corroborate.
    """
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")
    record = await db.payroll_records.find_one({"_id": _oid(record_id)})
    if not record:
        raise HTTPException(status_code=404, detail="Not found")
    if record.get("on_hold"):
        raise HTTPException(
            status_code=400,
            detail="This salary is on hold and cannot be marked as paid. Release the hold first.",
        )
    if record.get("status") == "paid":
        raise HTTPException(status_code=409, detail="This payslip is already marked as paid.")
    # A draft has never been reviewed by anyone. Bulk publish refuses them; this
    # route must too, or it becomes the way round that.
    if record.get("status") == "draft":
        raise HTTPException(status_code=400, detail=(
            "This payslip is still a draft — nobody has reviewed the figures. Save adjustments "
            "to move it to Processed before marking it paid."))
    # Nothing left the bank for a zero or negative net, so it cannot be recorded
    # as paid — the NEFT file withholds these for the same reason.
    if round(float(record.get("net_salary") or 0), 2) <= 0:
        raise HTTPException(status_code=400, detail=(
            "This payslip has a net of zero or less, so there is no payment to record. "
            "Correct the deductions first."))

    data = data or FinalizeRequest()
    reason = (data.reason or "").strip()
    sent_ids, _ = await _exported_amounts(record.get("period", ""))
    outside = record.get("employee_id") not in sent_ids
    if outside and not reason:
        raise HTTPException(status_code=400, detail=(
            "This employee was not in any NEFT file for this month, so the system cannot "
            "confirm the payment. Give a reason (for example: paid by cheque, or transferred "
            "manually) to record it anyway."))

    now = datetime.now(timezone.utc).isoformat()
    res = await db.payroll_records.update_one(
        # Re-assert the preconditions in the filter itself: the checks above are
        # a read, and a publish or an exit hold can land between read and write.
        {"_id": _oid(record_id), "status": {"$ne": "paid"}, "on_hold": {"$ne": True}},
        {"$set": {"status": "paid",
                  "paid_at": _clean_payment_date(data.payment_date) or now,
                  "published_at": now,
                  "published_by": current_user.get("employee_id") or current_user.get("username"),
                  "paid_outside_neft": bool(outside),
                  "paid_exception_reason": reason if outside else None}},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=409, detail=(
            "This payslip changed while you were marking it paid — it is now on hold or already "
            "paid. Reopen the payslip to see its current state."))
    return {"message": "Payroll marked as paid", "paid_outside_neft": bool(outside)}


@router.post("/{record_id}/reopen")
async def reopen_payslip(record_id: str, data: ReopenRequest,
                         current_user: dict = Depends(get_current_user)):
    """Take a record back out of `paid` so it can be corrected.

    Editing a paid record used to do this silently — Save Adjustments set the
    status back to `processed` and pulled the payslip out of the employee's view
    with no warning. Reopening is now a deliberate act with a reason attached.
    """
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")
    reason = (data.reason or "").strip()
    if not reason:
        raise HTTPException(status_code=400, detail="Give a reason for reopening this payslip.")
    record = await db.payroll_records.find_one({"_id": _oid(record_id)})
    if not record:
        raise HTTPException(status_code=404, detail="Not found")
    if record.get("status") != "paid":
        raise HTTPException(status_code=400, detail="This payslip is not marked as paid.")
    await db.payroll_records.update_one(
        {"_id": _oid(record_id)},
        {"$set": {"status": "processed",
                  "reopened_at": datetime.now(timezone.utc).isoformat(),
                  "reopened_by": current_user.get("employee_id") or current_user.get("username"),
                  "reopen_reason": reason,
                  "reopened_from_paid_at": record.get("paid_at")},
         # The exception fields describe a payment that has just been withdrawn.
         # Leaving them behind would tag a later, ordinary NEFT payment as a
         # manual cheque.
         "$unset": {"paid_at": "", "published_at": "",
                    "paid_outside_neft": "", "paid_exception_reason": ""}},
    )
    record = await db.payroll_records.find_one({"_id": _oid(record_id)})
    return pay_to_dict(record)


class ReleaseHoldRequest(BaseModel):
    note: Optional[str] = None


@router.post("/{record_id}/release-hold")
async def release_hold(record_id: str, data: ReleaseHoldRequest,
                       current_user: dict = Depends(get_current_user)):
    """Release a held salary so it enters the next NEFT export.

    Reaching `completed` on the exit only makes a record *eligible*; nothing is
    paid until an admin calls this. Releasing before the exit is complete is
    allowed but recorded as an override, so the trail shows it was a judgement
    call rather than the normal path.
    """
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")
    record = await db.payroll_records.find_one({"_id": _oid(record_id)})
    if not record:
        raise HTTPException(status_code=404, detail="Not found")
    if not record.get("on_hold"):
        raise HTTPException(status_code=400, detail="This salary is not on hold.")

    override = not record.get("hold_eligible_at")
    if override and not (data.note or "").strip():
        raise HTTPException(
            status_code=400,
            detail="The exit is not complete yet. Give a reason to release this salary early.",
        )

    await db.payroll_records.update_one(
        {"_id": _oid(record_id)},
        {"$set": {
            "on_hold": False,
            "released_at": datetime.now(timezone.utc).isoformat(),
            "released_by": current_user.get("employee_id") or current_user.get("username"),
            "release_note": (data.note or "").strip(),
            "release_override": override,
        }},
    )
    record = await db.payroll_records.find_one({"_id": _oid(record_id)})
    return pay_to_dict(record)


async def _exported_amounts(period: str) -> tuple:
    """(employee_ids sent to the bank, last amount sent per employee).

    The UNION of every export for the period, not just the latest: someone whose
    bank details were corrected between two downloads appears only in the second
    file, and they were genuinely sent. Later files win on the amount.

    `rows` only exists on exports taken after per-employee amounts were logged,
    so the amount map is best-effort — membership is not.
    """
    exports = await db.payroll_exports.find(
        {"period": period, "voided_at": {"$exists": False}}
    ).sort("exported_at", 1).to_list(200)
    ids, amounts = set(), {}
    for ex in exports:
        ids.update(ex.get("employee_ids") or [])
        for row in (ex.get("rows") or []):
            eid = row.get("employee_id")
            if eid:
                ids.add(eid)
                amounts[eid] = float(row.get("net_salary") or 0)
    return ids, amounts


async def _score_publish(period: str) -> dict:
    """Who would be marked paid, who would not, and why. Writes nothing.

    A record is only marked paid if it was actually sent to the bank. "Paid" used
    to mean "somebody clicked a button"; gating it on NEFT membership makes it
    mean "was in a file HR downloaded and gave to the bank".
    """
    _parse_period(period)
    sent_ids, sent_amounts = await _exported_amounts(period)
    records = await db.payroll_records.find({"period": period}).to_list(4000)

    to_pay, mismatched = [], []
    skipped = {"held": [], "draft": [], "not_in_neft": [], "already_paid": [],
               "nonpositive": []}
    for r in records:
        eid = r.get("employee_id")
        row = {"record_id": str(r.get("_id")), "employee_id": eid,
               "employee_name": r.get("employee_name", "") or eid,
               "net_salary": round(float(r.get("net_salary") or 0), 2)}
        if r.get("status") == "paid":
            skipped["already_paid"].append(row)
        elif r.get("on_hold"):
            skipped["held"].append(row)
        elif r.get("status") == "draft":
            skipped["draft"].append(row)
        elif row["net_salary"] <= 0:
            skipped["nonpositive"].append(row)
        elif eid not in sent_ids:
            skipped["not_in_neft"].append(row)
        else:
            sent = sent_amounts.get(eid)
            # Warn, never block: a changed amount is a fact to look into, and
            # blocking would strand the record with no way forward.
            if sent is not None and round(sent, 2) != row["net_salary"]:
                row = {**row, "sent_amount": round(sent, 2),
                       "difference": round(row["net_salary"] - sent, 2)}
                mismatched.append(row)
            to_pay.append(row)

    return {
        "period": period,
        "any_export": bool(sent_ids),
        "will_pay": len(to_pay),
        "rows": to_pay,
        "mismatched": mismatched,
        "skipped": skipped,
        "total_amount": round(sum(x["net_salary"] for x in to_pay), 2),
    }


@router.post("/approve")
async def approve_period(period: str, current_user: dict = Depends(get_current_user)):
    """Move a month's draft records to `processed` — approved for payment.

    This step used to be done, unknowingly, by "Mark All Paid": it moved draft
    records straight to `paid`, which is why `paid` meant nothing. Now that paid
    means "the bank sent this", the approval it was standing in for needs to
    exist on its own — otherwise a freshly processed month is all drafts, the
    NEFT sheet excludes drafts, and nothing can ever be exported or paid.

    Held records are approved too: a hold stops the money leaving, not the
    figures being reviewed, and the NEFT export and publish both skip them
    independently.
    """
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")
    _parse_period(period)
    drafts = await db.payroll_records.find(
        {"period": period, "status": "draft"}, {"_id": 1, "employee_id": 1}).to_list(4000)
    if not drafts:
        return {"period": period, "approved": 0, "employee_ids": [],
                "message": "Nothing to approve — no draft records for this month."}
    res = await db.payroll_records.update_many(
        {"period": period, "status": "draft"},
        {"$set": {"status": "processed",
                  "approved_at": datetime.now(timezone.utc).isoformat(),
                  "approved_by": current_user.get("employee_id") or current_user.get("username")}},
    )
    return {"period": period, "approved": res.modified_count,
            "employee_ids": sorted(d.get("employee_id") for d in drafts if d.get("employee_id"))}


@router.get("/publish/preview")
async def preview_publish(period: str, current_user: dict = Depends(get_current_user)):
    """What Mark All Paid would do. Read-only."""
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")
    return await _score_publish(period)


@router.post("/publish")
async def publish_payslips(period: str, payment_date: str = None,
                           current_user: dict = Depends(get_current_user)):
    """Mark as paid every `processed` record that was actually sent to the bank.

    Held records are skipped — marking one paid would publish a payslip for money
    that was never sent. Drafts are skipped too; they are already absent from the
    NEFT sheet, so this is belt-and-braces.

    Someone paid outside the bank run (no bank details, cheque, manual transfer)
    is never in the file and is never caught here. That is deliberate: they go
    through POST /{record_id}/finalize with a written reason, one at a time.
    """
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")
    scored = await _score_publish(period)

    if not scored["any_export"]:
        raise HTTPException(status_code=409, detail=(
            f"No NEFT file has been downloaded for {period}, so nothing can be confirmed as "
            f"paid. Export the sheet and complete the bank transfer first. If someone was "
            f"paid another way, mark that payslip paid individually with a reason."))

    when = _clean_payment_date(payment_date) or datetime.now(timezone.utc).isoformat()
    actor = current_user.get("employee_id") or current_user.get("username")
    ids = [ObjectId(x["record_id"]) for x in scored["rows"]]
    published = 0
    if ids:
        res = await db.payroll_records.update_many(
            # Scoring happened before this write. Repeating the conditions here
            # stops a record that became held or paid in between from being
            # marked paid anyway.
            {"_id": {"$in": ids}, "status": "processed", "on_hold": {"$ne": True}},
            {"$set": {"status": "paid", "paid_at": when,
                      "published_at": datetime.now(timezone.utc).isoformat(),
                      "published_by": actor}},
        )
        published = res.modified_count
    return {**scored, "published": published, "paid_at": when,
            # kept for older callers that read this field
            "held_skipped": len(scored["skipped"]["held"])}


@router.get("/export/neft")
async def export_neft(period: str, confirm_reexport: bool = False,
                      current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")

    # RE-EXPORT GUARD. Downloading this sheet used to leave no trace at all, so
    # a second download and a second upload paid the whole company twice with
    # nothing anywhere to notice. Every export is now logged, and a repeat is
    # refused unless the caller says explicitly that it means to do it.
    #
    # Enforced here rather than in the browser so a script, a retry or a second
    # tab cannot walk around it.
    prior = await db.payroll_exports.find(
        {"period": period, "voided_at": {"$exists": False}}
    ).sort("exported_at", -1).to_list(50)
    if prior and not confirm_reexport:
        last = prior[0]
        raise HTTPException(
            status_code=409,
            detail=(
                f"The NEFT sheet for {period} has already been exported "
                f"{len(prior)} time(s). The last was on "
                f"{str(last.get('exported_at', ''))[:19].replace('T', ' ')} UTC by "
                f"{last.get('exported_by') or 'unknown'}, covering {last.get('row_count', '?')} "
                f"employee(s) and \u20b9{last.get('total_amount', 0):,.2f}.\n\n"
                "If that file was already sent to the bank, downloading and uploading "
                "it again pays everyone a SECOND time. Only continue if you are sure "
                "the earlier file was not submitted."
            ),
        )

    # The check above is a READ. Two downloads firing together both passed it and
    # both walked away with a full file — the precise "download twice, upload
    # twice, pay everyone twice" case the guard exists to prevent.
    #
    # Take an exclusive lock on the period for the duration of the build. The
    # filter matches only when the lock is free (or gone stale after a crash);
    # when it is held the filter matches nothing, the upsert tries to insert a
    # second row for the period, and the unique index in server.py turns that
    # into a DuplicateKeyError — which is the loser being refused.
    _lock_now = datetime.now(timezone.utc)
    _stale_before = (_lock_now - timedelta(minutes=5)).isoformat()
    # A token unique to THIS claim. Releasing on the period alone meant a slow
    # export whose lock had gone stale would, on finishing, clear the lock of
    # whoever had taken it over — and a third export could then walk in while
    # the second was still building.
    _lock_token = f"{_lock_now.isoformat()}|{uuid.uuid4().hex}"
    try:
        await db.payroll_export_state.find_one_and_update(
            {"period": period,
             "$or": [{"export_in_progress": {"$ne": True}},
                     {"export_claimed_at": {"$lt": _stale_before}}]},
            {"$set": {"period": period,
                      "export_in_progress": True,
                      "export_claimed_at": _lock_now.isoformat(),
                      "export_lock_token": _lock_token,
                      "export_claimed_by": current_user.get("employee_id") or current_user.get("username")}},
            upsert=True, return_document=ReturnDocument.AFTER,
        )
    except DuplicateKeyError:
        raise HTTPException(status_code=409, detail=(
            f"Another export of {period} is being prepared right now. Wait for it to finish, "
            f"then check whether that file was sent before downloading again — only one of "
            f"them is the file to give the bank."))

    # Everything from here to the release runs under the lock. The heaviest
    # database work in this route sits in this stretch, so a client abort or a
    # Mongo blip here used to leak the lock and jam the period for five minutes.
    try:
        all_records = await db.payroll_records.find({"period": period}).to_list(1000)

        # This file is the only thing that actually moves money, so every exclusion from
        # it is reported back in a header and surfaced by the UI. Four things keep a
        # record out, and NONE of them may be silent:
        #   1. on hold      -- resignation accepted, pending exit clearance
        #   2. still draft  -- nobody has reviewed the auto-calculated figure yet
        #   3. bank account not verified
        #   4. account number or IFSC missing -- see _bank() below
        # Nothing is written into the sheet itself; the bank parses it.
        emp_ids_all = [r.get("employee_id") for r in all_records if r.get("employee_id")]
        emp_docs = await db.employees.find(
            {"employee_id": {"$in": emp_ids_all}},
            {"employee_id": 1, "bank_details": 1, "first_name": 1, "last_name": 1, "_id": 0}
        ).to_list(None)
        # `or {}` and not .get("bank_details", {}): a document with the key present
        # and set to null returns None from the default form, and the AttributeError
        # is raised outside the try below — one such employee 500s the export and
        # NOBODY gets paid that month.
        bank_verified_ids = {
            e["employee_id"] for e in emp_docs
            if (e.get("bank_details") or {}).get("verified") is True
        }
        # Live account/IFSC/name, keyed by employee.
        #
        # The payroll record froze a COPY of the bank details when the month was
        # processed, but "is this account verified?" is read live from the employee
        # master. Those two disagree whenever HR fills in or corrects bank details
        # after processing: the employee is verified, so they pass the filter, while
        # the frozen copy is still empty — and the row reached the bank with no
        # account number at all.
        #
        # THE EMPLOYEE MASTER IS THE ONLY SOURCE OF PAYMENT INSTRUCTIONS. The frozen
        # copy is history, not an instruction, and is deliberately NOT used as a
        # fallback. Two reasons, both of which put money in the wrong account:
        #
        #   - Falling back FIELD BY FIELD splices a pair that never existed. A live
        #     account number with last month's ICICI IFSC becomes an intra-ICICI
        #     transfer with the IFSC cell blanked, so there is no routing data left
        #     to catch the mistake.
        #   - Clearing the account number on the master is how HR retracts a wrong
        #     account. Reviving the frozen copy pays the very account they removed.
        #
        # An incomplete master pair is therefore an exclusion, reported by name.
        live_bank = {}
        live_name = {}
        for e in emp_docs:
            bd = e.get("bank_details") or {}
            live_bank[e["employee_id"]] = (
                "".join(str(bd.get("account_number") or "").split()),
                str(bd.get("ifsc_code") or "").strip().upper(),
            )
            live_name[e["employee_id"]] = " ".join(
                f"{e.get('first_name') or ''} {e.get('last_name') or ''}".split())

        def _eid(r):
            return (r.get("employee_id") or "").strip()

        def _bank(r):
            """(account, ifsc) from the employee master, as a pair or not at all."""
            return live_bank.get(_eid(r), ("", ""))

        def _payee_name(r):
            """Live name, so the payee cannot belong to an older state than the account."""
            return live_name.get(_eid(r)) or r.get("employee_name", "")

        def clean_name(name: str) -> str:
            if not name:
                return ""
            # Allow letters and spaces only, uppercase, max 32 chars
            cleaned = "".join(ch for ch in name.upper() if ch.isalpha() or ch == " ")
            cleaned = " ".join(cleaned.split())
            return cleaned[:32]

        held_records = [r for r in all_records if r.get("on_hold")]
        rest = [r for r in all_records if not r.get("on_hold")]
        # `draft` means the figure came straight out of the auto-calculation and no human
        # has checked it. The app already refuses to mark a draft as *paid*; it must not
        # be wired to an actual bank transfer either. Opening the payslip and clicking
        # "Save Adjustments" moves a record to `processed` and clears this.
        draft_records = [r for r in rest if r.get("status") == "draft"]
        # An ALLOW-LIST, not "anything except draft". The old test let `cancelled`,
        # `reversed`, an empty string and a record with no status field at all into
        # the bank file, because none of them are literally "draft".
        reviewed = [r for r in rest if r.get("status") in _PAYABLE_STATUSES]
        badstatus_records = [r for r in rest
                             if r.get("status") != "draft" and r.get("status") not in _PAYABLE_STATUSES]

        # DUPLICATES — two records for one employee in one month means the salary
        # goes out twice. Deduplicate before anything else touches the list.
        #
        # Identical amounts are the normal shape of an accidental double-process, so
        # those are collapsed to one row and reported: withholding pay over a
        # clerical duplicate helps nobody. Records that DISAGREE on the amount are a
        # different matter — nobody but a human can say which is right — so those are
        # excluded and named.
        by_emp = {}
        for r in reviewed:
            by_emp.setdefault(_eid(r), []).append(r)

        def _net(r):
            try:
                return round(float(r.get("net_salary") or 0), 2)
            except (TypeError, ValueError):
                return None

        deduped, duplicate_records, conflicting_records = [], [], []
        for eid, group in by_emp.items():
            if len(group) == 1:
                deduped.append(group[0])
                continue
            if len({_net(r) for r in group}) == 1:
                deduped.append(group[0])
                duplicate_records.append(group[0])
            else:
                conflicting_records.extend(group)
        reviewed = deduped
        verified_records = [r for r in reviewed if _eid(r) in bank_verified_ids]
        # A blank account number in this file is money that does not arrive, so it is
        # a fourth exclusion rather than a blank cell the bank has to reject. IFSC is
        # required too: it is what identifies an ICICI account, and ICICI rows are the
        # ones where the IFSC cell is deliberately left empty further down.
        # A blank beneficiary NAME is the same failure as a blank account: the bank
        # rejects the row, and it used to be written out silently. Checked with the
        # same cleaner the sheet uses, so what is validated is what gets sent.
        def _payable(r):
            return all(_bank(r)) and bool(clean_name(_payee_name(r)))

        incomplete_records = [r for r in verified_records if not _payable(r)]
        payable_records = [r for r in verified_records if _payable(r)]

        # WITHHELD: a net at or below zero. Deductions can legitimately exceed a
        # month's salary (an advance being recovered), but a zero row wastes a bank
        # transaction and a NEGATIVE row instructs the bank to take money OUT of the
        # employee's account. Neither belongs in a salary file — the excess has to be
        # carried to the next month by a human, so this is reported, never sent.
        def _net(r):
            try:
                return round(float(r.get("net_salary") or 0), 2)
            except (TypeError, ValueError):
                return 0.0

        nonpositive_records = [r for r in payable_records if _net(r) <= 0]
        records = [r for r in payable_records if _net(r) > 0]

        # INCLUDED, but flagged: verified with no account-holder name from the bank.
        # A verification used to be recorded from a Perfios response that contained
        # no verification data at all, and those look exactly like this — verified
        # true, name blank. They are paid, not withheld, because a blank name can
        # also be a legitimate bank response; but they are worth re-verifying before
        # the money moves.
        unnamed_ids = {
            e["employee_id"] for e in emp_docs
            if (e.get("bank_details") or {}).get("verified") is True
            and not ((e.get("bank_details") or {}).get("verified_name") or "").strip()
        }
        unnamed_records = [r for r in records if _eid(r) in unnamed_ids]
        # Already marked paid, yet still in this file. Legitimate if the period is
        # published before the sheet is pulled — but it is also exactly what a
        # re-export looks like, so HR is told rather than left to notice.
        alreadypaid_records = [r for r in records if r.get("status") == "paid"]
        # INCLUDED, and flagged for a different reason: a human overrode the bank
        # check rather than Perfios confirming it. Legitimate — Perfios can refuse to
        # answer — but it must be visible on the file that moves money.
        manual_ids = {
            e["employee_id"] for e in emp_docs
            if (e.get("bank_details") or {}).get("verified_manually") is True
        }
        manual_records = [r for r in records if _eid(r) in manual_ids]

        # Reported over every non-held record, INCLUDING drafts, so a record blocked for
        # two reasons reports both at once. Reporting only the reviewed ones would mean
        # fixing the draft, re-downloading, and only then learning the bank is unverified.
        # This overlaps `draft_records` on purpose -- both problems are real and both
        # need fixing.
        unverified_records = [r for r in rest if _eid(r) not in bank_verified_ids]

        settings_doc = await db.app_settings.find_one({"key": "company"}) or {}
        # NEFT debit account is fixed by company policy — always use 019005008108
        debit_account = "019005008108"
        txn_type = (settings_doc.get("transaction_type") or "NFT").strip()
        short_code = (settings_doc.get("company_short_code") or "RMF0001").strip()

        # Period -> "Apr26" (locale-independent)
        _MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        try:
            y, m = period.split("-")
            month_short = _MONTHS[int(m) - 1]
            period_label = f"{month_short.upper()}{y[-2:]}"  # e.g. "APR26" (uppercase, 2-digit year)
        except Exception:
            period_label = period

        remark_full = f"Salary {period_label}"  # per-employee ID prepended inside the loop
        remark_suffix = remark_full

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"NEFT_{period}"
            headers = [
                "Transaction type \n(Within Bank (WIB)/\nNEFT (NFT)/\nRTGS (RTG)/\nIMPS (IFC))",
                "Amount (\u20b9)\n(Should not be more than 15 digit including decimals and paise)",
                "Debit Account no\nShould be exactly 12 digit",
                "IFSC (Always 11 character alphanumeric and 5th character always 0 (zero)) (For ICICI bank accounts keep it blank)",
                "Beneficiary Account No (Max length for other bank 34 character alphanumeric and for ICICI Bank 12 digit number )",
                "Beneficiary Name (Max length 32 Character) (No Special Character is allowed but Space is allowed)",
                "Remarks for Client\n(should not be more than 21 characters)",
                "Remarks for Beneficiary\n(should not be more than 30 characters)",
            ]
            ws.append(headers)
            # Header styling
            header_fill = PatternFill("solid", fgColor="1E2A47")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            thin = Side(border_style="thin", color="CCCCCC")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws.row_dimensions[1].height = 90
            widths = [22, 18, 18, 16, 26, 28, 22, 28]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

            for r in records:
                # `records` is already filtered: not held, not draft, bank verified.
                emp_id = _eid(r)
                net_amount = round(float(r.get("net_salary", 0) or 0), 2)
                beneficiary_acct, ifsc = _bank(r)
                name = clean_name(_payee_name(r))
                row_remark = f"{emp_id} {remark_suffix}"   # e.g. "RMF0001 Salary APR26"
                row_remark_client = row_remark[:21]
                row_remark_beneficiary = row_remark[:30]
                # ICICI Bank: set transaction type to WIB and leave IFSC blank
                is_icici = ifsc.startswith("ICIC")
                ws.append([
                    "WIB" if is_icici else txn_type,
                    net_amount,
                    debit_account,
                    "" if is_icici else ifsc,
                    beneficiary_acct,
                    name,
                    row_remark_client,
                    row_remark_beneficiary,
                ])
            # Body styling
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=8):
                for cell in row:
                    cell.alignment = Alignment(vertical="center", wrap_text=False)
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            buffer = io.BytesIO()
            # Lock all cells — enable sheet protection so no one can tamper
            from openpyxl.styles import Protection
            for row in ws.iter_rows():
                for cell in row:
                    cell.protection = Protection(locked=True)
            ws.protection.sheet   = True
            ws.protection.password = "RadhyaFinance"
            ws.protection.insertRows     = True
            ws.protection.insertColumns  = True
            ws.protection.deleteRows     = True
            ws.protection.deleteColumns  = True
            ws.protection.sort           = True
            ws.protection.autoFilter     = True
            wb.save(buffer)
            buffer.seek(0)
            filename = f"NEFT_{short_code}_{period}.xlsx"
            held_amount = round(sum(float(r.get("net_salary") or 0) for r in held_records), 2)

            def _ids(rows, limit=60):
                # Employee IDs, not just a count — so HR can act on the omission instead of
                # working out who is absent by reading a spreadsheet.
                #
                # Truncated by WHOLE ids, never mid-identifier, and the remainder is
                # stated rather than silently dropped: a half-written employee id is
                # worse than an honest "+12 more". The cap also keeps the six id
                # headers together well inside a proxy's default header buffer —
                # blow that and the whole download 502s with no diagnosis.
                ids = sorted(_eid(r) for r in rows if _eid(r))
                shown, extra = ids[:limit], len(ids) - limit
                return ",".join(shown) + (f",+{extra} more" if extra > 0 else "")

            _exposed = [
                "X-Payroll-Held-Count", "X-Payroll-Held-Amount", "X-Payroll-Held-Ids",
                "X-Payroll-Draft-Count", "X-Payroll-Draft-Ids",
                "X-Payroll-Unverified-Count", "X-Payroll-Unverified-Ids",
                "X-Payroll-Incomplete-Count", "X-Payroll-Incomplete-Ids",
                "X-Payroll-Unnamed-Count", "X-Payroll-Unnamed-Ids",
                "X-Payroll-Manual-Count", "X-Payroll-Manual-Ids",
                "X-Payroll-Duplicate-Count", "X-Payroll-Duplicate-Ids",
                "X-Payroll-Conflicting-Count", "X-Payroll-Conflicting-Ids",
                "X-Payroll-Badstatus-Count", "X-Payroll-Badstatus-Ids",
                "X-Payroll-Alreadypaid-Count", "X-Payroll-Alreadypaid-Ids",
                "X-Payroll-Nonpositive-Count", "X-Payroll-Nonpositive-Ids",
                "X-Payroll-Previous-Exports",
                "X-Payroll-Included-Count",
            ]
            response = Response(
                content=buffer.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                    "X-Payroll-Held-Count": str(len(held_records)),
                    "X-Payroll-Held-Amount": str(held_amount),
                    "X-Payroll-Held-Ids": _ids(held_records),
                    "X-Payroll-Draft-Count": str(len(draft_records)),
                    "X-Payroll-Draft-Ids": _ids(draft_records),
                    "X-Payroll-Unverified-Count": str(len(unverified_records)),
                    "X-Payroll-Unverified-Ids": _ids(unverified_records),
                    "X-Payroll-Incomplete-Count": str(len(incomplete_records)),
                    "X-Payroll-Incomplete-Ids": _ids(incomplete_records),
                    "X-Payroll-Unnamed-Count": str(len(unnamed_records)),
                    "X-Payroll-Unnamed-Ids": _ids(unnamed_records),
                    "X-Payroll-Manual-Count": str(len(manual_records)),
                    "X-Payroll-Manual-Ids": _ids(manual_records),
                    "X-Payroll-Duplicate-Count": str(len(duplicate_records)),
                    "X-Payroll-Duplicate-Ids": _ids(duplicate_records),
                    "X-Payroll-Conflicting-Count": str(len(conflicting_records)),
                    "X-Payroll-Conflicting-Ids": _ids(conflicting_records),
                    "X-Payroll-Badstatus-Count": str(len(badstatus_records)),
                    "X-Payroll-Badstatus-Ids": _ids(badstatus_records),
                    "X-Payroll-Alreadypaid-Count": str(len(alreadypaid_records)),
                    "X-Payroll-Alreadypaid-Ids": _ids(alreadypaid_records),
                    "X-Payroll-Nonpositive-Count": str(len(nonpositive_records)),
                    "X-Payroll-Nonpositive-Ids": _ids(nonpositive_records),
                    "X-Payroll-Previous-Exports": str(len(prior)),
                    "X-Payroll-Included-Count": str(len(records)),
                    # Browsers hide custom headers from JS unless they are exposed.
                    "Access-Control-Expose-Headers": ", ".join(_exposed),
                },
            )

            # Logged only once the response object exists — header construction can
            # itself fail, and a marker written before that would claim a bank file
            # was delivered when the download 500'd and produced nothing.
            #
            # An export with no rows sent nothing to anyone either. Logging that
            # would tell `Mark All Paid` a bank file exists for the period, and would
            # 409 the next genuine export as a re-export.
            if records:
                await db.payroll_exports.insert_one({
                    "period": period,
                    "exported_at": datetime.now(timezone.utc).isoformat(),
                    "exported_by": current_user.get("employee_id") or current_user.get("username"),
                    "row_count": len(records),
                    "total_amount": round(sum(float(r.get("net_salary") or 0) for r in records), 2),
                    "employee_ids": sorted(_eid(r) for r in records if _eid(r)),
                    # Per-employee amounts as SENT. Without these, "was this employee
                    # in the file?" is answerable but "does the record still say what
                    # the bank was told?" is not — and a recalculation after export
                    # makes exactly that go out of step.
                    "rows": [{"employee_id": _eid(r),
                              "net_salary": round(float(r.get("net_salary") or 0), 2)}
                             for r in records if _eid(r)],
                    "was_reexport": bool(prior),
                })
            return response
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))
    finally:
        # Always released, including on a 500 or a client disconnect — otherwise
        # one failed download blocks the period until the lock goes stale. The
        # token match means this can only ever release ITS OWN claim.
        await db.payroll_export_state.update_one(
            {"period": period, "export_lock_token": _lock_token},
            {"$set": {"export_in_progress": False}})


@router.get("/duplicates")
async def payroll_duplicates(current_user: dict = Depends(get_current_user)):
    """Employees holding more than one payroll record for the same month.

    Read-only. Two records for one month is a salary paid twice, and until the
    unique (employee_id, period) index can be built — it cannot be, while
    duplicates exist — this is how you find and clear them.

    Also reports whether the amounts agree: matching amounts are a clerical
    double-process and the NEFT export collapses them to one row; disagreeing
    amounts need a human to decide which is right, and the export withholds them.
    """
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")
    pipeline = [
        {"$group": {
            "_id": {"employee_id": "$employee_id", "period": "$period"},
            "count": {"$sum": 1},
            "nets": {"$push": "$net_salary"},
            "statuses": {"$push": "$status"},
            "ids": {"$push": "$_id"},
        }},
        {"$match": {"count": {"$gt": 1}}},
        {"$sort": {"_id.period": -1, "_id.employee_id": 1}},
    ]
    rows = await db.payroll_records.aggregate(pipeline).to_list(2000)
    out = []
    for r in rows:
        nets = [round(float(n or 0), 2) for n in r.get("nets", [])]
        out.append({
            "employee_id": r["_id"]["employee_id"],
            "period": r["_id"]["period"],
            "count": r["count"],
            "net_salaries": nets,
            "amounts_agree": len(set(nets)) == 1,
            "statuses": r.get("statuses", []),
            "record_ids": [str(i) for i in r.get("ids", [])],
            "duplicate_total": round(sum(nets) - (nets[0] if nets else 0), 2),
        })
    return {
        "duplicate_groups": len(out),
        "extra_records": sum(r["count"] - 1 for r in out),
        "amount_at_risk": round(sum(r["duplicate_total"] for r in out), 2),
        "rows": out,
    }


async def _score_lop_recalc(period: str) -> dict:
    """Re-run the LOP calculation for a period and report what WOULD change.

    Pure scoring — writes nothing. Both the preview and the apply call this, so
    the numbers HR confirms are by construction the numbers that get written.

    Skips, and says so rather than silently dropping:
      * `paid` records — the money is gone; rewriting the record would only make
        it disagree with the bank.
      * records whose LOP was set by hand (`lop_source == "manual"`) — HR made a
        deliberate decision and this must not quietly undo it.
    Held records ARE rescored: they are unpaid and still owed.
    """
    year, month = _parse_period(period)
    days_in_month = _cal.monthrange(year, month)[1]
    records = await db.payroll_records.find({"period": period}).to_list(4000)

    rows, skipped_paid, skipped_manual, negative_clamped = [], [], [], []
    for r in records:
        eid = r.get("employee_id")
        name = r.get("employee_name", "") or eid
        if r.get("status") == "paid":
            skipped_paid.append({"employee_id": eid, "employee_name": name})
            continue
        if (r.get("lop_source") or "auto") == "manual":
            skipped_manual.append({"employee_id": eid, "employee_name": name,
                                   "lop_days": float(r.get("lop_days") or 0)})
            continue
        emp = await db.employees.find_one({"employee_id": eid})
        if not emp:
            continue

        joining = emp.get("joining_date") or f"{period}-01"
        lwd = emp.get("last_working_day") or None
        lop_new, non_emp_new = await calculate_lop_days(eid, year, month, joining, lwd)
        # Price against the salary the month was RUN on. Falling back to the live
        # master is only for records processed before salary_basis existed; those
        # are flagged so HR can see which rows are being re-rated at today's pay.
        basis = r.get("salary_basis")
        priced_on = "run" if basis else "current"
        emp_for_pricing = {"designation": r.get("designation") or emp.get("designation"),
                           "salary": basis} if basis else emp
        comp = calc_payroll_components(emp_for_pricing, days_in_month, lop_new, non_emp_new)

        tds = float(r.get("tds") or 0)
        other_ded = float(r.get("other_deductions") or 0)
        other_add = float(r.get("other_additions") or 0)
        net_new = round(comp["gross_payable"] - comp["epf_employee"]
                        - comp["esic_employee"] - tds - other_ded + other_add)
        # The same rule PUT enforces: a negative salary is never written. More
        # LOP on a record already carrying a large deduction can reach one, and
        # this path writes net_salary directly.
        if net_new < 0:
            net_new = 0
            negative_clamped.append(eid)
        net_old = float(r.get("net_salary") or 0)
        lop_old = float(r.get("lop_days") or 0)
        non_emp_old = float(r.get("non_employed_days") or 0)

        if lop_new == lop_old and non_emp_new == non_emp_old and net_new == net_old:
            continue

        rows.append({
            "record_id": str(r.get("_id")),
            "employee_id": eid,
            "employee_name": name,
            "status": r.get("status"),
            "on_hold": bool(r.get("on_hold")),
            "lop_before": lop_old, "lop_after": lop_new,
            "non_employed_before": non_emp_old, "non_employed_after": non_emp_new,
            "net_before": net_old, "net_after": net_new,
            "delta": round(net_new - net_old, 2),
            "priced_on": priced_on,
            "_update": {**comp, "net_salary": net_new,
                        "total_deductions": round(comp["epf_employee"] + comp["esic_employee"]
                                                  + tds + other_ded, 2)},
        })

    # Has money already gone out for this period? If so a recalculation cannot
    # recover anything — it only reveals what was overpaid or underpaid.
    export = await db.payroll_exports.find_one(
        {"period": period, "voided_at": {"$exists": False}}, sort=[("exported_at", -1)])
    overpaid = round(sum(-x["delta"] for x in rows if x["delta"] < 0), 2)
    underpaid = round(sum(x["delta"] for x in rows if x["delta"] > 0), 2)

    return {
        "period": period,
        "already_exported": bool(export),
        "exported_at": (export or {}).get("exported_at"),
        "changed": len(rows),
        "skipped_paid": skipped_paid,
        "skipped_manual": skipped_manual,
        "net_delta": round(sum(x["delta"] for x in rows), 2),
        "overpaid_total": overpaid,
        "underpaid_total": underpaid,
        "negative_clamped": negative_clamped,
        "rows": rows,
    }


def _strip_internal(scored: dict) -> dict:
    """Drop the write payload before the result goes over the wire."""
    return {**scored, "rows": [{k: v for k, v in r.items() if k != "_update"}
                               for r in scored["rows"]]}


@router.get("/recalculate-lop/preview")
async def preview_lop_recalc(period: str, current_user: dict = Depends(get_current_user)):
    """What Recalculate LOP would change. Read-only — nothing is written."""
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")
    return _strip_internal(await _score_lop_recalc(period))


@router.post("/recalculate-lop")
async def apply_lop_recalc(period: str, confirm_after_export: bool = False,
                           current_user: dict = Depends(get_current_user)):
    """Apply the recalculation to every unpaid, non-manual record in the period.

    Refuses once a NEFT file has been exported unless explicitly confirmed. At
    that point the money is with the bank, so rewriting the records does not
    recover a rupee — it makes the payroll disagree with what was actually sent.
    That can be the right call, but it has to be a decision, not a side effect.
    """
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")
    scored = await _score_lop_recalc(period)

    if scored["already_exported"] and not confirm_after_export:
        raise HTTPException(status_code=409, detail=(
            f"A NEFT file for {period} was already exported on "
            f"{str(scored['exported_at'])[:10]}. Recalculating now will not recover "
            f"money already sent — it will only change the records to disagree with it "
            f"(₹{scored['overpaid_total']:,.2f} overpaid, ₹{scored['underpaid_total']:,.2f} "
            f"underpaid across {scored['changed']} employee(s)). "
            f"Re-send with confirm_after_export=true to record it anyway."))

    now = datetime.now(timezone.utc).isoformat()
    actor = current_user.get("employee_id") or current_user.get("username")
    applied = 0
    for row in scored["rows"]:
        # Conditions repeated from the scoring: between preview and apply a
        # record can be paid or hand-edited, and it must not be clobbered.
        res = await db.payroll_records.update_one(
            {"_id": _oid(row["record_id"]), "status": {"$ne": "paid"},
             "lop_source": {"$ne": "manual"}},
            {"$set": {**row["_update"], "lop_source": "auto",
                      "lop_recalculated_at": now, "lop_recalculated_by": actor}},
        )
        applied += res.modified_count
    return {**_strip_internal(scored), "applied": applied}


@router.get("/epf-shortfall")
async def payroll_epf_shortfall(period: str = None, current_user: dict = Depends(get_current_user)):
    """Existing payroll records whose EPF was computed under the old rule.

    Read-only — nothing is recalculated or written. Every record is re-scored
    against the correct rule (12% of the basic actually earned, capped at 1,800
    on the contribution) using the record's OWN stored basic, not the employee
    master, so a salary revision since the run cannot distort the comparison.

    Only records that are genuinely wrong appear. Full-attendance months are
    unaffected, and so is anyone whose master held the true uncapped 12% — for
    them the old formula already produced the right number.

    The error is one-directional: the old rule capped an already-capped figure,
    so it always UNDER-deducted. Every row here is money short-remitted to EPFO
    on both the employee and the employer side.
    """
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")

    q = {"period": period} if period else {}
    records = await db.payroll_records.find(q).sort("period", -1).to_list(20000)

    rows = []
    for r in records:
        stored = round(float(r.get("epf_employee") or 0), 2)
        # 0 means the employee was exempt at process time — not a shortfall.
        if stored <= 0:
            continue
        basic_paid = float(r.get("basic") or 0)
        correct = round(min(basic_paid * 0.12, 1800), 2)
        if correct == stored:
            continue
        rows.append({
            "record_id": str(r.get("_id")),
            "employee_id": r.get("employee_id"),
            "employee_name": r.get("employee_name", ""),
            "period": r.get("period"),
            "status": r.get("status"),
            "on_hold": bool(r.get("on_hold")),
            "lop_days": r.get("lop_days", 0),
            "basic_paid": round(basic_paid, 2),
            "epf_charged": stored,
            "epf_correct": correct,
            "shortfall_employee": round(correct - stored, 2),
            # employer mirrors the employee side, so the remittance gap is double
            "shortfall_total": round((correct - stored) * 2, 2),
        })

    by_period: dict = {}
    for row in rows:
        p = by_period.setdefault(row["period"], {"period": row["period"], "records": 0, "shortfall_total": 0.0})
        p["records"] += 1
        p["shortfall_total"] = round(p["shortfall_total"] + row["shortfall_total"], 2)

    return {
        "affected_records": len(rows),
        "shortfall_employee": round(sum(r["shortfall_employee"] for r in rows), 2),
        "shortfall_total": round(sum(r["shortfall_total"] for r in rows), 2),
        "unpaid_records": sum(1 for r in rows if r["status"] != "paid"),
        "paid_records": sum(1 for r in rows if r["status"] == "paid"),
        "by_period": sorted(by_period.values(), key=lambda x: x["period"], reverse=True),
        "rows": rows,
    }


@router.get("/export-history")
async def payroll_export_history(period: str = None, current_user: dict = Depends(get_current_user)):
    """Every NEFT download, so "was this month already sent?" has an answer."""
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")
    q = {"period": period} if period else {}
    rows = await db.payroll_exports.find(q, {"_id": 0, "employee_ids": 0}).sort("exported_at", -1).to_list(200)
    return rows


@router.get("/export/salary-register")
async def export_salary_register(period: str, current_user: dict = Depends(get_current_user)):
    """Master Monthly Salary Register — comprehensive multi-column Excel summary
    of every processed payroll record for the period (YYYY-MM).
    Includes: Employee details, Paid/Leave days, Earnings breakup, Gross,
    Deductions breakup, Net Salary, Employer Contributions, Monthly CTC,
    and Bank details. A totals row is appended at the bottom."""
    if current_user.get("role") not in ["hr_admin", "management"]:
        raise HTTPException(status_code=403, detail="Access denied")

    records = await db.payroll_records.find({"period": period}).sort("employee_id", 1).to_list(2000)
    if not records:
        raise HTTPException(status_code=404, detail=f"No payroll records found for period {period}.")

    # Enrich with employee master data (joining date, PAN, UAN, ESI)
    emp_ids = [r.get("employee_id") for r in records if r.get("employee_id")]
    employees = await db.employees.find(
        {"employee_id": {"$in": emp_ids}},
        {"_id": 0}
    ).to_list(2000)
    emp_map = {e["employee_id"]: e for e in employees}

    _MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    try:
        y, m = period.split("-")
        period_label = f"{_MONTHS[int(m)-1]} {y}"
    except Exception:
        period_label = period

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Salary Register {period}"

    # Styling helpers
    navy_fill = PatternFill("solid", fgColor="1E2A47")
    orange_fill = PatternFill("solid", fgColor="E85B1E")
    grey_fill = PatternFill("solid", fgColor="F1F5F9")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    section_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, color="1E2A47", size=14)
    subtitle_font = Font(italic=True, color="64748B", size=9)
    total_font = Font(bold=True, color="1E2A47", size=10)
    total_fill = PatternFill("solid", fgColor="FEF3E2")
    thin = Side(border_style="thin", color="CCCCCC")
    medium = Side(border_style="medium", color="1E2A47")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Column layout (grouped):
    # Identity (1-7): Sr, Emp ID, Name, Designation, Department, DOJ, Status
    # Attendance (8-11): Working Days, Paid Days, LOP, Not Employed
    # Earnings (12-17): Basic, HRA, Special, Canteen, Conveyance, Other Add
    # Gross (18): Gross Salary
    # Deductions (19-23): EPF Emp, ESIC Emp, TDS, Other Ded, Total Ded
    # Net (24): Net Salary
    # Employer (25-27): EPF Empr, ESIC Empr, Gratuity
    # CTC (28): Monthly CTC
    # Statutory (29-31): PAN, UAN, ESI No
    # Bank (32-34): Bank Name, A/c No, IFSC
    #
    # "Not Employed" is separate from LOP so the row adds up: Working − LOP −
    # Not Employed = Paid. A joiner shows 26 not-employed and 0 LOP, never
    # 26 LOP — the register is a compliance record, and the difference between
    # "not on the rolls" and "absent without leave" is not cosmetic.
    group_headers = [
        ("Identity", 1, 7, navy_fill),
        ("Attendance", 8, 11, orange_fill),
        ("Earnings (₹)", 12, 17, navy_fill),
        ("Gross", 18, 18, orange_fill),
        ("Deductions (₹)", 19, 23, navy_fill),
        ("Net Pay", 24, 24, orange_fill),
        ("Employer Cost (₹)", 25, 27, navy_fill),
        ("CTC", 28, 28, orange_fill),
        ("Statutory IDs", 29, 31, navy_fill),
        ("Bank Details", 32, 34, orange_fill),
        ("Payment", 35, 35, navy_fill),
    ]
    col_headers = [
        "Sr", "Employee ID", "Name", "Designation", "Department", "Joining Date", "Status",
        "Working Days", "Paid Days", "LOP Days", "Not Employed",
        "Basic", "HRA", "Special Allow.", "Canteen", "Conveyance", "Other Add.",
        "Gross Salary",
        "EPF (Emp)", "ESIC (Emp)", "TDS", "Other Ded.", "Total Ded.",
        "Net Salary",
        "EPF (Empr)", "ESIC (Empr)", "Gratuity",
        "Monthly CTC",
        "PAN", "UAN", "ESI No.",
        "Bank Name", "Account No.", "IFSC",
        "Payment Status",
    ]
    n_cols = len(col_headers)  # 35

    # Title rows
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(row=1, column=1, value=f"RADHYA MICRO FINANCE PRIVATE LIMITED — Monthly Salary Register").font = title_font
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    held_count = sum(1 for r in records if r.get("on_hold"))
    held_note = f"  |  {held_count} salary on hold" if held_count == 1 else (
        f"  |  {held_count} salaries on hold" if held_count else "")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.cell(row=2, column=1, value=f"Period: {period_label}  |  Generated: {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M UTC')}  |  {len(records)} employees{held_note}").font = subtitle_font
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # Group header row (row 3)
    for label, start, end, fill in group_headers:
        if start == end:
            cell = ws.cell(row=3, column=start, value=label)
        else:
            ws.merge_cells(start_row=3, start_column=start, end_row=3, end_column=end)
            cell = ws.cell(row=3, column=start, value=label)
        cell.fill = fill
        cell.font = section_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[3].height = 20

    # Column header row (row 4)
    for i, h in enumerate(col_headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = grey_fill
        c.font = Font(bold=True, color="1E2A47", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[4].height = 30

    # Totals accumulator (money columns + the two day columns)
    money_cols = [12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28]
    DAY_COLS = {10: "lop", 11: "non_employed"}   # formatted 0.## and totalled
    totals = {c: 0.0 for c in money_cols}
    day_totals = {c: 0.0 for c in DAY_COLS}

    # Data rows start at row 5
    r_idx = 5
    for sr, rec in enumerate(records, 1):
        emp = emp_map.get(rec.get("employee_id"), {})
        working = rec.get("working_days", 26) or 26
        paid = rec.get("present_days", working) or working
        non_emp = float(rec.get("non_employed_days") or 0)
        # Prefer stored lop_days (set explicitly by HR — supports half days);
        # fall back to derived value if missing. The fallback nets off
        # non-employed days so a joiner is never back-filled as absent.
        if rec.get("lop_days") is not None:
            lop = float(rec.get("lop_days"))
        else:
            lop = max(0.0, float(working) - float(paid) - non_emp)
        gross = float(rec.get("gross_payable") or rec.get("gross_salary") or 0)
        epf_e = float(rec.get("epf_employee") or 0)
        esic_e = float(rec.get("esic_employee") or 0)
        tds = float(rec.get("tds") or 0)
        other_d = float(rec.get("other_deductions") or 0)
        other_a = float(rec.get("other_additions") or 0)
        total_ded = round(epf_e + esic_e + tds + other_d, 2)
        net = float(rec.get("net_salary") or 0)
        epf_r = float(rec.get("epf_employer") or 0)
        esic_r = float(rec.get("esic_employer") or 0)
        grat = float(rec.get("gratuity_monthly") or 0)
        ctc = float(rec.get("ctc_monthly") or 0)

        bank = emp.get("bank_details", {}) or {}
        # Held rows stay in the register — this is the compliance record of what was
        # earned, and totals below are the full liability. The column is what tells
        # you the money didn't actually go out.
        if rec.get("on_hold"):
            pay_status = "HELD — ready to release" if rec.get("hold_eligible_at") else "HELD — exit in progress"
        elif rec.get("released_at"):
            pay_status = "Released"
        else:
            pay_status = (rec.get("status") or "").replace("_", " ").title()

        row_values = [
            sr,
            rec.get("employee_id", ""),
            rec.get("employee_name", "").strip() or f"{emp.get('first_name','')} {emp.get('last_name','')}".strip(),
            rec.get("designation", "") or emp.get("designation", ""),
            rec.get("department", "") or emp.get("department", ""),
            emp.get("joining_date", ""),
            (emp.get("status", "") or "").replace("_", " ").title(),
            working, paid, lop, non_emp,
            float(rec.get("basic") or 0),
            float(rec.get("hra") or 0),
            float(rec.get("special_allowance") or 0),
            float(rec.get("canteen_allowance") or 0),
            float(rec.get("conveyance_allowance") or 0),
            other_a,
            gross,
            epf_e, esic_e, tds, other_d, total_ded,
            net,
            epf_r, esic_r, grat,
            ctc,
            emp.get("pan_number", "") or "",
            emp.get("uan_number", "") or "",
            emp.get("esi_number", "") or "",
            bank.get("bank_name", "") or rec.get("bank_name", ""),
            bank.get("account_number", "") or rec.get("bank_account", ""),
            bank.get("ifsc_code", "") or rec.get("ifsc_code", ""),
            pay_status,
        ]

        for col_idx, val in enumerate(row_values, 1):
            c = ws.cell(row=r_idx, column=col_idx, value=val)
            c.border = border
            c.alignment = Alignment(vertical="center", horizontal="right" if col_idx in money_cols else "left")
            if col_idx in money_cols:
                c.number_format = '#,##0'
                totals[col_idx] = round(totals.get(col_idx, 0) + float(val or 0), 2)
            elif col_idx in DAY_COLS:  # day counts — show 0.5/1.5 when fractional
                c.number_format = '0.##'
                c.alignment = Alignment(vertical="center", horizontal="right")
                day_totals[col_idx] += float(val or 0)
        r_idx += 1

    # Totals row
    totals_row = r_idx
    for col_idx in range(1, n_cols + 1):
        c = ws.cell(row=totals_row, column=col_idx)
        c.fill = total_fill
        c.border = Border(top=medium, bottom=medium, left=thin, right=thin)
        c.font = total_font
        if col_idx == 1:
            c.value = "TOTAL"
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx in DAY_COLS:
            c.value = round(day_totals[col_idx], 2)
            c.number_format = '0.##'
            c.alignment = Alignment(horizontal="right", vertical="center")
        elif col_idx in totals:
            c.value = round(totals[col_idx], 2)
            c.number_format = '#,##0'
            c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[totals_row].height = 22

    # Column widths
    widths = [
        4, 12, 26, 20, 18, 12, 14,         # Identity
        9, 9, 7, 13,                        # Attendance
        11, 10, 13, 11, 12, 11,             # Earnings
        12,                                 # Gross
        11, 11, 10, 11, 11,                 # Deductions
        12,                                 # Net
        11, 11, 10,                         # Employer
        13,                                 # CTC
        13, 14, 14,                         # Statutory IDs
        18, 18, 13,                         # Bank
        24,                                 # Payment Status
    ]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes at data start
    ws.freeze_panes = "C5"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"Salary_Register_{period}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{record_id}/payslip/pdf")
async def download_payslip_pdf(record_id: str, current_user: dict = Depends(get_current_user)):
    record = await db.payroll_records.find_one({"_id": _oid(record_id)})
    if not record:
        raise HTTPException(status_code=404, detail="Payroll record not found")
    # Permission: HR/management can download any; everyone else (managers, employee, field_agent) only their own
    if current_user.get("role") not in ["hr_admin", "management"]:
        if record.get("employee_id") != current_user.get("employee_id"):
            raise HTTPException(status_code=403, detail="Access denied")
        # Gate: only after processing AND month-end
        if not _is_payslip_visible_to_employee(record):
            raise HTTPException(
                status_code=403,
                detail="Your payslip for this period is not available yet. It will be released after HR processes the payroll and the month ends.",
            )
    employee = await db.employees.find_one({"employee_id": record.get("employee_id")})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    # Remove MongoDB _id before passing to PDF builder
    record.pop("_id", None)
    employee.pop("_id", None)
    try:
        from services.payslip_pdf import build_payslip_pdf
        pdf_bytes = build_payslip_pdf(record, employee)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {str(e)}")
    emp_name = f"{employee.get('first_name','')}_{employee.get('last_name','')}".strip("_")
    period   = record.get("period", "unknown")
    filename = f"Payslip_{emp_name}_{period}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{record_id}")
async def get_payslip(record_id: str, current_user: dict = Depends(get_current_user)):
    record = await db.payroll_records.find_one({"_id": _oid(record_id)})
    if not record:
        raise HTTPException(status_code=404, detail="Not found")
    role = current_user.get("role")
    if role not in ["hr_admin", "management"]:
        if record.get("employee_id") != current_user.get("employee_id"):
            raise HTTPException(status_code=403, detail="Access denied")
        if not _is_payslip_visible_to_employee(record):
            raise HTTPException(status_code=403, detail="Payslip not yet released.")
    return pay_to_dict(record)
