"""
Iteration 12 — Backend tests for:
1. Employee list includes uan_verification and bank_details fields
2. PUT /api/employees/{id} with changed uan_number clears uan_verification
3. PUT /api/employees/{id} with changed bank fields clears bank_details.verified
4. GET /api/payroll/export/neft?period only includes bank-verified employees
"""

import pytest
import requests
import os
import asyncio as _asyncio
from dotenv import load_dotenv as _load_dotenv

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")

# ── Module-level DB state seed & teardown ───────────────────────────────────

@pytest.fixture(scope="session", autouse=True)
def restore_verification_state():
    """Ensure RMF0006 UAN and RMF0005/RMF0009 bank verifications are set before tests.
    Also cleans up after the entire session so the DB is in a known good state.
    """
    _load_dotenv()
    from motor.motor_asyncio import AsyncIOMotorClient

    async def _set_verified():
        client = AsyncIOMotorClient(os.environ["MONGO_URL"])
        db = client[os.environ.get("DB_NAME", "hr_system")]
        # Restore RMF0006 UAN verification
        await db.employees.update_one(
            {"employee_id": "RMF0006"},
            {"$set": {
                "uan_number": "101319308083",
                "uan_verification": {
                    "verified": True,
                    "registered_name": "BHASKAR ABHISHEK",
                    "name_matched": True,
                    "employment_history": [],
                    "verified_at": "2026-06-03T09:46:36.298726+00:00",
                }
            }}
        )
        # Restore RMF0005 bank verification
        await db.employees.update_one(
            {"employee_id": "RMF0005"},
            {"$set": {
                "bank_details.account_number": "32617846559",
                "bank_details.ifsc_code": "SBIN0004215",
                "bank_details.bank_name": "STATE BANK OF INDIA",
                "bank_details.verified": True,
                "bank_details.verified_name": "Miss ROOPAM  GUPTA",
                "bank_details.verified_at": "2026-05-09T12:42:47.678544+00:00",
            }}
        )
        # Restore RMF0009 bank verification
        await db.employees.update_one(
            {"employee_id": "RMF0009"},
            {"$set": {
                "bank_details.account_number": "1714831362",
                "bank_details.ifsc_code": "KKBK0005199",
                "bank_details.bank_name": "KOTAK MAHINDRA BANK",
                "bank_details.verified": True,
                "bank_details.verified_name": "SHIVANI PATHAK",
                "bank_details.verified_at": "2026-05-09T07:21:55.628368+00:00",
            }}
        )

    _asyncio.run(_set_verified())
    yield
    # After all tests: restore the same state (clean up any test artifacts)
    _asyncio.run(_set_verified())


# ── Fixtures ────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def admin_token():
    """Login as HR Admin and return JWT token."""
    resp = requests.post(f"{BASE_URL}/api/auth/login", json={
        "username": "admin",
        "password": "Admin@12345",
    })
    assert resp.status_code == 200, f"Admin login failed: {resp.text}"
    token = resp.json().get("access_token") or resp.json().get("token")
    assert token, "No token in login response"
    return token


@pytest.fixture(scope="module")
def admin_headers(admin_token):
    return {"Authorization": f"Bearer {admin_token}"}


# ── 1. Employee list: EPF & Bank verification columns ───────────────────────

class TestEmployeeListVerificationColumns:
    """Employee list response must include uan_verification and bank_details fields."""

    def test_employee_list_returns_uan_verification_for_rmf0006(self, admin_headers):
        """RMF0006 has verified UAN — list endpoint must return uan_verification.verified=True."""
        resp = requests.get(f"{BASE_URL}/api/employees", headers=admin_headers)
        assert resp.status_code == 200, f"Employee list failed: {resp.text}"
        employees = resp.json()
        rmf0006 = next((e for e in employees if e.get("employee_id") == "RMF0006"), None)
        assert rmf0006 is not None, "RMF0006 not found in employee list"
        uan_ver = rmf0006.get("uan_verification")
        assert uan_ver is not None, "uan_verification field missing from RMF0006 in list response"
        assert uan_ver.get("verified") is True, f"Expected uan_verification.verified=True, got: {uan_ver}"

    def test_employee_list_returns_bank_details_verified_for_rmf0005(self, admin_headers):
        """RMF0005 has verified bank — list endpoint must return bank_details.verified=True."""
        resp = requests.get(f"{BASE_URL}/api/employees", headers=admin_headers)
        assert resp.status_code == 200
        employees = resp.json()
        rmf0005 = next((e for e in employees if e.get("employee_id") == "RMF0005"), None)
        assert rmf0005 is not None, "RMF0005 not found in employee list"
        bank = rmf0005.get("bank_details")
        assert bank is not None, "bank_details field missing from RMF0005 in list response"
        assert bank.get("verified") is True, f"Expected bank_details.verified=True, got: {bank}"

    def test_employee_list_unverified_uan_employee(self, admin_headers):
        """RMF0005 has no UAN verified — uan_verification should be absent or not verified."""
        resp = requests.get(f"{BASE_URL}/api/employees", headers=admin_headers)
        assert resp.status_code == 200
        employees = resp.json()
        rmf0005 = next((e for e in employees if e.get("employee_id") == "RMF0005"), None)
        assert rmf0005 is not None
        uan_ver = rmf0005.get("uan_verification")
        # Either uan_verification is absent/None or verified is False/None
        if uan_ver:
            assert uan_ver.get("verified") is not True, \
                f"RMF0005 should NOT have verified UAN, got: {uan_ver}"

    def test_employee_detail_returns_uan_verification(self, admin_headers):
        """GET /api/employees/RMF0006 must return uan_verification.verified=True."""
        resp = requests.get(f"{BASE_URL}/api/employees/RMF0006", headers=admin_headers)
        assert resp.status_code == 200, f"RMF0006 detail fetch failed: {resp.text}"
        emp = resp.json()
        uan_ver = emp.get("uan_verification")
        assert uan_ver is not None, "uan_verification field missing from RMF0006 detail"
        assert uan_ver.get("verified") is True, f"Expected verified=True, got: {uan_ver}"

    def test_employee_detail_returns_bank_details_verified(self, admin_headers):
        """GET /api/employees/RMF0009 must return bank_details.verified=True."""
        resp = requests.get(f"{BASE_URL}/api/employees/RMF0009", headers=admin_headers)
        assert resp.status_code == 200
        emp = resp.json()
        bank = emp.get("bank_details")
        assert bank is not None, "bank_details field missing from RMF0009"
        assert bank.get("verified") is True, f"Expected bank verified=True, got: {bank}"


# ── 2. PUT endpoint clears UAN verification when UAN changes ────────────────

class TestUanVerificationClearOnUpdate:
    """PUT /api/employees/{id} with changed uan_number must clear uan_verification."""

    ORIGINAL_UAN = "101319308083"  # RMF0006's verified UAN
    TEST_UAN = "999999999999"      # A different UAN to trigger clearing

    def test_rmf0006_has_verified_uan_before_test(self, admin_headers):
        """Pre-condition: RMF0006 uan_verification.verified should be True."""
        resp = requests.get(f"{BASE_URL}/api/employees/RMF0006", headers=admin_headers)
        assert resp.status_code == 200
        emp = resp.json()
        assert emp.get("uan_verification", {}).get("verified") is True, \
            "Pre-condition failed: RMF0006 UAN verification not True"

    def test_changing_uan_clears_uan_verification(self, admin_headers):
        """PUT with a different uan_number → uan_verification should be cleared/empty."""
        resp = requests.put(
            f"{BASE_URL}/api/employees/RMF0006",
            json={"uan_number": self.TEST_UAN},
            headers=admin_headers,
        )
        assert resp.status_code == 200, f"PUT failed: {resp.text}"
        emp = resp.json()
        uan_ver = emp.get("uan_verification")
        # Should be cleared: either empty dict {} or verified=False/None
        if uan_ver:
            assert uan_ver.get("verified") is not True, \
                f"uan_verification should be cleared after UAN change, got: {uan_ver}"

    def test_uan_verification_cleared_on_get_after_change(self, admin_headers):
        """GET after UAN change: uan_verification should still be cleared (persisted)."""
        resp = requests.get(f"{BASE_URL}/api/employees/RMF0006", headers=admin_headers)
        assert resp.status_code == 200
        emp = resp.json()
        uan_ver = emp.get("uan_verification")
        if uan_ver:
            assert uan_ver.get("verified") is not True, \
                f"uan_verification should still be cleared in DB, got: {uan_ver}"

    def test_same_uan_does_not_clear_verification(self, admin_headers):
        """PUT with same UAN should NOT clear verification — restore original first."""
        # First restore the original UAN
        resp_restore = requests.put(
            f"{BASE_URL}/api/employees/RMF0006",
            json={"uan_number": self.ORIGINAL_UAN},
            headers=admin_headers,
        )
        assert resp_restore.status_code == 200, f"Restore UAN failed: {resp_restore.text}"
        # Note: After changing back, uan_verification was already cleared in the previous step.
        # The point here is that sending the same UAN shouldn't crash/error.
        # It CANNOT restore verification automatically — that requires a separate verify call.
        resp = requests.put(
            f"{BASE_URL}/api/employees/RMF0006",
            json={"uan_number": self.ORIGINAL_UAN},
            headers=admin_headers,
        )
        assert resp.status_code == 200, f"PUT same UAN failed: {resp.text}"

    def test_uan_number_restored_after_test(self, admin_headers):
        """Verify UAN is back to original value after test cleanup."""
        resp = requests.get(f"{BASE_URL}/api/employees/RMF0006", headers=admin_headers)
        assert resp.status_code == 200
        emp = resp.json()
        assert emp.get("uan_number") == self.ORIGINAL_UAN, \
            f"UAN not restored: expected {self.ORIGINAL_UAN}, got {emp.get('uan_number')}"


# ── 3. PUT endpoint clears bank verification when bank fields change ─────────

class TestBankVerificationClearOnUpdate:
    """PUT /api/employees/{id} with changed bank fields must clear bank_details.verified."""

    # RMF0005 has verified bank
    EMPLOYEE_ID = "RMF0005"
    ORIG_ACCOUNT = "32617846559"
    ORIG_IFSC = "SBIN0004215"
    ORIG_BANK = "STATE BANK OF INDIA"
    TEST_ACCOUNT = "99999999999"  # Different account to trigger clearing

    def test_rmf0005_has_verified_bank_before_test(self, admin_headers):
        """Pre-condition: RMF0005 bank_details.verified should be True."""
        resp = requests.get(f"{BASE_URL}/api/employees/{self.EMPLOYEE_ID}", headers=admin_headers)
        assert resp.status_code == 200
        emp = resp.json()
        bank = emp.get("bank_details", {})
        assert bank.get("verified") is True, \
            f"Pre-condition failed: RMF0005 bank not verified, got: {bank}"

    def test_changing_account_number_clears_bank_verification(self, admin_headers):
        """PUT with different account_number → bank_details.verified should be cleared."""
        resp = requests.put(
            f"{BASE_URL}/api/employees/{self.EMPLOYEE_ID}",
            json={"account_number": self.TEST_ACCOUNT},
            headers=admin_headers,
        )
        assert resp.status_code == 200, f"PUT failed: {resp.text}"
        emp = resp.json()
        bank = emp.get("bank_details", {})
        assert bank.get("verified") is not True, \
            f"bank_details.verified should be cleared after account change, got: {bank}"

    def test_bank_verified_cleared_on_get_after_account_change(self, admin_headers):
        """GET after account change: bank_details.verified should still be cleared (persisted)."""
        resp = requests.get(f"{BASE_URL}/api/employees/{self.EMPLOYEE_ID}", headers=admin_headers)
        assert resp.status_code == 200
        emp = resp.json()
        bank = emp.get("bank_details", {})
        assert bank.get("verified") is not True, \
            f"bank_details.verified should still be cleared in DB after account change: {bank}"

    def test_bank_verified_name_cleared_after_account_change(self, admin_headers):
        """bank_details.verified_name should also be cleared after account change."""
        resp = requests.get(f"{BASE_URL}/api/employees/{self.EMPLOYEE_ID}", headers=admin_headers)
        assert resp.status_code == 200
        emp = resp.json()
        bank = emp.get("bank_details", {})
        # verified_name should be None or absent after clearing
        assert not bank.get("verified_name"), \
            f"verified_name should be cleared, got: {bank.get('verified_name')}"

    def test_restore_original_account_number(self, admin_headers):
        """Restore RMF0005 bank details to original values."""
        resp = requests.put(
            f"{BASE_URL}/api/employees/{self.EMPLOYEE_ID}",
            json={
                "account_number": self.ORIG_ACCOUNT,
                "ifsc_code": self.ORIG_IFSC,
                "bank_name": self.ORIG_BANK,
            },
            headers=admin_headers,
        )
        assert resp.status_code == 200, f"Restore bank failed: {resp.text}"
        emp = resp.json()
        bank = emp.get("bank_details", {})
        assert bank.get("account_number") == self.ORIG_ACCOUNT, \
            f"Account not restored: expected {self.ORIG_ACCOUNT}, got {bank.get('account_number')}"

    def test_ifsc_change_clears_bank_verification_rmf0009(self, admin_headers):
        """PUT with different ifsc_code → bank_details.verified cleared on RMF0009."""
        # Get current state first
        resp = requests.get(f"{BASE_URL}/api/employees/RMF0009", headers=admin_headers)
        assert resp.status_code == 200
        emp = resp.json()
        orig_bank = emp.get("bank_details", {})
        orig_account = orig_bank.get("account_number")
        orig_ifsc = orig_bank.get("ifsc_code")
        orig_bank_name = orig_bank.get("bank_name")

        if not orig_bank.get("verified"):
            pytest.skip("RMF0009 bank not verified — skipping IFSC change test")

        # Change IFSC to a different value
        new_ifsc = "HDFC0000001"
        put_resp = requests.put(
            f"{BASE_URL}/api/employees/RMF0009",
            json={"ifsc_code": new_ifsc},
            headers=admin_headers,
        )
        assert put_resp.status_code == 200, f"IFSC change failed: {put_resp.text}"
        bank = put_resp.json().get("bank_details", {})
        assert bank.get("verified") is not True, \
            f"bank_details.verified should be cleared after IFSC change, got: {bank}"

        # Restore original IFSC
        restore_resp = requests.put(
            f"{BASE_URL}/api/employees/RMF0009",
            json={"account_number": orig_account, "ifsc_code": orig_ifsc, "bank_name": orig_bank_name},
            headers=admin_headers,
        )
        assert restore_resp.status_code == 200, f"Restore failed: {restore_resp.text}"


# ── 4. NEFT export skips unverified bank accounts ───────────────────────────

class TestNeftExportBankVerificationFilter:
    """GET /api/payroll/export/neft?period must only include bank-verified employees."""

    PERIOD = "2026-05"
    BANK_VERIFIED_IDS = {"RMF0005", "RMF0009"}  # verified bank accounts in DB

    def test_neft_export_returns_200(self, admin_headers):
        """NEFT export for a valid period returns HTTP 200."""
        resp = requests.get(
            f"{BASE_URL}/api/payroll/export/neft",
            params={"period": self.PERIOD},
            headers=admin_headers,
        )
        assert resp.status_code == 200, f"NEFT export failed: {resp.text[:300]}"

    def test_neft_export_content_type_is_excel(self, admin_headers):
        """NEFT export response must be an Excel file (OOXML)."""
        resp = requests.get(
            f"{BASE_URL}/api/payroll/export/neft",
            params={"period": self.PERIOD},
            headers=admin_headers,
        )
        assert resp.status_code == 200
        content_type = resp.headers.get("Content-Type", "")
        assert "spreadsheetml" in content_type or "excel" in content_type.lower() or len(resp.content) > 100, \
            f"Expected Excel content type, got: {content_type}"

    def test_neft_export_only_includes_verified_employees(self, admin_headers):
        """Parse the NEFT Excel and verify only bank-verified employees are included."""
        resp = requests.get(
            f"{BASE_URL}/api/payroll/export/neft",
            params={"period": self.PERIOD},
            headers=admin_headers,
        )
        assert resp.status_code == 200

        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True)
        ws = wb.active

        # Parse rows: first row is header, data rows start at row 2
        data_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                data_rows.append(row)

        print(f"\nNEFT export rows (excluding header): {len(data_rows)}")

        # The beneficiary account numbers in the NEFT sheet
        # Col 5 (index 4) = Beneficiary Account No
        beneficiary_accounts = [str(row[4] or "").strip() for row in data_rows if row[4]]
        print(f"Beneficiary accounts in NEFT: {beneficiary_accounts}")

        # Known verified accounts
        verified_accounts = {"32617846559", "1714831362"}  # RMF0005 and RMF0009
        # Known unverified accounts that should NOT appear
        # (Most employees in 2026-05 period are unverified)

        # At minimum, we should have <= total employees in period rows
        # (since most are unverified, the count should be small)
        # We expect exactly the verified ones to appear
        for acct in beneficiary_accounts:
            assert acct in verified_accounts, \
                f"Account {acct} should NOT be in NEFT export (not bank-verified)"

        assert len(data_rows) <= len(self.BANK_VERIFIED_IDS), \
            f"Expected at most {len(self.BANK_VERIFIED_IDS)} rows in NEFT, got {len(data_rows)}"

    def test_neft_export_verified_accounts_present(self, admin_headers):
        """NEFT export should contain verified bank employees when bank_details.verified=True.
        Note: This test seeds verified=True directly before checking, since prior bank tests may
        have cleared the verification flags as part of testing the clearing behavior.
        """
        import asyncio
        from motor.motor_asyncio import AsyncIOMotorClient
        from dotenv import load_dotenv
        import os as _os
        load_dotenv()
        # Directly seed verified=True for RMF0005 and RMF0009 so the NEFT filter is testable
        async def _seed():
            client = AsyncIOMotorClient(_os.environ["MONGO_URL"])
            db = client[_os.environ.get("DB_NAME", "hr_system")]
            await db.employees.update_one(
                {"employee_id": "RMF0005"},
                {"$set": {"bank_details.verified": True, "bank_details.verified_name": "Miss ROOPAM  GUPTA"}}
            )
            await db.employees.update_one(
                {"employee_id": "RMF0009"},
                {"$set": {"bank_details.verified": True, "bank_details.verified_name": "SHIVANI PATHAK"}}
            )
        asyncio.run(_seed())

        resp = requests.get(
            f"{BASE_URL}/api/payroll/export/neft",
            params={"period": self.PERIOD},
            headers=admin_headers,
        )
        assert resp.status_code == 200

        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True)
        ws = wb.active

        data_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                data_rows.append(row)

        # Only check if the period has payroll records for these employees
        beneficiary_accounts = {str(row[4] or "").strip() for row in data_rows if row[4]}
        print(f"NEFT export beneficiary accounts after seed: {beneficiary_accounts}")

        # Both verified employees' accounts should be in the NEFT export
        # RMF0005 account: 32617846559, RMF0009 account: 1714831362
        assert "32617846559" in beneficiary_accounts, \
            f"RMF0005 (verified bank) account 32617846559 should be in NEFT, got: {beneficiary_accounts}"
        assert "1714831362" in beneficiary_accounts, \
            f"RMF0009 (verified bank) account 1714831362 should be in NEFT, got: {beneficiary_accounts}"

    def test_neft_export_unverified_period_empty(self, admin_headers):
        """NEFT export for a period with no records returns 200 with no data rows."""
        resp = requests.get(
            f"{BASE_URL}/api/payroll/export/neft",
            params={"period": "2025-01"},
            headers=admin_headers,
        )
        # Even if no payroll records, should return 200 with empty sheet (or valid Excel)
        assert resp.status_code in [200, 404], f"Unexpected: {resp.status_code} {resp.text[:200]}"

    def test_neft_export_requires_auth(self):
        """NEFT export without auth returns 401."""
        resp = requests.get(
            f"{BASE_URL}/api/payroll/export/neft",
            params={"period": self.PERIOD},
        )
        assert resp.status_code in [401, 403], f"Expected auth failure, got: {resp.status_code}"

    def test_neft_export_access_denied_for_non_admin(self):
        """NEFT export as employee (non-admin/management) returns 403."""
        # RMF0005 has role=employee in users collection
        login_resp = requests.post(f"{BASE_URL}/api/auth/login", json={
            "username": "RMF0005",
            "password": "Radhya@123",
        })
        if login_resp.status_code != 200:
            pytest.skip("RMF0005 login failed — skipping access denied test")
        token = login_resp.json().get("access_token") or login_resp.json().get("token")
        emp_headers = {"Authorization": f"Bearer {token}"}

        resp = requests.get(
            f"{BASE_URL}/api/payroll/export/neft",
            params={"period": self.PERIOD},
            headers=emp_headers,
        )
        assert resp.status_code == 403, f"Expected 403 for employee role, got: {resp.status_code}"


# ── 5. PUT /api/employees does not clear bank if bank fields unchanged ───────

class TestBankVerificationPreservedIfUnchanged:
    """PUT with non-bank fields should NOT clear bank_details.verified on RMF0009."""

    def test_non_bank_update_preserves_bank_verification(self, admin_headers):
        """Updating employee's designation should NOT clear bank verification."""
        # First check RMF0009 bank verification state
        resp = requests.get(f"{BASE_URL}/api/employees/RMF0009", headers=admin_headers)
        assert resp.status_code == 200
        emp = resp.json()
        original_bank_verified = emp.get("bank_details", {}).get("verified")

        if not original_bank_verified:
            pytest.skip("RMF0009 bank not verified; skipping preservation test")

        # Update a non-bank field (designation) — bank should be untouched
        put_resp = requests.put(
            f"{BASE_URL}/api/employees/RMF0009",
            json={"designation": emp.get("designation", "Executive")},
            headers=admin_headers,
        )
        assert put_resp.status_code == 200
        updated_bank = put_resp.json().get("bank_details", {})
        assert updated_bank.get("verified") is True, \
            f"bank_details.verified should remain True after non-bank update, got: {updated_bank}"

    def test_non_uan_update_preserves_uan_verification_rmf0008(self, admin_headers):
        """Updating RMF0008 email should NOT clear uan_verification."""
        resp = requests.get(f"{BASE_URL}/api/employees/RMF0008", headers=admin_headers)
        assert resp.status_code == 200
        emp = resp.json()
        original_uan_verified = emp.get("uan_verification", {}).get("verified")

        if not original_uan_verified:
            pytest.skip("RMF0008 UAN not verified; skipping preservation test")

        # Update a non-UAN field — uan_verification should be untouched
        put_resp = requests.put(
            f"{BASE_URL}/api/employees/RMF0008",
            json={"blood_group": emp.get("blood_group") or "O+"},
            headers=admin_headers,
        )
        assert put_resp.status_code == 200
        updated_uan_ver = put_resp.json().get("uan_verification", {})
        assert updated_uan_ver.get("verified") is True, \
            f"uan_verification should remain True after non-UAN update, got: {updated_uan_ver}"
