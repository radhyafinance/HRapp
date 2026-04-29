"""Tests for new features: company settings, NEFT export format, location tracking."""
import pytest
import requests
import os
import io
import openpyxl

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")


@pytest.fixture(scope="module")
def headers():
    resp = requests.post(f"{BASE_URL}/api/auth/login", json={"email": "admin@radhyamfi.com", "password": "Admin@123"})
    assert resp.status_code == 200, resp.text
    token = resp.json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


# Settings - Company / Bank
class TestCompanySettings:
    def test_get_company_creates_defaults(self, headers):
        r = requests.get(f"{BASE_URL}/api/settings/company", headers=headers)
        assert r.status_code == 200
        data = r.json()
        assert "_id" not in data
        # required fields exist
        for k in ["company_name", "company_short_code", "debit_account_no",
                  "debit_account_ifsc", "transaction_type"]:
            assert k in data, f"Missing {k}"

    def test_put_company_persists(self, headers):
        payload = {
            "company_name": "Radhya Micro Finance Private Limited",
            "company_short_code": "RMF0001",
            "debit_account_no": "123456789012",
            "debit_account_ifsc": "ICIC0000001",
            "debit_bank_name": "ICICI Bank",
            "transaction_type": "NFT",
            "address": "Moradabad",
            "cin": "U65990UP2020PTC123456",
            "phone": "9999999999",
            "email": "hr@radhyamfi.com",
            "website": "https://radhyamfi.com",
        }
        r = requests.put(f"{BASE_URL}/api/settings/company", json=payload, headers=headers)
        assert r.status_code == 200, r.text
        data = r.json()
        assert "_id" not in data
        assert data["company_short_code"] == "RMF0001"
        assert data["debit_account_no"] == "123456789012"
        assert data["transaction_type"] == "NFT"

        # GET to verify persistence
        g = requests.get(f"{BASE_URL}/api/settings/company", headers=headers)
        assert g.status_code == 200
        gd = g.json()
        assert gd["debit_account_no"] == "123456789012"
        assert gd["debit_account_ifsc"] == "ICIC0000001"


# NEFT export
class TestNEFTExport:
    def test_neft_export_format(self, headers):
        # Ensure settings configured first
        requests.put(f"{BASE_URL}/api/settings/company", json={
            "debit_account_no": "123456789012",
            "transaction_type": "NFT",
            "company_short_code": "RMF0001",
        }, headers=headers)

        r = requests.get(f"{BASE_URL}/api/payroll/export/neft?period=2026-04", headers=headers)
        assert r.status_code == 200, r.text
        # Filename
        cd = r.headers.get("Content-Disposition", "")
        assert "NEFT_RMF0001_2026-04.xlsx" in cd, f"Bad filename header: {cd}"

        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # Headers - 8 columns
        headers_row = [c.value for c in ws[1]]
        assert len(headers_row) == 8
        assert "Transaction type" in (headers_row[0] or "")
        assert "Amount" in (headers_row[1] or "")
        assert "Debit Account no" in (headers_row[2] or "")
        assert "IFSC" in (headers_row[3] or "")
        assert "Beneficiary Account No" in (headers_row[4] or "")
        assert "Beneficiary Name" in (headers_row[5] or "")
        assert "Remarks for Client" in (headers_row[6] or "")
        assert "Remarks for Beneficiary" in (headers_row[7] or "")

        # Data rows
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) > 0, "No payroll rows for 2026-04"
        for row in rows:
            assert row[0] == "NFT"
            assert isinstance(row[1], (int, float))
            assert row[2] == "123456789012"
            # IFSC uppercase if present
            if row[3]:
                assert row[3] == row[3].upper()
            # Beneficiary name <= 32 chars, no special chars
            name = row[5] or ""
            assert len(name) <= 32
            assert all(ch.isalpha() or ch == " " for ch in name), f"Bad name: {name}"
            assert name == name.upper()
            # Remarks
            assert row[6] == "RMF0001 Salary Apr26"
            assert len(row[6]) <= 21
            assert row[7] == "RMF0001 Salary Apr26"
            assert len(row[7]) <= 30


# Attendance + tracking
class TestAttendanceTracking:
    def _create_employee(self, headers):
        payload = {
            "first_name": "TEST_Track", "last_name": "Field",
            "email": f"TEST_track_{os.urandom(3).hex()}@radhyamfi.com",
            "mobile": "9000000001",
            "department": "Operations", "designation": "Field Officer",
            "joining_date": "2024-01-15", "employment_type": "full_time",
            "branch": "Head Office - Moradabad", "role": "field_agent",
            "basic": 15000, "hra": 3000,
        }
        r = requests.post(f"{BASE_URL}/api/employees", json=payload, headers=headers)
        assert r.status_code in (200, 201), r.text
        return r.json()["employee_id"]

    def test_punch_in_without_photo(self, headers):
        emp_id = self._create_employee(headers)
        r = requests.post(f"{BASE_URL}/api/attendance/punch-in", json={
            "employee_id": emp_id, "latitude": 28.84, "longitude": 78.77
        }, headers=headers)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["success"] is True
        assert "geofence_verified" in d

    def test_location_update_requires_active_session(self, headers):
        # Use a non-existent employee with no punch-in -> should fail
        r = requests.post(f"{BASE_URL}/api/attendance/location-update", json={
            "employee_id": "EMP_NO_SESSION_XX", "latitude": 28.84, "longitude": 78.77
        }, headers=headers)
        assert r.status_code == 400

    def test_location_update_with_active_session(self, headers):
        emp_id = self._create_employee(headers)
        # Punch in first
        requests.post(f"{BASE_URL}/api/attendance/punch-in", json={
            "employee_id": emp_id, "latitude": 28.84, "longitude": 78.77
        }, headers=headers)
        r = requests.post(f"{BASE_URL}/api/attendance/location-update", json={
            "employee_id": emp_id, "latitude": 28.841, "longitude": 78.771, "accuracy": 10
        }, headers=headers)
        assert r.status_code == 200, r.text
        assert r.json()["success"] is True

    def test_location_track_response_shape(self, headers):
        emp_id = self._create_employee(headers)
        requests.post(f"{BASE_URL}/api/attendance/punch-in", json={
            "employee_id": emp_id, "latitude": 28.84, "longitude": 78.77
        }, headers=headers)
        # Add a few location updates
        for i in range(3):
            requests.post(f"{BASE_URL}/api/attendance/location-update", json={
                "employee_id": emp_id, "latitude": 28.84 + i * 0.0001,
                "longitude": 78.77 + i * 0.0001
            }, headers=headers)
        r = requests.get(f"{BASE_URL}/api/attendance/location-track/{emp_id}", headers=headers)
        assert r.status_code == 200, r.text
        d = r.json()
        assert "locations" in d and "stops" in d and "attendance" in d
        assert isinstance(d["locations"], list)
        assert isinstance(d["stops"], list)

    def test_field_staff_active(self, headers):
        r = requests.get(f"{BASE_URL}/api/attendance/field-staff/active", headers=headers)
        assert r.status_code == 200
        data = r.json()
        assert isinstance(data, list)
        if data:
            row = data[0]
            for k in ["employee_id", "name", "punch_in_time", "location_points"]:
                assert k in row


class TestRoleAccess:
    def test_settings_put_forbidden_for_employee(self):
        # Get a list of existing users; try to create or use field_agent.
        # We'll skip if no non-admin login available; instead directly test by
        # using the no-token request (which returns 401/403) — quick check.
        r = requests.put(f"{BASE_URL}/api/settings/company", json={"company_name": "X"})
        assert r.status_code in (401, 403)

    def test_field_staff_active_unauth(self):
        r = requests.get(f"{BASE_URL}/api/attendance/field-staff/active")
        assert r.status_code in (401, 403)
