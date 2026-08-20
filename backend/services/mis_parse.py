"""
Turn the MIS Demand & Collection workbook into the numbers the daily close needs.

WHAT COMES OUT
--------------
Per branch, per day: how much was collected, how much of it is CASH sitting in
someone's hand, and how much is already in a bank. The cash figure is what the
branch has to deposit, and it is the only number the deposit slips have to match.

    Mobile App  = an officer posted it on the app -> physical cash
    MIS         = posted at Head Office          -> physical cash
    UPI         = paid digitally                 -> already banked, never deposited
    (blank)     = demand that was not collected

Getting this wrong in the obvious direction -- treating "total collected" as the
deposit target -- overstates every branch by about a quarter. Across July, UPI
was 24.1% of Rs 84.9 lakh.

TWO RULES THIS FILE EXISTS TO ENFORCE
-------------------------------------
1. AN UNRECOGNISED PAYMENT MODE IS NEVER GUESSED. If the MIS starts emitting a
   mode we have not seen -- "Bank Transfer", "Cash", anything -- it goes to
   `unknown` and raises a warning, and the day does not silently reconcile. The
   tempting default (treat it as cash) would tell a branch to deposit money that
   was never in its hands; the other default (treat it as digital) would hide a
   real shortfall. Neither is safe, so we refuse to choose.

2. NO CUSTOMER DATA IS READ. The workbook carries member names, husbands' names,
   phone numbers, member IDs and loan numbers. A cash reconciliation needs none
   of it, so only nine columns are ever touched (see _NEEDED). Putting 3,000
   borrowers' phone numbers into a closing tool would be an easy accident to
   make by loading the sheet wholesale, and a bad one to explain afterwards.

   `status` and `Posted_by_user` were added later and do not breach this: one is
   the state of a loan, the other is which of OUR staff keyed the entry. The
   loan number would have made the grouping below exact, and is deliberately
   still not read -- see the note on `special`.

3. A LOAN'S STATUS IS NOT A TRANSACTION'S NATURE. `status` describes the loan
   when the report was GENERATED, not what happened at the counter. On 30-Jul a
   field officer collected an ordinary Rs 2,620 instalment at 17:20; Head Office
   had preclosed that loan at 09:54, so the report labels the afternoon's real
   cash `preclosed` too. Excluding money by status alone would have taken Rs
   2,620 of a branch manager's cash out of their deposit target -- and money
   quietly removed from a target is money nobody ever looks for.

   So this file only REPORTS status and who posted it. What that means for the
   deposit figure is decided in routes/closing.py, where it can be checked
   against who actually works at Head Office.

Columns are located BY HEADER NAME, not by position. The ad-hoc analysis this
was built from used fixed indices, which works right up until the vendor inserts
a column and every branch's cash figure silently becomes its penalty collection.
"""

import io
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Optional

import openpyxl

logger = logging.getLogger(__name__)

SHEET_MEMBER = "Member wise report"

# Physical cash, which somebody has to walk into a bank.
CASH_MODES = {"mobile app", "mis"}
# Already in an account by the time we see it.
DIGITAL_MODES = {"upi"}

_NEEDED = (
    "Branch",
    "Transaction_Date",
    "Transaction_Time",
    "Total_Demand",
    "TotalDemandAsPerCDS",
    "Total_Collect",
    "Posted_by",
    "status",
    "Posted_by_user",
)

# Loan states where the money may not be the branch's to bank.
#
#   deathclose -- the borrower died and the insurer settles the outstanding
#                 principal. It lands in the report looking exactly like a
#                 collection; no branch manager ever touched it.
#   preclosed  -- either the borrower really did pay off the loan (cash, to the
#                 BM) or the balance was netted off against a new loan, in which
#                 case nothing moved at all. The workbook cannot tell these
#                 apart; a person at Head Office can.
#
# Rs 5,55,190 of July's Rs 64,43,918 of "cash" was one of these -- 8.6%, across
# 20 of 69 branch-days, and on 20-Jul it was CHANDPUR's entire stated deposit.
UNSETTLED_STATUSES = {"preclosed", "deathclose"}

# "002 - CHANDPUR" -> ("002", "CHANDPUR")
_BRANCH_RE = re.compile(r"^\s*(\d+)\s*-\s*(.+?)\s*$")


class MISParseError(Exception):
    """The workbook was not the shape we rely on."""


@dataclass
class BranchDay:
    branch_code: str
    branch_name: str
    day: date
    cash: float = 0.0            # Mobile App + MIS -> must be deposited
    upi: float = 0.0             # already banked
    unknown: float = 0.0         # an unrecognised mode; see rule 1
    demand: float = 0.0
    demand_cds: float = 0.0
    txn_count: int = 0
    uncollected_count: int = 0
    last_posted_at: Optional[str] = None
    unknown_modes: set = field(default_factory=set)
    # Cash-mode collections on a preclosed or deathclose loan, GROUPED.
    #
    # Grouped rather than listed one by one because the only fields we are
    # willing to read -- time, amount, and which member of staff posted it --
    # collide: on 18-Jul two preclosures of Rs 5,343 were posted in the same
    # second by the same user. The loan number would separate them and is not
    # read (rule 2).
    #
    # It does not matter. Two identical rows are INTERCHANGEABLE here: if Head
    # Office says one of the two Rs 5,343 preclosures was netted off, which one
    # is a question with no consequence — the arithmetic is the same either way.
    # So the group carries a count, and Head Office answers "how many", which for
    # 33 of July's 40 groups is simply "the one".
    special: dict = field(default_factory=dict)

    @property
    def collected(self) -> float:
        return round(self.cash + self.upi + self.unknown, 2)

    @property
    def expected_deposit(self) -> float:
        """What the deposit slips for this branch and day must add up to.

        Unknown-mode money is deliberately excluded: we do not know whether it is
        in a drawer or in a bank, and inventing an answer here is how a
        reconciliation tool starts lying.
        """
        return round(self.cash, 2)

    def as_dict(self) -> dict:
        return {
            "branch_code": self.branch_code,
            "branch_name": self.branch_name,
            "date": self.day.isoformat(),
            "cash": round(self.cash, 2),
            "upi": round(self.upi, 2),
            "unknown": round(self.unknown, 2),
            "collected": self.collected,
            "expected_deposit": self.expected_deposit,
            "demand": round(self.demand, 2),
            "demand_cds": round(self.demand_cds, 2),
            "txn_count": self.txn_count,
            "uncollected_count": self.uncollected_count,
            "last_posted_at": self.last_posted_at,
            "unknown_modes": sorted(self.unknown_modes),
            # Sorted so the content digest that decides "did anything change?"
            # does not flip on dict ordering alone and re-alert every pull.
            "special": [self.special[k] for k in sorted(self.special)],
        }


@dataclass
class DayReport:
    days: list                       # the dates actually present in the file
    branches: list                   # list[BranchDay]
    warnings: list = field(default_factory=list)

    @property
    def total_cash(self) -> float:
        return round(sum(b.cash for b in self.branches), 2)

    @property
    def total_upi(self) -> float:
        return round(sum(b.upi for b in self.branches), 2)

    @property
    def total_collected(self) -> float:
        return round(sum(b.collected for b in self.branches), 2)

    def as_dict(self) -> dict:
        return {
            "days": [d.isoformat() for d in self.days],
            "branches": [b.as_dict() for b in self.branches],
            "totals": {
                "cash": self.total_cash,
                "upi": self.total_upi,
                "collected": self.total_collected,
                "expected_deposit": self.total_cash,
            },
            "warnings": self.warnings,
        }


def _num(v, bad: list = None) -> float:
    """A number, or 0.0 — but never silently.

    The original version swallowed every unparseable cell as zero. That is the
    most dangerous default in this file: one `Total_Collect` exported as the
    string "2,680.00" removed Rs 2,680 from a branch's deposit target with no
    warning anywhere, which is money a branch is then told not to bank. The
    module already refuses to guess an unrecognised payment MODE; an
    unparseable AMOUNT deserves exactly the same treatment, and did not get it.

    `bad` collects the offending values so the caller can raise a warning.
    """
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        if bad is not None and len(bad) < 5:   # a few examples are enough
            bad.append(str(v)[:40])
        elif bad is not None:
            bad.append(None)                   # keep counting, stop collecting
        return 0.0


def _norm_time(ts: str) -> str:
    """'9:27:29' -> '09:27:29'. Left alone if it is not a clock string."""
    parts = ts.split(":")
    if len(parts) in (2, 3) and all(p.strip().isdigit() for p in parts):
        return ":".join(p.strip().rjust(2, "0") for p in parts)
    return ts


def _as_date(v) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).date()
    except ValueError:
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d").date()
        except ValueError:
            return None


def _split_branch(raw) -> tuple[str, str]:
    s = ("" if raw is None else str(raw)).strip()
    m = _BRANCH_RE.match(s)
    if m:
        return m.group(1), m.group(2)
    # Unrecognised shape: keep the whole string as the name rather than dropping
    # the row. A branch we cannot code is still a branch holding real cash.
    return "", s


def parse_demand_collection(data: bytes, expect_day: Optional[date] = None) -> DayReport:
    """Parse the workbook. `data` is the raw .xlsx bytes.

    `expect_day` is the day we asked the MIS for. It is checked rather than
    trusted: a stale session or a mis-set filter returns a perfectly valid
    workbook for the WRONG date, which is the one failure that would quietly
    reconcile a branch against somebody else's money.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise MISParseError(f"could not open the workbook: {e}") from e

    try:
        if SHEET_MEMBER not in wb.sheetnames:
            raise MISParseError(
                f"sheet {SHEET_MEMBER!r} missing (found: {', '.join(wb.sheetnames)})"
            )
        ws = wb[SHEET_MEMBER]
        rows = ws.iter_rows(values_only=True)

        try:
            header = next(rows)
        except StopIteration:
            raise MISParseError("the member sheet is empty")

        index = {}
        for i, name in enumerate(header):
            if name is None:
                continue
            index[str(name).strip()] = i
        missing = [c for c in _NEEDED if c not in index]
        if missing:
            raise MISParseError(f"missing expected column(s): {', '.join(missing)}")

        # Only these are ever pulled out of a row. See rule 2.
        i_branch = index["Branch"]
        i_date = index["Transaction_Date"]
        i_time = index["Transaction_Time"]
        i_demand = index["Total_Demand"]
        i_cds = index["TotalDemandAsPerCDS"]
        i_collect = index["Total_Collect"]
        i_mode = index["Posted_by"]
        i_status = index["status"]
        i_user = index["Posted_by_user"]

        by_branch: dict[str, BranchDay] = {}
        days: set = set()
        warnings: list = []
        unknown_modes: set = set()
        undated = 0
        undated_value = 0.0
        bad_numbers: list = []
        # Branch codes seen against more than one spelling of the name. Two
        # spellings of one branch ("002 - CHANDPUR" and "002 - Chandpur") become
        # two BranchDay objects sharing a code, and anything downstream keyed on
        # the code alone will keep one and lose the other.
        code_to_names: dict = {}
        # NO ROW-LEVEL DUPLICATE DETECTION, deliberately.
        #
        # The MIS does emit genuine duplicates -- two byte-identical transactions
        # exist in the July file -- and a repeated CASH row would overstate what a
        # branch is told to bank. But the only fields available without reading
        # customer identifiers are branch, date, time, amount and mode, and those
        # collide constantly: Transaction_Time is a batch stamp (a whole centre
        # posts at 09:27:29) and instalment amounts repeat across members. A first
        # attempt flagged 62 "duplicates" on a clean day, which is how warnings
        # get trained into noise.
        #
        # TXN_ID would work, but it is populated only on UPI rows (852/852 in
        # July, 0 on cash), so it cannot deduplicate the figure that matters.
        # Detecting this properly means ingesting a member or loan identifier,
        # which Rule 2 exists to prevent. The trade is made knowingly: no false
        # alarms, and duplicate cash rows are not caught here.

        # THE GUARD DID NOT MOVE, and that is deliberate.
        #
        # status (28) and Posted_by_user (30) sit to the right of Posted_by (29),
        # so widening this to the new columns would skip any row too short to
        # reach them — and a skipped row is CASH REMOVED from a branch's deposit
        # target with nothing on screen to say so. Every row in the July file is
        # full width, but that is one file's evidence for a rule that decides how
        # much money someone carries to a bank.
        #
        # So the guard still covers only the columns the figures need, and the
        # two new ones are read defensively below. A row too short to say who
        # posted it still contributes its cash; it just cannot be classified, and
        # `_cell` returning None routes it to "ask Head Office" rather than to
        # silence.
        def _cell(r, i):
            return r[i] if i < len(r) else None

        for row in rows:
            if row is None or len(row) <= i_mode:
                continue
            code, name = _split_branch(row[i_branch])
            if not name:
                continue

            d = _as_date(row[i_date])
            if d is None:
                # Count the MONEY, not just the rows. "3 rows skipped" reads as
                # housekeeping; "Rs 9,000 skipped" is the deposit target being
                # wrong, and it is the same sentence either way to whoever has to
                # find the difference at 9pm.
                undated += 1
                undated_value += _num(row[i_collect])
                continue
            days.add(d)

            key = f"{code}|{name}"
            bd = by_branch.get(key)
            if bd is None:
                bd = by_branch[key] = BranchDay(branch_code=code, branch_name=name, day=d)
            if code:
                code_to_names.setdefault(code, set()).add(name)

            bd.demand += _num(row[i_demand], bad_numbers)
            bd.demand_cds += _num(row[i_cds], bad_numbers)

            amount = _num(row[i_collect], bad_numbers)
            mode = ("" if row[i_mode] is None else str(row[i_mode])).strip()


            if not mode and amount == 0:
                # Demand raised, nothing collected. Normal, and not a cash event.
                bd.uncollected_count += 1
                continue

            bd.txn_count += 1
            norm = mode.lower()
            if norm in CASH_MODES:
                bd.cash += amount
            elif norm in DIGITAL_MODES:
                bd.upi += amount
            else:
                bd.unknown += amount
                bd.unknown_modes.add(mode or "(blank)")
                unknown_modes.add(mode or "(blank)")

            t = row[i_time]
            ts = ""
            if t is not None:
                # Zero-pad before comparing. These are clock strings compared as
                # strings, which is fine only while every one is HH:MM:SS -- the
                # day the MIS emits "9:27:29", it sorts ABOVE "21:14:03" and the
                # "last posted at" stamp jumps backwards to the morning.
                ts = _norm_time(str(t).strip())
                if ts and (bd.last_posted_at is None or ts > bd.last_posted_at):
                    bd.last_posted_at = ts

            # Only CASH-mode rows are recorded. A preclosure paid by UPI is
            # already in a bank and was never going to be deposited, so it
            # cannot inflate anybody's deposit target.
            _st = _cell(row, i_status)
            status = ("" if _st is None else str(_st)).strip()
            if norm in CASH_MODES and amount > 0 and status.lower() in UNSETTLED_STATUSES:
                _pu = _cell(row, i_user)
                poster = ("" if _pu is None else str(_pu)).strip()
                k = f"{ts}|{amount:.2f}|{status.lower()}|{poster}"
                g = bd.special.get(k)
                if g is None:
                    g = bd.special[k] = {
                        "key": k, "time": ts, "amount": round(amount, 2),
                        "status": status.lower(), "posted_by_user": poster, "count": 0,
                    }
                g["count"] += 1
    finally:
        wb.close()

    if not by_branch:
        warnings.append("the report contained no collection rows at all")

    if undated:
        warnings.append(
            f"{undated} row(s) had no usable transaction date and were skipped"
            + (f" — Rs {undated_value:,.2f} of collection is NOT in these totals"
               if undated_value else ""))

    if bad_numbers:
        shown = [x for x in bad_numbers if x][:5]
        warnings.append(
            f"{len(bad_numbers)} amount cell(s) could not be read as a number and were "
            f"counted as zero"
            + (f" (e.g. {', '.join(repr(x) for x in shown)})" if shown else "")
            + " -- the deposit target is understated by whatever they held"
        )

    dupe_codes = {c: sorted(n) for c, n in code_to_names.items() if len(n) > 1}
    if dupe_codes:
        warnings.append(
            "one branch code appears under more than one name: "
            + "; ".join(f"{c} -> {', '.join(n)}" for c, n in dupe_codes.items())
        )

    if unknown_modes:
        # Loud on purpose. Someone has to decide whether this is cash.
        warnings.append(
            "unrecognised payment mode(s): "
            + ", ".join(sorted(unknown_modes))
            + " -- this money is excluded from the deposit target until classified"
        )

    ordered_days = sorted(days)
    if expect_day is not None:
        if not days:
            warnings.append(f"expected data for {expect_day.isoformat()} but the report was empty")
        else:
            wrong = [d for d in ordered_days if d != expect_day]
            if wrong:
                warnings.append(
                    f"expected only {expect_day.isoformat()} but the report also covers "
                    + ", ".join(d.isoformat() for d in wrong)
                )

    return DayReport(
        days=ordered_days,
        branches=sorted(by_branch.values(), key=lambda b: (b.branch_code, b.branch_name)),
        warnings=warnings,
    )


def storage_conflicts(report: DayReport, day: date) -> list:
    """Reasons this report must NOT be written as one day's per-branch rows.

    Storage keys a document per (date, branch_code). Anything that lets two
    BranchDay objects collide on that key means one branch's cash silently
    replaces another's — and the wrong figure is the one that survives, because
    the correct total only ever existed in a return value nobody displays.

    Three ways it happens, all reproduced against real data:

    * TWO SPELLINGS OF ONE BRANCH ("002 - CHANDPUR" / "002 - Chandpur") share a
      code, so the second overwrites the first. Chandpur was told to bank
      Rs 59,850 less than it collected, with no warning anywhere.
    * A `Branch` CELL WITHOUT THE "00n - " PREFIX yields an empty code, so every
      such branch collapses into a single document keyed on "". Four branches
      became three and Rs 154,702 disappeared.
    * A MULTI-DAY REPORT files each branch under whichever date its first row
      happened to carry, so the day asked for stores nothing and reads as a
      holiday, while the next poll collides on the unique index forever.

    Refusing to store is the right answer for all three: a loud failure that
    keeps yesterday's correct figures beats a quiet one that publishes wrong
    ones. `check_expected_branches` cannot catch any of these — it inspects the
    parsed report, which still looks complete; the loss happens downstream.
    """
    problems = []

    if report.days and (len(report.days) > 1 or report.days[0] != day):
        problems.append(
            "the report does not cover exactly " + day.isoformat()
            + " (it covers " + ", ".join(d.isoformat() for d in report.days) + ")"
        )

    seen = {}
    for b in report.branches:
        seen.setdefault(b.branch_code, []).append(b.branch_name)
    for code, names in seen.items():
        if len(names) > 1:
            label = f"branch code {code!r}" if code else "branches with no branch code"
            problems.append(
                f"{label} covers more than one branch name ({', '.join(sorted(names))}) "
                "— storing them would overwrite one with the other"
            )
    return problems


def check_expected_branches(report: DayReport, expected: set) -> list:
    """Warn if a branch we bank for is absent from the report.

    The fetcher requests branch checkboxes by index without being able to see
    them, so a branch that stops being requested produces a valid report that is
    simply missing a branch. Nothing else in the pipeline would notice: the
    totals would still add up, the slips for that branch would just never match
    anything. This is the guard for that.

    A branch legitimately has no row on a day it did not collect, so this cannot
    be an error -- but it must be visible.
    """
    seen = {b.branch_name.strip().upper() for b in report.branches}
    absent = sorted(x for x in expected if x.strip().upper() not in seen)
    if not absent:
        return []
    return [f"no collections reported for: {', '.join(absent)} (verify this is a genuine zero day)"]
